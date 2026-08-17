import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from PIL import Image

from workbook_exporter import export_workbook_payload


class WorkbookExporterTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.output = self.root / "exports" / "result.xlsx"
        self.images = self.root / "images"
        self.images.mkdir()
        self.png_paths = []
        for index in range(1, 7):
            image_path = self.images / f"preview-{index}.png"
            Image.new("RGB", (160, 120), (index * 20, 80, 120)).save(image_path)
            self.png_paths.append(image_path)
        self.webp_path = self.images / "preview.webp"
        Image.new("RGB", (160, 120), (180, 90, 40)).save(self.webp_path, format="WEBP")

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def _payload(self) -> dict:
        return {
            "output_path": str(self.output),
            "overview": [
                ["字段", "值"],
                ["来源商品ID", "12345678901234567890"],
                ["来源商品标题", "测试商品"],
            ],
            "main": [
                {
                    "source_path": str(self.png_paths[index * 2]),
                    "source_public_url": "https://cdn.example/source.png" if index == 0 else "",
                    "output_path": str(self.png_paths[index * 2 + 1]),
                    "output_public_url": "https://cdn.example/output.png" if index == 0 else "",
                    "generation_status": "生成成功",
                }
                for index in range(3)
            ],
            "detail": [
                {
                    "source_path": str(self.webp_path),
                    "source_public_url": "",
                    "output_path": "",
                    "output_public_url": "",
                    "generation_status": "未生成",
                }
            ],
            "sku": [
                {
                    "product_id": "12345678901234567890",
                    "sku_label": "大号白色",
                    "spec_text": "大号",
                    "color_text": "白色",
                    "price": "￥1,299.50",
                    "parse_status": "parsed",
                    "source_path": str(self.png_paths[0]),
                    "source_public_url": "https://cdn.example/sku-source.png",
                    "output_path": str(self.png_paths[1]),
                    "output_public_url": "",
                    "generation_status": "生成成功",
                }
            ],
            "parameters": [
                {
                    "type": "商品参数",
                    "name": "产品说明",
                    "value": "很长的商品参数" * 80,
                    "handling": "采集原值",
                }
            ],
            "title": {"long_title": "很长的商品标题" * 40, "short_title": "短标题"},
            "videos": [],
        }

    def test_exports_exact_sheet_order_headers_and_values(self) -> None:
        result = export_workbook_payload(self.output, self._payload())

        self.assertEqual(result, self.output)
        workbook = load_workbook(self.output)
        try:
            self.assertEqual(
                workbook.sheetnames,
                ["总览", "主图", "详情图", "SKU", "商品参数", "标题", "视频"],
            )
            expected_headers = {
                "总览": ["字段", "值"],
                "主图": ["序号", "采集图缩略图", "采集图路径", "生成图缩略图", "生成图路径", "生成状态"],
                "详情图": ["序号", "采集图缩略图", "采集图路径", "生成图缩略图", "生成图路径", "生成状态"],
                "SKU": [
                    "序号", "商品ID", "SKU标签", "规格", "颜色", "价格",
                    "采集图缩略图", "采集图路径", "生成图缩略图", "生成图路径", "生成图状态",
                ],
                "商品参数": ["类型", "参数名", "参数值", "处理方式"],
                "标题": ["序号", "长标题", "短标题"],
                "视频": ["序号", "视频名称", "公网播放地址", "访问说明"],
            }
            for sheet_name, headers in expected_headers.items():
                self.assertEqual([cell.value for cell in workbook[sheet_name][1]], headers)
                self.assertEqual(workbook[sheet_name].freeze_panes, "A2")

            sku = workbook["SKU"]
            self.assertNotIn("解析状态", [cell.value for cell in sku[1]])
            self.assertEqual(sku["B2"].value, "12345678901234567890")
            self.assertEqual(sku["B2"].number_format, "@")
            self.assertEqual(sku["F2"].value, 1299.5)
            self.assertEqual(sku["H2"].value, "https://cdn.example/sku-source.png")
            self.assertEqual(sku["J2"].value, "../images/preview-2.png")
            self.assertEqual(len(workbook["主图"]._images), 6)
            self.assertEqual(len(workbook["详情图"]._images), 1)
            self.assertEqual(len(sku._images), 2)
            self.assertGreater(workbook["标题"].row_dimensions[2].height, 22)
            self.assertGreater(workbook["商品参数"].row_dimensions[2].height, 22)
        finally:
            workbook.close()
        with zipfile.ZipFile(self.output) as archive:
            media = [name for name in archive.namelist() if name.startswith("xl/media/")]
            self.assertTrue(media)
            self.assertTrue(all(name.endswith(".png") for name in media))

    def test_prefers_public_urls_and_uses_relative_local_fallbacks(self) -> None:
        export_workbook_payload(self.output, self._payload())

        workbook = load_workbook(self.output, read_only=False)
        try:
            main = workbook["主图"]
            self.assertEqual(main["C2"].value, "https://cdn.example/source.png")
            self.assertEqual(main["E2"].value, "https://cdn.example/output.png")
            self.assertEqual(main["C3"].value, "../images/preview-3.png")
            self.assertEqual(main["E3"].value, "../images/preview-4.png")
        finally:
            workbook.close()

    def test_replaces_output_atomically_and_cleans_temporary_file_on_failure(self) -> None:
        self.output.parent.mkdir(parents=True)
        existing = Workbook()
        existing.active["A1"] = "existing"
        existing.save(self.output)
        existing.close()

        with patch("workbook_exporter.Workbook.save", side_effect=OSError("save failed")):
            with self.assertRaisesRegex(OSError, "save failed"):
                export_workbook_payload(self.output, self._payload())

        workbook = load_workbook(self.output)
        try:
            self.assertEqual(workbook.active["A1"].value, "existing")
        finally:
            workbook.close()
        self.assertEqual(list(self.output.parent.glob(f".{self.output.stem}.*{self.output.suffix}")), [])

        export_workbook_payload(self.output, self._payload())
        replacement = load_workbook(self.output)
        try:
            self.assertEqual(replacement.sheetnames, ["总览", "主图", "详情图", "SKU", "商品参数", "标题", "视频"])
        finally:
            replacement.close()
        self.assertEqual(list(self.output.parent.glob(f".{self.output.stem}.*{self.output.suffix}")), [])


if __name__ == "__main__":
    unittest.main()
