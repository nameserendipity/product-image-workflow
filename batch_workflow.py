from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
import threading
import uuid
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from hashlib import sha256
from dataclasses import dataclass, replace
from datetime import datetime
from pathlib import Path
from typing import Any, Callable
from urllib.parse import parse_qs, urlparse
from urllib.request import Request, url2pathname, urlopen

from openpyxl import load_workbook

from agent_flow import DEFAULT_MAIN_IMAGES
from image_workflows import (
    ApiSettings,
    DetailViewPlan,
    ImageTask,
    IdentitySource,
    ProductTitleClient,
    VisionClient,
    WorkflowRunner,
    build_detail_tasks,
    load_identity_sources,
    load_manifest_tasks,
    resolve_identity_image,
    materialize_sku_screenshot_references,
)
from kuaishou_parameters import ensure_kuaishou_product_parameters
from oss_uploader import OssUploader, upload_generation_records, upload_video_if_needed
from platform_urls import (
    TAOBAO_SHORT_HOSTS,
    is_douyin_product_host,
    is_kuaishou_product_host,
    is_taobao_host,
    is_tmall_host,
    kuaishou_product_id,
)
from product_identity import ProductIdentity, ProductIdentityError, ProductIdentityResolver
from shared_library_cache import SharedLibraryCache
from shared_library_client import (
    LockLease,
    SharedLibraryClient,
    SharedLibraryLockBusy,
    SharedLibraryUnavailable,
)
from shared_package_builder import SharedPackageBuilder, materialize_reused_package
from spreadsheet_inputs import extract_embedded_images, normalize_header
from workbook_exporter import export_workbook_payload


class BatchCollectionPaused(RuntimeError):
    """The collector needs user action, so the batch must not continue."""


BatchCallback = Callable[[dict[str, Any]], None]
ASSET_TYPES = ("main", "sku", "detail")
SHEET_TITLES = ("总览", "主图", "详情图", "SKU", "商品参数", "标题", "视频")
INVALID_FILENAME = re.compile(r'[<>:"/\\|?*\x00-\x1f]+')
BATCH_RESULTS_FILENAME = "batch-results.json"
SUPPLEMENT_LIMITS = {"main": 999, "sku": 8, "detail": 15}


@dataclass(frozen=True, slots=True)
class BatchItem:
    sequence: int
    row_number: int
    title: str
    product_image: Path
    source_url: str
    purchase_price: Any
    shipping_fee: Any


@dataclass(frozen=True, slots=True)
class DirectLinkBatchItem:
    sequence: int
    row_number: int
    source_url: str
    platform: str
    title: str = ""
    validation_error: str = ""
    purchase_price: Any = ""
    shipping_fee: Any = ""
    sku_screenshot: Path | None = None
    manual_skus: tuple[dict[str, Any], ...] = ()
    sku_screenshot_error: str = ""


@dataclass(frozen=True, slots=True)
class DirectReplaceBatchItem:
    sequence: int
    sheet_name: str
    row_number: int
    product_image: Path | None
    source_url: str
    platform: str
    title: str = ""
    validation_error: str = ""
    manual_skus: tuple[dict[str, Any], ...] = ()
    purchase_price: Any = ""
    shipping_fee: Any = ""


@dataclass(frozen=True, slots=True)
class SupplementWorkbookContext:
    workbook_path: Path
    item: BatchItem | DirectLinkBatchItem | DirectReplaceBatchItem
    item_root: Path
    source_manifest: Path
    generated_root: Path
    generation_mode: str


def _cell_link(cell: Any) -> str:
    hyperlink = getattr(cell, "hyperlink", None)
    return str(hyperlink.target) if hyperlink and hyperlink.target else str(cell.value or "")


def _direct_link_platform(value: str) -> tuple[str, str]:
    parsed = urlparse(value.strip())
    host = (parsed.hostname or "").lower()
    if parsed.scheme not in {"http", "https"} or not host:
        return "invalid", "不是有效的 HTTP 商品链接"
    if host in TAOBAO_SHORT_HOSTS:
        return "taobao", ""
    if host in {"u.jd.com", "3.cn"}:
        return "jd", ""
    if host == "v.douyin.com":
        return "douyin", ""
    if is_taobao_host(host):
        product_id = parse_qs(parsed.query).get("id", [""])[0]
        return ("taobao", "") if product_id else ("invalid", "淘宝商品链接缺少商品 ID")
    if is_tmall_host(host):
        product_id = parse_qs(parsed.query).get("id", [""])[0]
        return ("tmall", "") if product_id else ("invalid", "天猫商品链接缺少商品 ID")
    if host == "jd.com" or host.endswith(".jd.com"):
        product_id = re.search(r"/(\d+)\.html", parsed.path)
        return ("jd", "") if product_id else ("invalid", "京东商品链接缺少商品 ID")
    if is_douyin_product_host(host):
        product_id = parse_qs(parsed.query).get("id", [""])[0]
        return ("douyin", "") if product_id else ("invalid", "抖音商品链接缺少商品 ID")
    if is_kuaishou_product_host(host):
        return ("kuaishou", "") if kuaishou_product_id(value) else ("invalid", "快手商品链接路径或商品 ID 无效")
    return "unsupported", "不支持的商品链接平台"


def resolve_direct_item_url(value: str, timeout: float = 15.0) -> str:
    platform, validation_error = _direct_link_platform(value)
    if validation_error:
        raise ValueError(validation_error)
    host = (urlparse(value.strip()).hostname or "").lower()
    short_hosts = TAOBAO_SHORT_HOSTS | {"u.jd.com", "3.cn", "v.douyin.com"}
    if host not in short_hosts:
        return value.strip()
    request = Request(
        value.strip(),
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0 Safari/537.36"
            ),
            "Range": "bytes=0-0",
        },
    )
    try:
        with urlopen(request, timeout=timeout) as response:
            resolved = str(response.geturl() or "").strip()
    except Exception as error:
        raise ValueError(f"短链接解析失败：{error}") from error
    _, resolved_error = _direct_link_platform(resolved)
    if resolved_error or (urlparse(resolved).hostname or "").lower() in short_hosts:
        raise ValueError("短链接未解析到受支持的商品详情页")
    return resolved


def _local_image_reference(value: Any, workbook_path: Path) -> Path | None:
    text = str(value or "").strip()
    if text.lower().startswith("file://"):
        candidate = Path(url2pathname(urlparse(text).path))
    else:
        candidate = Path(text).expanduser()
        if not candidate.is_absolute():
            candidate = workbook_path.parent / candidate
    if candidate.is_file() and candidate.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}:
        return candidate.resolve()
    return None


def _stage_sku_screenshot(
    screenshot: Path | None,
    output_dir: Path | None,
    sequence: int,
    row_number: int,
) -> Path | None:
    if screenshot is None:
        return None
    if output_dir is None:
        return screenshot
    target_dir = output_dir / "source-images" / "sku-screenshots"
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / f"{sequence:03d}-row-{row_number:04d}{screenshot.suffix.lower()}"
    shutil.copy2(screenshot, target)
    return target.resolve()


def extract_direct_link_items(
    workbook_path: Path,
    output_dir: Path | None = None,
) -> list[DirectLinkBatchItem]:
    workbook = load_workbook(workbook_path, data_only=True)
    try:
        sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
        embedded_images = (
            extract_embedded_images(workbook_path, output_dir / "source-images" / "embedded")
            if output_dir is not None
            else {}
        )
        link_column = 0
        title_column = 0
        header_row = 0
        accepted_headers = {
            "商品链接",
            "对标链接",
            "竞品链接",
            "淘宝链接",
            "天猫链接",
            "京东链接",
            "抖音链接",
            "快手链接",
            "url",
        }
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 10)):
            for cell in row:
                label = re.sub(r"\s+", "", str(cell.value or "")).lower()
                if label in accepted_headers:
                    link_column = cell.column
                    header_row = cell.row
                if label in {"标题", "商品标题", "商品名称", "名称"}:
                    title_column = cell.column
            if link_column:
                break

        if not link_column:
            link_column = 1
            first_value = _cell_link(sheet.cell(1, link_column)).strip()
            if not first_value or _direct_link_platform(first_value)[0] == "invalid" and "://" not in first_value:
                return []

        items: list[DirectLinkBatchItem] = []
        start_row = header_row + 1 if header_row else 1
        for row_number in range(start_row, sheet.max_row + 1):
            source_url = _cell_link(sheet.cell(row_number, link_column)).strip()
            if not source_url:
                continue
            platform, validation_error = _direct_link_platform(source_url)
            title = str(sheet.cell(row_number, title_column).value or "").strip() if title_column else ""
            screenshot: Path | None = None
            for cell in sheet[row_number]:
                if cell.column == link_column:
                    continue
                screenshot = _local_image_reference(_cell_link(cell), workbook_path)
                if screenshot is None:
                    screenshot = next(
                        iter(embedded_images.get((sheet.title, row_number, cell.column), [])),
                        None,
                    )
                if screenshot is not None:
                    break
            sequence = len(items) + 1
            items.append(
                DirectLinkBatchItem(
                    sequence=sequence,
                    row_number=row_number,
                    source_url=source_url,
                    platform=platform,
                    title=title,
                    validation_error=validation_error,
                    sku_screenshot=_stage_sku_screenshot(screenshot, output_dir, sequence, row_number),
                )
            )
        return items
    finally:
        workbook.close()


DIRECT_REPLACE_HEADERS = {
    "image": {"商品图", "产品图", "我方商品图", "1688商品图"},
    "link": {"商品链接", "对标链接", "竞品链接", "url", "淘宝链接", "天猫链接", "京东链接", "抖音链接", "快手链接"},
    "title": {"标题", "商品标题", "商品名称", "名称"},
    "sku_name": {"sku名称", "sku名", "规格名称"},
    "color": {"颜色", "颜色分类"},
    "spec": {"规格", "型号", "尺码", "容量"},
    "price": {"价格", "sku价格"},
    "sku_image": {"sku参考图", "sku图"},
}


def _row_urls(sheet: Any, row_number: int) -> list[str]:
    values: list[str] = []
    for cell in sheet[row_number]:
        value = _cell_link(cell).strip()
        parsed = urlparse(value)
        if parsed.scheme in {"http", "https"} and parsed.hostname:
            values.append(value)
    return list(dict.fromkeys(values))


def _manual_sku(sheet: Any, row_number: int, columns: dict[str, int], images: dict[tuple[str, int, int], list[Path]]) -> dict[str, Any] | None:
    values = {
        name: str(sheet.cell(row_number, columns.get(name, 0)).value or "").strip() if columns.get(name) else ""
        for name in ("sku_name", "color", "spec", "price")
    }
    reference_images = images.get((sheet.title, row_number, columns.get("sku_image", 0)), []) if columns.get("sku_image") else []
    if not any(values.values()) and not reference_images:
        return None
    values["reference_image"] = str(reference_images[0]) if len(reference_images) == 1 else ""
    values["source_status"] = "reference_image" if values["reference_image"] else "text_conditioned"
    return values


def extract_direct_replace_items(workbook_path: Path, output_dir: Path) -> list[DirectReplaceBatchItem]:
    images = extract_embedded_images(workbook_path, output_dir / "source-images")
    workbook = load_workbook(workbook_path, data_only=False)
    try:
        items: list[DirectReplaceBatchItem] = []
        for sheet in workbook.worksheets:
            if sheet.max_row < 1:
                continue
            header_row = 0
            columns: dict[str, int] = {}
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 10)):
                candidate: dict[str, int] = {}
                for cell in row:
                    label = normalize_header(cell.value)
                    for name, accepted in DIRECT_REPLACE_HEADERS.items():
                        if label in accepted and name not in candidate:
                            candidate[name] = cell.column
                if "image" in candidate and "link" in candidate:
                    header_row = row[0].row
                    columns = candidate
                    break

            if header_row:
                drafts: list[dict[str, Any]] = []
                for row_number in range(header_row + 1, sheet.max_row + 1):
                    source_url = _cell_link(sheet.cell(row_number, columns["link"])).strip()
                    product_images = images.get((sheet.title, row_number, columns["image"]), [])
                    sku = _manual_sku(sheet, row_number, columns, images)
                    if not source_url and not product_images and sku is None:
                        continue
                    if drafts and source_url and source_url == drafts[-1]["source_url"] and not product_images:
                        if sku is not None:
                            drafts[-1]["manual_skus"].append(sku)
                        continue
                    platform, platform_error = _direct_link_platform(source_url) if source_url else ("invalid", "缺少商品链接")
                    errors: list[str] = []
                    if len(product_images) != 1:
                        errors.append("缺少我方商品图" if not product_images else "图片配对冲突")
                    if platform_error:
                        errors.append(platform_error)
                    drafts.append(
                        {
                            "sheet_name": sheet.title,
                            "row_number": row_number,
                            "product_image": product_images[0] if len(product_images) == 1 else None,
                            "source_url": source_url,
                            "platform": platform,
                            "title": str(sheet.cell(row_number, columns.get("title", 0)).value or "").strip() if columns.get("title") else "",
                            "validation_error": "；".join(dict.fromkeys(errors)),
                            "manual_skus": [sku] if sku is not None else [],
                        }
                    )
                for draft in drafts:
                    items.append(
                        DirectReplaceBatchItem(
                            sequence=len(items) + 1,
                            manual_skus=tuple(draft.pop("manual_skus")),
                            **draft,
                        )
                    )
                continue

            image_rows: dict[int, list[Path]] = defaultdict(list)
            for (sheet_name, row_number, _), paths in images.items():
                if sheet_name == sheet.title:
                    image_rows[row_number].extend(paths)
            candidate_rows = sorted(set(image_rows) | {row for row in range(1, sheet.max_row + 1) if _row_urls(sheet, row)})
            for row_number in candidate_rows:
                product_images = image_rows.get(row_number, [])
                urls = _row_urls(sheet, row_number)
                errors: list[str] = []
                if len(product_images) != 1 or len(urls) != 1:
                    if len(product_images) > 1 or len(urls) > 1:
                        errors.append("图片或链接配对冲突")
                    if not product_images:
                        errors.append("缺少我方商品图")
                    if not urls:
                        errors.append("缺少商品链接")
                source_url = urls[0] if len(urls) == 1 else ""
                platform, platform_error = _direct_link_platform(source_url) if source_url else ("invalid", "")
                if platform_error:
                    errors.append(platform_error)
                items.append(
                    DirectReplaceBatchItem(
                        sequence=len(items) + 1,
                        sheet_name=sheet.title,
                        row_number=row_number,
                        product_image=product_images[0] if len(product_images) == 1 else None,
                        source_url=source_url,
                        platform=platform,
                        validation_error="；".join(dict.fromkeys(errors)),
                    )
                )
        return items
    finally:
        workbook.close()


def normalize_direct_manifest(source_manifest: Path, target_path: Path) -> Path:
    source = json.loads(source_manifest.read_text(encoding="utf-8"))
    grouped = {asset_type: [] for asset_type in ASSET_TYPES}
    for image in source.get("images", []):
        asset_type = str(image.get("type") or "")
        path = str(image.get("path") or "")
        if asset_type not in grouped or not path:
            continue
        normalized = dict(image)
        normalized["local_path"] = path
        normalized["status"] = "ok" if Path(path).is_file() else "missing"
        grouped[asset_type].append(normalized)
    product_id = str(source.get("product_id") or "")
    source_url = str(source.get("source_url") or "")
    document = dict(source)
    document["products"] = [
        {
            "product_id": product_id,
            "title": str(source.get("product_title") or source.get("title") or ""),
            "item_url": source_url,
            "types": {
                asset_type: {"images": grouped[asset_type]}
                for asset_type in ASSET_TYPES
            },
        }
    ]
    document["extended_assets"] = document["products"]
    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_text(json.dumps(document, ensure_ascii=False, indent=2), encoding="utf-8")
    return target_path


def resolve_supplement_workbook(workbook_path: Path, project_root: Path) -> SupplementWorkbookContext:
    workbook_path = workbook_path.resolve()
    if not workbook_path.is_file() or workbook_path.suffix.lower() != ".xlsx":
        raise ValueError("请选择程序此前导出的单商品 XLSX 表格。")
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if "总览" not in workbook.sheetnames:
            raise ValueError("该表格不是程序导出的单商品结果表。")
        overview = workbook["总览"]
        values: dict[str, Any] = {}
        for row in overview.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            key = str(row[0] or "").strip()
            if key:
                values[key] = row[1] if len(row) > 1 else None
    finally:
        workbook.close()

    item_root = workbook_path.parent
    manifest_candidates = [
        item_root / "manifest.json",
        item_root / "collected" / "direct-manifest.json",
        item_root / "collected" / "main-image-manifest.json",
    ]
    source_manifest = next((path for path in manifest_candidates if path.is_file()), None)
    if source_manifest is None:
        raise ValueError("表格相邻目录中找不到原采集 manifest，无法安全补图。")
    try:
        source_document = json.loads(source_manifest.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        source_document = {}

    direct_manifest = source_manifest.name in {"manifest.json", "direct-manifest.json"}
    source_url = str(values.get("来源商品链接") or values.get("1688商品链接") or "").strip()
    title = str(values.get("source_title") or values.get("来源商品标题") or values.get("淘宝商品标题") or "").strip()
    platform_label = str(values.get("来源平台") or "").strip()
    platform = {"淘宝": "taobao", "天猫": "tmall", "京东": "jd", "抖音": "douyin", "快手": "kuaishou"}.get(platform_label, "")
    if source_url and not platform:
        platform, _ = _direct_link_platform(source_url)
    own_product_value = str(values.get("我方商品图") or "").strip()
    sequence_match = re.match(r"(\d+)-row-(\d+)", item_root.name)
    sequence = int(sequence_match.group(1)) if sequence_match else 1
    row_number = int(values.get("source_row") or (sequence_match.group(2) if sequence_match else 1))
    if direct_manifest and own_product_value:
        configured_product = Path(own_product_value).expanduser()
        if not configured_product.is_absolute():
            configured_product = item_root / configured_product
        product_image = configured_product if configured_product.is_file() else next(
            (
                path
                for path in sorted((item_root.parent / "source-images").glob(f"{sequence:03d}-row-{row_number:04d}.*"))
                if path.is_file()
            ),
            None,
        )
        if product_image is None:
            raise ValueError("该表格缺少可定位的我方商品图，无法安全补图。")
        item: BatchItem | DirectLinkBatchItem | DirectReplaceBatchItem = DirectReplaceBatchItem(
            sequence,
            str(values.get("source_sheet") or ""),
            row_number,
            product_image,
            source_url,
            platform,
            title,
            manual_skus=tuple(
                {
                    "sku_name": str(variant.get("sku_label") or ""),
                    "color": str(variant.get("color_text") or ""),
                    "spec": str(variant.get("spec_text") or variant.get("net_content") or ""),
                    "price": str(variant.get("after_coupon_price") or variant.get("list_price") or ""),
                    "reference_image": str(variant.get("reference_image") or ""),
                    "source_status": str(variant.get("source_status") or "text_conditioned"),
                }
                for variant in source_document.get("sku_variants", [])[:SUPPLEMENT_LIMITS["sku"]]
                if isinstance(variant, dict)
            ),
        )
        generation_mode = "own_product"
    elif direct_manifest and platform in {"taobao", "tmall", "jd", "douyin", "kuaishou"}:
        item: BatchItem | DirectLinkBatchItem = DirectLinkBatchItem(
            1,
            int(values.get("source_row") or 1),
            source_url,
            platform,
            title,
            manual_skus=tuple(
                {
                    "sku_name": str(variant.get("sku_label") or ""),
                    "color": str(variant.get("color_text") or ""),
                    "spec": str(variant.get("spec_text") or variant.get("net_content") or ""),
                    "price": str(variant.get("after_coupon_price") or variant.get("list_price") or ""),
                    "reference_image": str(variant.get("reference_image") or ""),
                    "source_status": str(variant.get("source_status") or "text_conditioned"),
                    "visual_confidence": variant.get("visual_confidence", ""),
                    "quality_note": str(variant.get("quality_note") or ""),
                }
                for variant in source_document.get("sku_variants", [])[:SUPPLEMENT_LIMITS["sku"]]
                if isinstance(variant, dict)
            ),
        )
        generation_mode = "competitor_reference"
    else:
        source_images = item_root.parent / "source-images"
        product_image = next(
            (
                path
                for path in sorted(source_images.glob(f"{sequence:03d}-row-{row_number:04d}.*"))
                if path.is_file()
            ),
            None,
        )
        if product_image is None:
            raise ValueError("该表格缺少可定位的商品身份图，无法安全补图。")
        item = BatchItem(sequence, row_number, title, product_image, source_url, "", "")
        generation_mode = "own_product"

    generated_root = item_root / "generated"
    if not (generated_root / "analysis.json").is_file() and direct_manifest:
        try:
            source = json.loads(source_manifest.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            source = {}
        source_paths = {
            _normalized_file_path(image.get("path") or image.get("local_path") or "")
            for image in source.get("images", [])
            if isinstance(image, dict) and (image.get("path") or image.get("local_path"))
        }
        global_generated = project_root / "outputs" / "generated"
        for analysis_path in sorted(global_generated.glob("*/analysis.json"), key=lambda path: path.stat().st_mtime, reverse=True):
            try:
                records = json.loads(analysis_path.read_text(encoding="utf-8")).get("records", [])
            except (OSError, json.JSONDecodeError):
                continue
            record_sources = {
                _normalized_file_path(record.get("source_path") or "")
                for record in records
                if isinstance(record, dict) and record.get("source_path")
            }
            if record_sources and record_sources.issubset(source_paths):
                generated_root = analysis_path.parent
                break
    return SupplementWorkbookContext(
        workbook_path=workbook_path,
        item=item,
        item_root=item_root,
        source_manifest=source_manifest,
        generated_root=generated_root,
        generation_mode=generation_mode,
    )


class DirectLinkCollector:
    def __init__(
        self,
        project_root: Path,
        profile_dir: Path,
        browser_executable: str = "",
        max_main_images: int | None = DEFAULT_MAIN_IMAGES,
        callback: BatchCallback | None = None,
    ) -> None:
        self.project_root = project_root
        self.profile_dir = profile_dir
        self.browser_executable = browser_executable.strip()
        self.max_main_images = max_main_images
        self.callback = callback
        self.process: subprocess.Popen[str] | None = None

    def cancel(self) -> None:
        process = self.process
        if process and process.poll() is None:
            subprocess.run(["taskkill", "/PID", str(process.pid), "/F"], check=False, capture_output=True)

    def collect(self, item: DirectLinkBatchItem | DirectReplaceBatchItem, item_root: Path) -> Path:
        collector_executable = self.project_root / "store_insight_collector.exe"
        collector_script = self.project_root / "store_insight_collector.py"
        command = [str(collector_executable)] if collector_executable.is_file() else [sys.executable, str(collector_script)]
        item_url = resolve_direct_item_url(item.source_url)
        command.extend(
            [
                item_url,
                "--output",
                str(item_root / "collected"),
                "--auto-launch",
                "--reuse-existing-cdp",
                "--profile-dir",
                str(self.profile_dir),
                "--types",
                *ASSET_TYPES,
            ]
        )
        if self.browser_executable:
            command.extend(["--browser-executable", self.browser_executable])
        if self.max_main_images is not None:
            command.extend(["--max-main-images", str(self.max_main_images)])
        self.process = subprocess.Popen(
            command,
            cwd=str(self.project_root),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        manifest: Path | None = None
        pause_message = ""
        assert self.process.stdout is not None
        for line in self.process.stdout:
            message = line.strip()
            if message and self.callback:
                self.callback({"stage": "collecting", "sequence": item.sequence, "message": message})
            if message.startswith("[collector] manifest: "):
                manifest = Path(message.removeprefix("[collector] manifest: ").strip())
            if any(
                marker in message
                for marker in (
                    "登录页面",
                    "login was not completed",
                    "平台验证",
                    "访问被拒",
                    "访问受限",
                    "请求太频繁",
                    "操作过于频繁",
                )
            ):
                pause_message = message
        exit_code = self.process.wait()
        self.process = None
        if exit_code != 0:
            if pause_message:
                raise BatchCollectionPaused(pause_message)
            raise RuntimeError(f"直接链接采集失败，退出码 {exit_code}")
        if not manifest or not manifest.is_file():
            raise RuntimeError("直接链接采集结束但未生成 manifest")
        return manifest


def _safe_name(value: str, fallback: str, limit: int = 72) -> str:
    cleaned = INVALID_FILENAME.sub("_", re.sub(r"\s+", " ", value).strip()).strip(" ._")
    return (cleaned or fallback)[:limit]


def _normalized_file_path(path: Path | str) -> str:
    return os.path.normcase(os.path.abspath(str(path)))


def _same_file_path(first: Path | str, second: Path | str) -> bool:
    try:
        return os.path.samefile(first, second)
    except OSError:
        return _normalized_file_path(first) == _normalized_file_path(second)


def load_batch_results(workbook_path: Path, output_root: Path) -> list[dict[str, Any]]:
    checkpoint_path = output_root / BATCH_RESULTS_FILENAME
    try:
        document = json.loads(checkpoint_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    if not _same_file_path(document.get("source", ""), workbook_path):
        return []
    items = document.get("items")
    if not isinstance(items, list):
        return []
    return [item for item in items if isinstance(item, dict) and int(item.get("sequence") or 0) > 0]


def save_batch_results(
    workbook_path: Path,
    output_root: Path,
    results: list[dict[str, Any]],
    *,
    batch_mode: str | None = None,
    run_mode: str | None = None,
    total: int | None = None,
) -> Path:
    output_root.mkdir(parents=True, exist_ok=True)
    checkpoint_path = output_root / BATCH_RESULTS_FILENAME
    temporary_path = checkpoint_path.with_suffix(".json.tmp")
    temporary_path.write_text(
        json.dumps(
            {
                "source": str(workbook_path.resolve()),
                "updated_at": datetime.now().isoformat(timespec="seconds"),
                "items": sorted(results, key=lambda item: int(item.get("sequence") or 0)),
                **({"batch_mode": batch_mode} if batch_mode else {}),
                **({"run_mode": run_mode} if run_mode else {}),
                **({"total": total} if total is not None else {}),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    temporary_path.replace(checkpoint_path)
    return checkpoint_path


def _record_output_is_valid(record: dict[str, Any]) -> bool:
    return (
        record.get("status") == "completed"
        and bool(str(record.get("output_path") or ""))
        and Path(str(record.get("output_path"))).is_file()
    )


def summarize_generation_result(
    records: list[dict[str, Any]], expected_categories: tuple[str, ...]
) -> tuple[int, int, tuple[str, ...]]:
    valid_records = [record for record in records if _record_output_is_valid(record)]
    valid_categories = {str(record.get("category") or "") for record in valid_records}
    missing_categories = tuple(
        category for category in expected_categories if category not in valid_categories
    )
    failed_count = len(records) - len(valid_records)
    return len(valid_records), failed_count, missing_categories


def _batch_result_is_valid(result: dict[str, Any], output_root: Path) -> bool:
    if result.get("status") != "completed":
        return False
    workbook = Path(str(result.get("workbook") or ""))
    if not workbook.is_file():
        return False
    analysis_path = workbook.parent / "generated" / "analysis.json"
    try:
        document = json.loads(analysis_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return False
    return any(
        isinstance(record, dict) and _record_output_is_valid(record)
        for record in document.get("records", [])
    )


def plan_supplement_ordinals(
    records: list[dict[str, Any]], category: str, count: int
) -> list[int]:
    """Return missing ordinals first, then append new ordinals after the current range."""
    if category not in SUPPLEMENT_LIMITS:
        raise ValueError(f"不支持补充生成类型：{category}")
    if count < 1:
        raise ValueError("补充数量必须至少为 1")
    limit = SUPPLEMENT_LIMITS[category]
    by_ordinal = {
        int(record.get("ordinal") or 0): record
        for record in records
        if record.get("category") == category and int(record.get("ordinal") or 0) > 0
    }
    missing = [
        ordinal
        for ordinal in range(1, max(by_ordinal, default=0) + 1)
        if not _record_output_is_valid(by_ordinal.get(ordinal, {}))
    ]
    next_ordinal = max(by_ordinal, default=0) + 1
    available = max(0, limit - next_ordinal + 1)
    selected = missing[:count]
    remaining = count - len(selected)
    if remaining > available:
        raise ValueError(
            f"{category} 图最多生成 {limit} 张，当前只能补充 {len(selected) + available} 张。"
        )
    selected.extend(range(next_ordinal, next_ordinal + remaining))
    return selected


def plan_all_supplement_ordinals(
    records: list[dict[str, Any]], manifest_path: Path | None = None
) -> dict[str, list[int]]:
    """Plan existing failed or missing output slots for every asset category."""
    source_counts = {category: 0 for category in ASSET_TYPES}
    if manifest_path is not None and manifest_path.is_file():
        try:
            document = json.loads(manifest_path.read_text(encoding="utf-8"))
            for entry in document.get("images", []):
                category = str(entry.get("type") or "") if isinstance(entry, dict) else ""
                if category in source_counts:
                    raw_path = Path(str(entry.get("path") or ""))
                    path = raw_path if raw_path.is_absolute() else manifest_path.parent / raw_path
                    if path.is_file():
                        source_counts[category] += 1
        except (OSError, json.JSONDecodeError):
            pass
    planned: dict[str, list[int]] = {}
    for category in ASSET_TYPES:
        category_records = [
            record for record in records
            if isinstance(record, dict) and record.get("category") == category
        ]
        ordinals = {
            int(record.get("ordinal") or 0)
            for record in category_records
            if int(record.get("ordinal") or 0) > 0
        }
        missing = [
            ordinal
            for ordinal in range(1, max(ordinals, default=0) + 1)
            if not _record_output_is_valid(
                next((record for record in category_records if int(record.get("ordinal") or 0) == ordinal), {})
            )
        ]
        if missing:
            planned[category] = missing
            continue
        if not category_records and source_counts[category]:
            target_count = source_counts[category]
            if category == "sku":
                target_count = min(8, max(3, target_count))
            elif category == "detail":
                target_count = min(15, max(6, target_count))
            planned[category] = list(range(1, target_count + 1))
    return planned


def merge_generation_records(
    analysis_path: Path, previous: list[dict[str, Any]], added: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    merged: dict[tuple[str, int], dict[str, Any]] = {}
    for record in [*previous, *added]:
        key = (str(record.get("category") or ""), int(record.get("ordinal") or 0))
        if key[0] and key[1] > 0:
            merged[key] = dict(record)
    ordered = sorted(merged.values(), key=lambda item: (item.get("category", ""), int(item.get("ordinal") or 0)))
    try:
        document = json.loads(analysis_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        document = {"schema_version": 3}
    document["records"] = ordered
    analysis_path.parent.mkdir(parents=True, exist_ok=True)
    analysis_path.write_text(json.dumps(document, ensure_ascii=False, indent=2), encoding="utf-8")
    return ordered


def plan_batch_retry_ordinals(
    manifest_path: Path,
    categories: tuple[str, ...],
    previous_records: list[dict[str, Any]],
    max_main_images: int | None,
) -> dict[str, list[int]]:
    expected_tasks = load_manifest_tasks(
        manifest_path,
        categories,
        max_main_images=max_main_images,
    )
    valid_keys = {
        (str(record.get("category") or ""), int(record.get("ordinal") or 0))
        for record in previous_records
        if _record_output_is_valid(record)
    }
    planned: dict[str, list[int]] = {}
    for task in expected_tasks:
        key = (task.category, task.ordinal)
        if key in valid_keys:
            continue
        planned.setdefault(task.category, []).append(task.ordinal)
    return planned


def _manifest_image_paths(manifest_path: Path, category: str) -> list[Path]:
    document = json.loads(manifest_path.read_text(encoding="utf-8"))
    paths: list[Path] = []
    for entry in document.get("images", []):
        if not isinstance(entry, dict) or entry.get("type") != category:
            continue
        raw_path = Path(str(entry.get("path") or ""))
        path = raw_path if raw_path.is_absolute() else manifest_path.parent / raw_path
        if path.is_file():
            paths.append(path.resolve())
    return paths


def _supplement_detail_plans(
    output_root: Path,
    ordinals: list[int],
    sources: list[IdentitySource],
) -> tuple[dict[str, Any] | None, Path | None, list[DetailViewPlan]]:
    dossier_path = output_root / "product-dossier.json"
    dossier: dict[str, Any] | None = None
    plans_by_ordinal: dict[int, DetailViewPlan] = {}
    if dossier_path.is_file():
        try:
            document = json.loads(dossier_path.read_text(encoding="utf-8"))
            dossier = document.get("dossier") if isinstance(document.get("dossier"), dict) else None
            for raw in document.get("detail_view_plans", []):
                if not isinstance(raw, dict):
                    continue
                ordinal = int(raw.get("ordinal") or 0)
                if ordinal < 1:
                    continue
                plans_by_ordinal[ordinal] = DetailViewPlan(
                    ordinal=ordinal,
                    view_type=str(raw.get("view_type") or "detail_closeup"),
                    focus=str(raw.get("focus") or "展示商品已验证的材质和工艺细节"),
                    supporting_source_index=(
                        int(raw["supporting_source_index"])
                        if raw.get("supporting_source_index") not in (None, "")
                        else None
                    ),
                    inferred_view=bool(raw.get("inferred_view")),
                    prohibited_inventions=tuple(raw.get("prohibited_inventions") or ()),
                )
        except (OSError, ValueError, TypeError, json.JSONDecodeError):
            pass
    if not sources:
        raise ValueError("没有可用于补充详情图的采集素材")
    plans: list[DetailViewPlan] = []
    for ordinal in ordinals:
        plan = plans_by_ordinal.get(ordinal)
        if plan is None:
            source = sources[(ordinal - 1) % len(sources)]
            plan = DetailViewPlan(
                ordinal=ordinal,
                view_type="detail_closeup",
                focus="展示参考素材中已经可见的材质、纹理、结构或使用细节，不新增未验证部件",
                supporting_source_index=source.index,
                inferred_view=True,
                prohibited_inventions=(),
            )
        plans.append(plan)
    return dossier, dossier_path if dossier_path.is_file() else None, plans


def _1688_image_columns(sheet: Any) -> set[int]:
    columns: set[int] = set()
    for row in sheet.iter_rows(max_row=min(sheet.max_row, 10)):
        for cell in row:
            label = re.sub(r"\s+", "", str(cell.value or "")).lower()
            if "1688" in label and "商品图" in label:
                columns.add(cell.column - 1)
    return columns


def extract_batch_items(workbook_path: Path, output_dir: Path) -> list[BatchItem]:
    workbook = load_workbook(workbook_path, data_only=True)
    try:
        sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
        anchored_images: list[tuple[int, int, Any]] = []
        for image in sheet._images:
            anchor = getattr(image, "anchor", None)
            marker = getattr(anchor, "_from", None)
            if marker is None:
                continue
            anchored_images.append((marker.row + 1, marker.col, image))

        preferred_columns = _1688_image_columns(sheet)
        if preferred_columns:
            anchored_images = [entry for entry in anchored_images if entry[1] in preferred_columns]
        anchored_images.sort(key=lambda entry: (entry[0], entry[1]))

        image_dir = output_dir / "source-images"
        image_dir.mkdir(parents=True, exist_ok=True)
        items: list[BatchItem] = []
        for sequence, (row_number, _, image) in enumerate(anchored_images, start=1):
            suffix = "." + str(getattr(image, "format", "png") or "png").lower().lstrip(".")
            image_path = image_dir / f"{sequence:03d}-row-{row_number:04d}{suffix}"
            image_path.write_bytes(image._data())
            items.append(
                BatchItem(
                    sequence=sequence,
                    row_number=row_number,
                    title=str(sheet.cell(row_number, 1).value or "").strip(),
                    product_image=image_path,
                    source_url=_cell_link(sheet.cell(row_number, 8)),
                    purchase_price=sheet.cell(row_number, 6).value,
                    shipping_fee=sheet.cell(row_number, 7).value,
                )
            )
        return items
    finally:
        workbook.close()


def _resolve_manifest_asset_path(raw_path: Any, source_manifest: Path) -> Path | None:
    value = str(raw_path or "").strip()
    if not value:
        return None
    path = Path(value)
    if path.is_file():
        return path.resolve()

    item_root = source_manifest.parent.parent
    relative_candidate = source_manifest.parent / path
    if relative_candidate.is_file():
        return relative_candidate.resolve()

    normalized_parts = [part for part in value.replace("/", "\\").split("\\") if part]
    try:
        collected_index = next(
            index for index, part in enumerate(normalized_parts)
            if part.lower() == "collected"
        )
    except StopIteration:
        return None
    rebased = item_root / "collected" / Path(*normalized_parts[collected_index + 1 :])
    return rebased.resolve() if rebased.is_file() else None


def _asset_images(
    document: dict[str, Any], asset_type: str, source_manifest: Path | None = None
) -> list[dict[str, Any]]:
    images: list[dict[str, Any]] = []
    if not document.get("extended_assets") and document.get("images"):
        for image in document.get("images", []):
            if not isinstance(image, dict) or image.get("type") != asset_type:
                continue
            path = str(image.get("local_path") or image.get("path") or "")
            if not path:
                continue
            normalized = dict(image)
            resolved = (
                _resolve_manifest_asset_path(path, source_manifest)
                if source_manifest
                else Path(path) if Path(path).is_file() else None
            )
            normalized["local_path"] = str(resolved or path)
            normalized["status"] = "ok" if resolved else "missing"
            if normalized["status"] == "ok":
                images.append(normalized)
        return images
    for product in document.get("extended_assets", []):
        type_entry = (product.get("types") or {}).get(asset_type) or {}
        for image in type_entry.get("images", []):
            if image.get("status") in {"ok", "metadata_only"}:
                normalized = dict(image)
                if source_manifest:
                    raw_path = normalized.get("local_path") or normalized.get("path")
                    resolved = _resolve_manifest_asset_path(raw_path, source_manifest)
                    if resolved:
                        normalized["local_path"] = str(resolved)
                        normalized["status"] = "ok"
                    elif normalized.get("status") == "ok":
                        continue
                images.append(normalized)
    return images


def _collected_manifest_is_valid(manifest_path: Path) -> bool:
    try:
        document = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return False
    if str(document.get("status") or "").lower() in {"partial", "paused", "empty", "failed"}:
        return False
    if any(_asset_images(document, asset_type, manifest_path) for asset_type in ASSET_TYPES):
        return True
    requested = {
        str(asset_type)
        for asset_type in document.get("requested_asset_types", [])
        if str(asset_type) in ASSET_TYPES
    }
    missing = {
        str(asset_type)
        for asset_type in document.get("missing_asset_types", [])
        if str(asset_type) in ASSET_TYPES
    }
    return bool(requested) and requested.issubset(missing)


def build_generation_manifest(source_manifest: Path, target_path: Path) -> Path:
    document = json.loads(source_manifest.read_text(encoding="utf-8"))
    flattened: list[dict[str, Any]] = []
    for asset_type in ASSET_TYPES:
        for image in _asset_images(document, asset_type, source_manifest):
            local_path = str(image.get("local_path") or "")
            if image.get("status") != "ok" or not local_path or not Path(local_path).is_file():
                continue
            flattened.append({"type": asset_type, "path": local_path})
    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "created_at": datetime.now().isoformat(timespec="seconds"),
                "source_manifest": str(source_manifest),
                "images": flattened,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return target_path


def merge_manual_sku_metadata(
    document: dict[str, Any],
    item: DirectLinkBatchItem | DirectReplaceBatchItem,
) -> dict[str, Any]:
    if not item.manual_skus:
        return document
    if document.get("sku_variants") and not (
        isinstance(item, DirectLinkBatchItem) and item.sku_screenshot
    ):
        return document
    updated = dict(document)
    variants: list[dict[str, Any]] = []
    screenshot_conditioned = any(
        raw.get("source_status") == "screenshot_thumbnail"
        for raw in item.manual_skus
        if isinstance(raw, dict)
    )
    for index, raw in enumerate(item.manual_skus[:SUPPLEMENT_LIMITS["sku"]], start=1):
        source_status = str(raw.get("source_status") or "text_conditioned")
        reference_image = str(raw.get("reference_image") or "")
        variants.append(
            {
                "source_index": str(index),
                "sku_id": "",
                "sku_label": str(raw.get("sku_name") or "").strip(),
                "color_text": str(raw.get("color") or "").strip(),
                "spec_text": str(raw.get("spec") or "").strip(),
                "list_price": str(raw.get("price") or "").strip(),
                "after_coupon_price": "",
                "reference_image": reference_image,
                "source_status": "reference_image" if reference_image else source_status,
                "visual_confidence": raw.get("visual_confidence", ""),
                "quality_note": str(raw.get("quality_note") or "").strip(),
                "parse_status": source_status or "manual",
            }
        )
    updated["sku_variants"] = variants
    updated["sku_metadata_status"] = (
        "screenshot"
        if screenshot_conditioned
        else "manual"
    )
    updated["sku_metadata_error"] = ""
    return updated


def add_manual_sku_generation_sources(
    generation_manifest: Path,
    item: DirectLinkBatchItem | DirectReplaceBatchItem,
    fallback_image: Path,
) -> Path:
    if not item.manual_skus:
        return generation_manifest
    document = json.loads(generation_manifest.read_text(encoding="utf-8"))
    images = [image for image in document.get("images", []) if isinstance(image, dict)]
    existing_sku = [image for image in images if image.get("type") == "sku"]
    if existing_sku:
        return generation_manifest
    non_sku = [image for image in images if image.get("type") != "sku"]
    manual_images: list[dict[str, Any]] = []
    for raw in item.manual_skus[:SUPPLEMENT_LIMITS["sku"]]:
        reference_image = Path(str(raw.get("reference_image") or ""))
        if reference_image.is_file():
            source_path = reference_image
        else:
            source_path = fallback_image
        if not source_path.is_file():
            continue
        manual_images.append(
            {
                "type": "sku",
                "path": str(source_path),
                "manual_sku": dict(raw),
            }
        )
    document["images"] = non_sku + manual_images
    generation_manifest.write_text(
        json.dumps(document, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return generation_manifest


def restore_collected_manifest(item_root: Path) -> tuple[Path, int] | None:
    collected_root = item_root / "collected"
    manifest_path = collected_root / "main-image-manifest.json"
    asset_manifest_path = collected_root / "_work" / "extended-competitor-assets.json"
    direct_manifest_path = collected_root / "direct-manifest.json"
    for source_path in (manifest_path, asset_manifest_path, direct_manifest_path):
        try:
            source = json.loads(source_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue

        if str(source.get("status") or "").lower() in {"partial", "paused", "empty", "failed"}:
            continue

        assets = source.get("extended_assets")
        if not isinstance(assets, list):
            assets = source.get("products")
        if not isinstance(assets, list):
            continue

        usable = 0
        for product in assets:
            for asset_type in ASSET_TYPES:
                for image in ((product.get("types") or {}).get(asset_type) or {}).get("images", []):
                    local_path = str(image.get("local_path") or "")
                    if image.get("status") == "ok" and local_path and Path(local_path).is_file():
                        usable += 1
        if not usable:
            continue

        if source_path == asset_manifest_path:
            restored = dict(source)
            restored["extended_assets"] = assets
            restored["updated_at"] = datetime.now().isoformat(timespec="seconds")
            manifest_path.write_text(json.dumps(restored, ensure_ascii=False, indent=2), encoding="utf-8")
            return manifest_path, usable
        return source_path, usable
    return None


def _file_sha256(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def find_prior_collected_manifest(product_image: Path, item_root: Path, output_root: Path) -> tuple[Path, int] | None:
    try:
        image_hash = _file_sha256(product_image)
    except OSError:
        return None
    source_images = sorted(
        output_root.parent.glob("*/source-images/*"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    for source_image in source_images:
        try:
            if source_image.resolve() == product_image.resolve() or _file_sha256(source_image) != image_hash:
                continue
        except OSError:
            continue
        candidate_root = source_image.parent.parent / source_image.stem
        if candidate_root.resolve() == item_root.resolve():
            continue
        restored = restore_collected_manifest(candidate_root)
        if restored:
            return restored
    return None


def _direct_collection_identity(value: str, product_id: str = "") -> tuple[str, str] | None:
    candidate = value.strip()
    platform, validation_error = _direct_link_platform(candidate)
    if validation_error:
        return None
    parsed = urlparse(candidate)
    host = (parsed.hostname or "").lower()
    resolved_product_id = str(product_id or "").strip()
    if not resolved_product_id:
        if platform in {"taobao", "tmall", "douyin", "kuaishou"}:
            resolved_product_id = parse_qs(parsed.query).get("id", [""])[0]
        elif platform == "jd":
            matched = re.search(r"/(\d+)\.html", parsed.path)
            resolved_product_id = matched.group(1) if matched else ""
    if not resolved_product_id and host in TAOBAO_SHORT_HOSTS | {"u.jd.com", "3.cn", "v.douyin.com"}:
        try:
            return _direct_collection_identity(resolve_direct_item_url(candidate))
        except ValueError:
            return None
    return (platform, resolved_product_id) if resolved_product_id else None


def find_prior_direct_collected_manifest(
    source_url: str,
    search_roots: tuple[Path, ...],
    exclude_root: Path | None = None,
) -> tuple[Path, int] | None:
    identity = _direct_collection_identity(source_url)
    if identity is None:
        return None
    excluded = exclude_root.resolve() if exclude_root else None
    candidates: set[Path] = set()
    for root in search_roots:
        if not root.is_dir():
            continue
        candidates.update(root.rglob("direct-manifest.json"))
        candidates.update(root.rglob("manifest.json"))

    def modified_at(path: Path) -> float:
        try:
            return path.stat().st_mtime
        except OSError:
            return 0.0

    for candidate in sorted(candidates, key=modified_at, reverse=True):
        resolved = candidate.resolve()
        if excluded and (resolved == excluded or excluded in resolved.parents):
            continue
        try:
            document = json.loads(candidate.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        products = document.get("products") or [{}]
        top_product = products[0] if isinstance(products[0], dict) else {}
        candidate_url = str(document.get("source_url") or top_product.get("item_url") or "")
        candidate_product_id = str(document.get("product_id") or top_product.get("product_id") or "")
        if _direct_collection_identity(candidate_url, candidate_product_id) != identity:
            continue
        if not _collected_manifest_is_valid(candidate):
            continue
        asset_count = sum(
            len(_asset_images(document, asset_type, candidate))
            for asset_type in ASSET_TYPES
        )
        return candidate, asset_count
    return None


def _normalized_path(value: Any) -> str:
    text = str(value or "")
    return os.path.normcase(os.path.abspath(text)) if text else ""


def _public_url(record: dict[str, Any], *names: str) -> str:
    for name in names:
        value = str(record.get(name) or "")
        if value:
            return value
    return ""


def _generation_status(record: dict[str, Any]) -> str:
    return "生成成功" if record.get("status") == "completed" else str(record.get("status") or "失败")


def _image_rows(source_images: list[dict[str, Any]], generated: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = [
        {
            "source_path": str(image.get("local_path") or ""),
            "source_public_url": _public_url(image, "public_url", "oss_url", "remote_url"),
            "output_path": "",
            "output_public_url": "",
            "generation_status": "未生成",
        }
        for image in source_images
    ]
    source_indexes = [index for index, row in enumerate(rows) if row["source_path"]]
    by_path: dict[str, list[int]] = {}
    for index in source_indexes:
        normalized = _normalized_path(rows[index]["source_path"])
        if normalized:
            by_path.setdefault(normalized, []).append(index)

    assigned_sources: set[int] = set()
    generated_only: list[dict[str, Any]] = []
    for record in sorted(generated, key=lambda value: int(value.get("ordinal") or 0)):
        source_path = str(record.get("source_path") or "")
        ordinal = int(record.get("ordinal") or 0)
        source_index = source_indexes[ordinal - 1] if 0 < ordinal <= len(source_indexes) else None
        if source_index in assigned_sources:
            source_index = None
        if source_index is None and source_path:
            source_index = next(
                (index for index in by_path.get(_normalized_path(source_path), []) if index not in assigned_sources),
                None,
            )
        if source_index is not None:
            assigned_sources.add(source_index)
            row = rows[source_index]
        else:
            row = {
                "source_path": source_path,
                "source_public_url": _public_url(record, "source_public_url", "public_url", "remote_url"),
                "output_path": "",
                "output_public_url": "",
                "generation_status": "未生成",
            }
            generated_only.append(row)
        row["output_path"] = str(record.get("output_path") or "")
        row["output_public_url"] = _public_url(record, "output_public_url", "public_url", "oss_url")
        row["generation_status"] = _generation_status(record)
    return rows + generated_only


def _sku_rows(
    source: dict[str, Any],
    generated: list[dict[str, Any]],
    top_product: dict[str, Any],
    include_metadata_only: bool = False,
) -> list[dict[str, Any]]:
    source_images = [
        image
        for image in _asset_images(source, "sku")
        if str(image.get("local_path") or "").strip()
    ][:8]
    source_rows: list[dict[str, Any]] = []
    for image in source_images:
        row = {
            "product_id": image.get("product_id") or top_product.get("product_id", ""),
            "sku_id": image.get("sku_id", ""),
            "sku_label": image.get("sku_label", ""),
            "spec_text": image.get("spec_text") or image.get("net_content", ""),
            "color_text": image.get("color_text", ""),
            "price": image.get("after_coupon_price") or image.get("list_price") or "",
            "parse_status": image.get("parse_status") or image.get("metadata_status", ""),
            "source_path": str(image.get("local_path") or ""),
            "source_public_url": _public_url(image, "public_url", "oss_url", "remote_url"),
            "output_path": "",
            "output_public_url": "",
            "generation_status": "未生成",
            "_source_index": str(image.get("source_index") or ""),
        }
        source_rows.append(row)

    variants = source.get("sku_variants") or []
    for variant in variants:
        source_index = str(variant.get("source_index") or "")
        sku_id = str(variant.get("sku_id") or "")
        existing = next(
            (row for row in source_rows if (source_index and row["_source_index"] == source_index) or (sku_id and row["sku_id"] == sku_id)),
            None,
        )
        if existing is None:
            if not include_metadata_only:
                continue
            existing = {
                "product_id": variant.get("product_id") or top_product.get("product_id", ""),
                "sku_id": sku_id,
                "sku_label": variant.get("sku_label", ""),
                "spec_text": variant.get("spec_text") or variant.get("net_content", ""),
                "color_text": variant.get("color_text", ""),
                "price": variant.get("after_coupon_price") or variant.get("list_price") or "",
                "parse_status": variant.get("parse_status", ""),
                "source_path": str(variant.get("reference_image") or ""),
                "source_public_url": "",
                "output_path": "",
                "output_public_url": "",
                "generation_status": "无 SKU 图片",
                "_source_index": source_index,
            }
            source_rows.append(existing)
        for field in ("sku_label", "spec_text", "color_text", "parse_status"):
            if variant.get(field) and not existing.get(field):
                existing[field] = variant[field]
        if not existing.get("price"):
            existing["price"] = variant.get("after_coupon_price") or variant.get("list_price") or ""
        if not existing.get("sku_id"):
            existing["sku_id"] = sku_id

    available_indexes = [index for index, row in enumerate(source_rows) if row.get("source_path")]
    rows = [dict(row) for row in source_rows]
    used: set[int] = set()
    for record in sorted(generated, key=lambda value: int(value.get("ordinal") or 0)):
        source_path = str(record.get("source_path") or "")
        ordinal = int(record.get("ordinal") or 0)
        manual_sku = record.get("manual_sku") if isinstance(record.get("manual_sku"), dict) else None
        source_index = None
        if manual_sku:
            expected = {
                "sku_label": str(manual_sku.get("sku_name") or "").strip(),
                "spec_text": str(manual_sku.get("spec") or "").strip(),
                "color_text": str(manual_sku.get("color") or "").strip(),
                "price": str(manual_sku.get("price") or "").strip(),
            }
            source_index = next(
                (
                    index
                    for index, candidate in enumerate(rows)
                    if index not in used
                    and any(expected.values())
                    and all(
                        not value or str(candidate.get(field) or "").strip() == value
                        for field, value in expected.items()
                    )
                ),
                None,
            )
            if source_index is None:
                manual_indexes = [
                    index
                    for index, candidate in enumerate(rows)
                    if index not in used and candidate.get("parse_status") == "manual"
                ]
                source_index = manual_indexes[ordinal - 1] if 0 < ordinal <= len(manual_indexes) else None
        elif 0 < ordinal <= len(available_indexes):
            source_index = available_indexes[ordinal - 1]
        if source_index in used:
            source_index = None
        if source_index is None and source_path:
            source_index = next(
                (
                    index
                    for index, candidate in enumerate(rows)
                    if index not in used
                    and _normalized_path(candidate.get("source_path")) == _normalized_path(source_path)
                ),
                None,
            )
        if source_index is None:
            source_index = next((index for index in available_indexes if index not in used), None)
        if source_index is None:
            continue
        else:
            used.add(source_index)
            row = rows[source_index]
        row["source_path"] = str(row.get("source_path") or source_path)
        row["output_path"] = str(record.get("output_path") or "")
        row["output_public_url"] = _public_url(record, "output_public_url", "public_url", "oss_url")
        row["generation_status"] = _generation_status(record)
        if source_index is None:
            rows.append(row)

    for row in rows:
        row.pop("_source_index", None)
    return rows[:SUPPLEMENT_LIMITS["sku"]]


def ensure_workbook_available(output_path: Path) -> None:
    lock_path = output_path.with_name(f"~${output_path.name}")
    if lock_path.exists():
        raise RuntimeError(
            f"Excel/WPS 正在占用结果表格：{output_path}。"
            "请关闭该表格后再重试补图，已生成的图片不会丢失。"
        )


def export_product_workbook(
    output_path: Path,
    item: BatchItem | DirectLinkBatchItem | DirectReplaceBatchItem,
    source_manifest: Path,
    generation_records: list[dict[str, Any]],
    titles: dict[str, Any] | None = None,
    project_root: Path | None = None,
    include_metadata_only_skus: bool = False,
) -> Path:
    source = json.loads(source_manifest.read_text(encoding="utf-8"))
    generated = {
        asset_type: sorted(
            [record for record in generation_records if record.get("category") == asset_type],
            key=lambda record: int(record.get("ordinal") or 0),
        )
        for asset_type in ASSET_TYPES
    }
    top_product = (source.get("products") or [{}])[0]
    source_main = _asset_images(source, "main")
    source_detail = _asset_images(source, "detail")
    main_rows = _image_rows(source_main, generated["main"])
    detail_rows = _image_rows(source_detail, generated["detail"])[:SUPPLEMENT_LIMITS["detail"]]
    sku_rows = _sku_rows(source, generated["sku"], top_product, include_metadata_only_skus)
    main_completed = sum(record.get("status") == "completed" for record in generated["main"])
    detail_completed = sum(record.get("status") == "completed" for record in generated["detail"])
    sku_completed = sum(record.get("generation_status") == "生成成功" for record in sku_rows)
    parameters = [
        {
            "type": "商品参数",
            "name": row.get("name", ""),
            "value": row.get("value", ""),
            "handling": row.get("handling") or "采集原值",
        }
        for row in source.get("product_parameters", [])
    ]
    video_url = str(source.get("main_video_url") or "")
    if isinstance(item, (DirectLinkBatchItem, DirectReplaceBatchItem)):
        platform_name = {"taobao": "淘宝", "tmall": "天猫", "jd": "京东", "douyin": "抖音", "kuaishou": "快手"}.get(item.platform, item.platform)
        source_overview = [
            ["来源平台", platform_name],
            ["来源商品链接", item.source_url],
            ["来源商品ID", str(top_product.get("product_id") or "")],
            ["来源商品标题", top_product.get("title", "")],
            ["商品当前价", source.get("current_price", "")],
            ["采集商品链接", top_product.get("item_url", "")],
        ]
        if isinstance(item, DirectReplaceBatchItem):
            source_overview.extend(
                [
                    ["来源工作表", item.sheet_name],
                    ["我方商品图", str(item.product_image or "")],
                ]
            )
    else:
        source_overview = [
            ["1688商品链接", item.source_url],
            ["1688拿货价", item.purchase_price],
            ["运费", item.shipping_fee],
            ["淘宝商品ID", str(top_product.get("product_id") or "")],
            ["淘宝商品标题", top_product.get("title", "")],
            ["淘宝商品链接", top_product.get("item_url", "")],
        ]
    payload = {
        "output_path": str(output_path),
        "overview": [
            ["字段", "值"],
            ["source_row", item.row_number],
            ["source_title", item.title],
            *source_overview,
            ["销量", top_product.get("sales_text", "")],
            ["main_image_count", len(source_main)],
            ["main_video_url", video_url or ""],
            ["main_video_status", source.get("main_video_status", "not_found")],
            ["main_image_exported", main_completed],
            ["detail_image_exported", detail_completed],
            ["sku_exported", sku_completed],
            ["parameter_exported", len(parameters)],
            ["title_exported", 1 if (titles or {}).get("long_title") else 0],
            ["export_generated_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ],
        "main": main_rows,
        "detail": detail_rows,
        "sku": sku_rows,
        "parameters": parameters,
        "title": titles or {},
        "videos": ([{"name": "商品主视频", "url": video_url}] if video_url else []),
    }
    return export_workbook_payload(output_path, payload)


class BatchRunner:
    def __init__(
        self,
        settings: ApiSettings,
        project_root: Path,
        profile_dir: Path,
        browser_executable: str = "",
        max_main_images: int | None = DEFAULT_MAIN_IMAGES,
        callback: BatchCallback | None = None,
        oss_uploader: OssUploader | None = None,
        batch_mode: str = "image_search",
        collect_only: bool = False,
        shared_library: SharedLibraryClient | None = None,
        shared_cache: SharedLibraryCache | None = None,
    ) -> None:
        if batch_mode not in {"image_search", "direct_link", "direct_replace"}:
            raise ValueError(f"Unknown batch mode: {batch_mode}")
        self.settings = settings
        self.project_root = project_root
        self.profile_dir = profile_dir
        self.browser_executable = browser_executable.strip()
        self.max_main_images = max_main_images
        self.callback = callback
        self.oss_uploader = oss_uploader
        self.batch_mode = batch_mode
        self.collect_only = collect_only
        self.shared_library = shared_library
        self.shared_cache = shared_cache
        self.shared_identity_resolver = ProductIdentityResolver(resolve_direct_item_url)
        self.cancel_event = threading.Event()
        self.process: subprocess.Popen[str] | None = None
        self.workflow_runner: WorkflowRunner | None = None
        self.workflow_runners: list[WorkflowRunner] = []
        self._supplement_runner_lock = threading.Lock()
        self.direct_collector = DirectLinkCollector(
            project_root,
            profile_dir,
            browser_executable,
            None,
            callback,
        )

    def _enrich_direct_link_skus(
        self,
        item: DirectLinkBatchItem | DirectReplaceBatchItem,
        item_root: Path,
    ) -> DirectLinkBatchItem | DirectReplaceBatchItem:
        if not isinstance(item, DirectLinkBatchItem) or not item.sku_screenshot or item.manual_skus:
            return item
        if self.settings is None:
            return replace(item, sku_screenshot_error="视觉模型未配置")
        try:
            self._emit(
                stage="analyzing_sku_screenshot",
                sequence=item.sequence,
                row=item.row_number,
                message="正在分析 SKU 截图并裁剪可用参考图",
            )
            analysis = VisionClient(self.settings).analyze_sku_screenshot(item.sku_screenshot)
            variants = materialize_sku_screenshot_references(
                item.sku_screenshot,
                analysis,
                item_root / "source-images" / "sku-screenshot-references",
            )
            self._emit(
                stage="sku_screenshot_ready",
                sequence=item.sequence,
                row=item.row_number,
                message=f"SKU 截图解析完成，共识别 {len(variants)} 条",
            )
            return replace(item, manual_skus=tuple(variants), sku_screenshot_error="")
        except Exception as error:
            message = f"SKU 截图解析失败：{type(error).__name__}: {error}"
            self._emit(
                stage="sku_screenshot_failed",
                sequence=item.sequence,
                row=item.row_number,
                message=message,
            )
            return replace(item, sku_screenshot_error=message)

    def _emit(self, **event: Any) -> None:
        if self.callback:
            self.callback(event)

    def cancel(self) -> None:
        self.cancel_event.set()
        if self.workflow_runner:
            self.workflow_runner.cancel()
        for runner in list(self.workflow_runners):
            runner.cancel()
        self.direct_collector.cancel()
        process = self.process
        if process and process.poll() is None:
            subprocess.run(["taskkill", "/PID", str(process.pid), "/F"], check=False, capture_output=True)

    def _collect(self, item: BatchItem, item_root: Path) -> Path:
        collector_executable = self.project_root / "same_item_collector.exe"
        collector_script = self.project_root / "same_item_collector.py"
        command = [str(collector_executable)] if collector_executable.is_file() else [sys.executable, str(collector_script)]
        command.extend([
            "--reference-image",
            str(item.product_image),
            "--output-dir",
            str(item_root / "collected"),
            "--max-count",
            "1",
            "--collect-store-insight-assets",
            "--store-insight-asset-types",
            "main,sku,detail",
            "--collect-main-video",
            "--top-product-only",
            "--browser-profile-dir",
            str(self.profile_dir),
        ])
        if self.browser_executable:
            command.extend(["--browser-executable", self.browser_executable])
        if self.max_main_images is not None:
            command.extend(["--main-image-asset-count", str(self.max_main_images)])
        self.process = subprocess.Popen(
            command,
            cwd=str(self.project_root),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        assert self.process.stdout is not None
        for line in self.process.stdout:
            message = line.strip()
            if message:
                self._emit(stage="collecting", sequence=item.sequence, message=message)
        exit_code = self.process.wait()
        self.process = None
        if self.cancel_event.is_set():
            raise RuntimeError("批处理已停止")
        if exit_code != 0:
            paused_manifest = item_root / "collected" / "collection-paused.json"
            legacy_manifest = item_root / "collected" / "main-image-manifest.json"
            for candidate in (paused_manifest, legacy_manifest):
                if not candidate.is_file():
                    continue
                try:
                    payload = json.loads(candidate.read_text(encoding="utf-8"))
                except json.JSONDecodeError:
                    continue
                if payload.get("status") == "paused":
                    raise BatchCollectionPaused(str(payload.get("message") or "采集需要人工处理"))
            raise RuntimeError(f"搜同款采集失败，退出码 {exit_code}")
        manifest = item_root / "collected" / "main-image-manifest.json"
        if not manifest.is_file():
            raise RuntimeError("搜同款采集结束但未生成 manifest")
        return manifest

    def supplement(
        self,
        workbook_path: Path,
        output_root: Path,
        sequence: int,
        category: str,
        count: int,
    ) -> dict[str, Any]:
        if self.batch_mode == "direct_link":
            items = extract_direct_link_items(workbook_path, output_root)
        elif self.batch_mode == "direct_replace":
            items = extract_direct_replace_items(workbook_path, output_root)
        else:
            items = extract_batch_items(workbook_path, output_root)
        item = next((candidate for candidate in items if candidate.sequence == sequence), None)
        if item is None:
            raise ValueError(f"找不到第 {sequence} 个商品")
        if category not in ASSET_TYPES:
            raise ValueError("补充类型必须是 main、sku 或 detail")
        item_root = output_root / f"{item.sequence:03d}-row-{item.row_number:04d}"
        restored = restore_collected_manifest(item_root)
        if not restored:
            raise ValueError("该商品没有可用的采集素材，请先完成采集")
        source_manifest, _ = restored
        generation_mode = "competitor_reference" if self.batch_mode == "direct_link" else "own_product"
        workbook_name = _safe_name(item.title, f"商品-{item.sequence:03d}") + ".xlsx"
        return self._supplement_item(
            item=item,
            item_root=item_root,
            source_manifest=source_manifest,
            generated_root=item_root / "generated",
            generation_mode=generation_mode,
            workbook_path=item_root / workbook_name,
            category=category,
            count=count,
        )

    def supplement_exported_workbook(self, workbook_path: Path, category: str, count: int) -> dict[str, Any]:
        context = resolve_supplement_workbook(workbook_path, self.project_root)
        return self._supplement_item(
            item=context.item,
            item_root=context.item_root,
            source_manifest=context.source_manifest,
            generated_root=context.generated_root,
            generation_mode=context.generation_mode,
            workbook_path=context.workbook_path,
            category=category,
            count=count,
        )

    def supplement_all(
        self,
        workbook_path: Path,
        output_root: Path,
        sequence: int,
    ) -> dict[str, Any]:
        if self.batch_mode == "direct_link":
            items = extract_direct_link_items(workbook_path, output_root)
        elif self.batch_mode == "direct_replace":
            items = extract_direct_replace_items(workbook_path, output_root)
        else:
            items = extract_batch_items(workbook_path, output_root)
        item = next((candidate for candidate in items if candidate.sequence == sequence), None)
        if item is None:
            raise ValueError(f"找不到第 {sequence} 个商品")
        item_root = output_root / f"{item.sequence:03d}-row-{item.row_number:04d}"
        restored = restore_collected_manifest(item_root)
        if not restored:
            raise ValueError("该商品没有可用的采集素材，请先完成采集")
        source_manifest, _ = restored
        workbook_name = _safe_name(item.title, f"商品-{item.sequence:03d}") + ".xlsx"
        return self._supplement_all_item(
            item=item,
            item_root=item_root,
            source_manifest=source_manifest,
            generated_root=item_root / "generated",
            generation_mode="competitor_reference" if self.batch_mode == "direct_link" else "own_product",
            workbook_path=item_root / workbook_name,
        )

    def supplement_all_exported_workbook(self, workbook_path: Path) -> dict[str, Any]:
        context = resolve_supplement_workbook(workbook_path, self.project_root)
        return self._supplement_all_item(
            item=context.item,
            item_root=context.item_root,
            source_manifest=context.source_manifest,
            generated_root=context.generated_root,
            generation_mode=context.generation_mode,
            workbook_path=context.workbook_path,
        )

    def _supplement_all_item(
        self,
        *,
        item: BatchItem | DirectLinkBatchItem | DirectReplaceBatchItem,
        item_root: Path,
        source_manifest: Path,
        generated_root: Path,
        generation_mode: str,
        workbook_path: Path,
    ) -> dict[str, Any]:
        ensure_workbook_available(workbook_path)
        generation_manifest = item_root / "generation-manifest.json"
        build_generation_manifest(source_manifest, generation_manifest)
        if isinstance(item, DirectReplaceBatchItem) and item.manual_skus:
            if item.product_image is None or not item.product_image.is_file():
                raise ValueError("该商品缺少可用的我方商品图，无法补充生成")
            add_manual_sku_generation_sources(generation_manifest, item, item.product_image)
        if isinstance(item, DirectLinkBatchItem) and item.manual_skus:
            fallback_image = resolve_identity_image(generation_manifest, None, generation_mode)
            add_manual_sku_generation_sources(generation_manifest, item, fallback_image)
        analysis_path = generated_root / "analysis.json"
        try:
            analysis_document = json.loads(analysis_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            analysis_document = {"records": []}
        previous_records = [
            record for record in analysis_document.get("records", []) if isinstance(record, dict)
        ]
        planned = plan_all_supplement_ordinals(previous_records, generation_manifest)
        planned = {
            category: ordinals
            for category, ordinals in planned.items()
            if ordinals and _manifest_image_paths(generation_manifest, category)
        }
        if not planned:
            return {
                "sequence": item.sequence,
                "row": item.row_number,
                "status": "completed",
                "supplemented": 0,
                "failed": 0,
                "categories": {},
                "workbook": str(workbook_path),
            }

        identity_image = resolve_identity_image(
            generation_manifest,
            item.product_image if isinstance(item, (BatchItem, DirectReplaceBatchItem)) else None,
            generation_mode,
        )
        detail_plans = None
        if "detail" in planned and generation_mode == "competitor_reference":
            sources = load_identity_sources(generation_manifest)
            _, _, detail_plans = _supplement_detail_plans(
                generated_root,
                planned["detail"],
                sources,
            )

        def run_category(category: str, ordinals: list[int]) -> tuple[str, list[dict[str, Any]]]:
            if self.cancel_event.is_set():
                return category, []
            runner = WorkflowRunner(
                self.settings,
                callback=lambda event: self._emit(sequence=item.sequence, **event),
            )
            with self._supplement_runner_lock:
                self.workflow_runners.append(runner)
            try:
                added = runner.run(
                    generation_manifest,
                    identity_image,
                    generated_root,
                    None,
                    (category,),
                    max_main_images=max(ordinals),
                    generation_mode=generation_mode,
                    identity_image=identity_image,
                    requested_ordinals={category: ordinals},
                    detail_plans=detail_plans if category == "detail" else None,
                    existing_records=previous_records,
                    persist_records=False,
                )
                return category, upload_generation_records(added, self.oss_uploader)
            except Exception as error:
                return category, [
                    {
                        "category": category,
                        "ordinal": ordinal,
                        "status": "failed",
                        "error": f"{type(error).__name__}: {error}",
                    }
                    for ordinal in ordinals
                ]
            finally:
                with self._supplement_runner_lock:
                    if runner in self.workflow_runners:
                        self.workflow_runners.remove(runner)

        added_by_category: dict[str, list[dict[str, Any]]] = {}
        with ThreadPoolExecutor(max_workers=len(planned), thread_name_prefix="supplement") as executor:
            futures = [executor.submit(run_category, category, ordinals) for category, ordinals in planned.items()]
            for future in as_completed(futures):
                category, added = future.result()
                added_by_category[category] = added
        added = [record for records in added_by_category.values() for record in records]
        merged = merge_generation_records(analysis_path, previous_records, added)
        supplemented = sum(record.get("status") == "completed" for record in added)
        failed = sum(record.get("status") == "failed" for record in added)
        if supplemented:
            titles: dict[str, Any] = {}
            titles_path = item_root / "titles.json"
            try:
                loaded_titles = json.loads(titles_path.read_text(encoding="utf-8"))
                if isinstance(loaded_titles, dict):
                    titles = loaded_titles
            except (OSError, json.JSONDecodeError):
                pass
            exported = export_product_workbook(
                workbook_path,
                item,
                source_manifest,
                merged,
                titles,
                self.project_root,
                include_metadata_only_skus=isinstance(item, (DirectLinkBatchItem, DirectReplaceBatchItem)),
            )
        else:
            exported = workbook_path
        return {
            "sequence": item.sequence,
            "row": item.row_number,
            "status": "partial" if failed else "completed",
            "supplemented": supplemented,
            "failed": failed,
            "categories": {
                category: {
                    "requested": len(planned.get(category, [])),
                    "completed": sum(record.get("status") == "completed" for record in records),
                    "failed": sum(record.get("status") == "failed" for record in records),
                }
                for category, records in added_by_category.items()
            },
            "workbook": str(exported),
        }

    def _supplement_item(
        self,
        *,
        item: BatchItem | DirectLinkBatchItem | DirectReplaceBatchItem,
        item_root: Path,
        source_manifest: Path,
        generated_root: Path,
        generation_mode: str,
        workbook_path: Path,
        category: str,
        count: int,
    ) -> dict[str, Any]:
        if category not in ASSET_TYPES:
            raise ValueError("补充类型必须是 main、sku 或 detail")
        ensure_workbook_available(workbook_path)
        generation_manifest = item_root / "generation-manifest.json"
        build_generation_manifest(source_manifest, generation_manifest)
        if isinstance(item, DirectReplaceBatchItem) and item.manual_skus:
            if item.product_image is None or not item.product_image.is_file():
                raise ValueError("该商品缺少可用的我方商品图，无法补充生成")
            add_manual_sku_generation_sources(generation_manifest, item, item.product_image)
        if isinstance(item, DirectLinkBatchItem) and item.manual_skus:
            fallback_image = resolve_identity_image(generation_manifest, None, generation_mode)
            add_manual_sku_generation_sources(generation_manifest, item, fallback_image)
        analysis_path = generated_root / "analysis.json"
        try:
            analysis_document = json.loads(analysis_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            analysis_document = {"records": []}
        previous_records = [
            record for record in analysis_document.get("records", []) if isinstance(record, dict)
        ]
        ordinals = plan_supplement_ordinals(previous_records, category, count)
        if not _manifest_image_paths(generation_manifest, category):
            raise ValueError(f"该商品没有可用的{category}采集素材，无法补充生成")

        identity_image = resolve_identity_image(
            generation_manifest,
            item.product_image if isinstance(item, (BatchItem, DirectReplaceBatchItem)) else None,
            generation_mode,
        )
        detail_plans = None
        if category == "detail" and generation_mode == "competitor_reference":
            sources = load_identity_sources(generation_manifest)
            _, _, detail_plans = _supplement_detail_plans(
                generated_root,
                ordinals,
                sources,
            )
        runner = WorkflowRunner(
            self.settings,
            callback=lambda event: self._emit(sequence=item.sequence, **event),
        )
        self.workflow_runner = runner
        try:
            added = runner.run(
                generation_manifest,
                identity_image,
                generated_root,
                None,
                (category,),
                max_main_images=max(ordinals),
                generation_mode=generation_mode,
                identity_image=identity_image,
                requested_ordinals={category: ordinals},
                detail_plans=detail_plans,
                existing_records=previous_records,
            )
        finally:
            self.workflow_runner = None
        added = upload_generation_records(added, self.oss_uploader)
        merged = merge_generation_records(analysis_path, previous_records, added)
        titles: dict[str, Any] = {}
        titles_path = item_root / "titles.json"
        try:
            loaded_titles = json.loads(titles_path.read_text(encoding="utf-8"))
            if isinstance(loaded_titles, dict):
                titles = loaded_titles
        except (OSError, json.JSONDecodeError):
            pass
        exported = export_product_workbook(
            workbook_path,
            item,
            source_manifest,
            merged,
            titles,
            self.project_root,
            include_metadata_only_skus=isinstance(item, (DirectLinkBatchItem, DirectReplaceBatchItem)),
        )
        return {
            "sequence": item.sequence,
            "row": item.row_number,
            "status": "completed",
            "category": category,
            "supplemented": len([record for record in added if record.get("status") == "completed"]),
            "workbook": str(exported),
        }

    def _shared_identity_for_item(
        self,
        item: BatchItem | DirectLinkBatchItem | DirectReplaceBatchItem,
    ) -> ProductIdentity | None:
        if (
            self.batch_mode != "direct_link"
            or self.collect_only
            or self.shared_library is None
            or not isinstance(item, DirectLinkBatchItem)
            or item.platform not in {"taobao", "tmall"}
        ):
            return None
        return self.shared_identity_resolver.resolve(item.source_url)

    def _reuse_shared_item(
        self,
        item: DirectLinkBatchItem,
        item_root: Path,
        identity: ProductIdentity,
        catalog: dict[str, Any],
    ) -> dict[str, Any]:
        if self.shared_library is None:
            raise SharedLibraryUnavailable("共享素材库暂时不可用")
        package_object = str(catalog.get("package_object") or "")
        package_sha256 = str(catalog.get("package_sha256") or "")
        reuse_root = self.project_root / "outputs" / "reused" / identity.product_key
        package_zip = self.shared_library.download(
            package_object,
            int(catalog.get("package_size") or 0),
            package_sha256,
            reuse_root / "complete-package.zip",
        )
        reused = materialize_reused_package(
            package_zip,
            reuse_root / "materialized",
        )
        title = item.title or str(reused.titles.get("long_title") or "")
        workbook_name = _safe_name(title, f"商品-{item.sequence:03d}") + ".xlsx"
        exported = export_product_workbook(
            item_root / workbook_name,
            item,
            reused.source_manifest,
            reused.generated_records,
            reused.titles,
            self.project_root,
            include_metadata_only_skus=True,
        )
        if self.shared_cache is not None:
            self.shared_cache.record_download(
                identity.product_key,
                package_object,
                package_sha256,
                reused.root,
            )
        return {
            "sequence": item.sequence,
            "row": item.row_number,
            "status": "completed",
            "workbook": str(exported),
            "generated": sum(
                record.get("status") == "completed"
                for record in reused.generated_records
            ),
            "shared_status": "reused",
        }

    def run(self, workbook_path: Path, output_root: Path) -> list[dict[str, Any]]:
        if self.batch_mode == "direct_link":
            items = extract_direct_link_items(workbook_path, output_root)
        elif self.batch_mode == "direct_replace":
            items = extract_direct_replace_items(workbook_path, output_root)
        else:
            items = extract_batch_items(workbook_path, output_root)
        if not items:
            if self.batch_mode == "direct_link":
                message = "输入 Excel 中没有可识别的商品链接"
            elif self.batch_mode == "direct_replace":
                message = "输入 Excel 中没有可配对的我方商品图和对标链接"
            else:
                message = "输入 Excel 的 E 列没有可用商品图"
            raise ValueError(message)
        checkpoint_options = {
            "batch_mode": self.batch_mode,
            "run_mode": "collect_only" if self.collect_only else "full",
            "total": len(items),
        }
        previous_results = load_batch_results(workbook_path, output_root)
        results_by_sequence = {
            int(item.get("sequence") or 0): dict(item)
            for item in previous_results
            if int(item.get("sequence") or 0) > 0
        }
        completed_sequences = {
            sequence
            for sequence, item in results_by_sequence.items()
            if _batch_result_is_valid(item, output_root)
            or (
                self.collect_only
                and item.get("status") == "collected"
                and _collected_manifest_is_valid(Path(str(item.get("manifest") or "")))
            )
        }
        resume_message = f"，将跳过 {len(completed_sequences)} 个已完成商品" if completed_sequences else ""
        self._emit(stage="ready", total=len(items), message=f"已识别 {len(items)} 个商品{resume_message}")
        for item in items:
            if self.cancel_event.is_set():
                break
            if item.sequence in completed_sequences:
                self._emit(
                    stage="resumed",
                    sequence=item.sequence,
                    total=len(items),
                    row=item.row_number,
                    message="已完成，跳过采集",
                )
                continue
            item_root = output_root / f"{item.sequence:03d}-row-{item.row_number:04d}"
            shared_identity: ProductIdentity | None = None
            shared_lease: LockLease | None = None
            shared_status = ""
            shared_task_id = uuid.uuid4().hex
            try:
                if self.batch_mode in {"direct_link", "direct_replace"} and item.validation_error:
                    raise ValueError(item.validation_error)
                if self.batch_mode == "direct_link":
                    item = self._enrich_direct_link_skus(item, item_root)
                try:
                    shared_identity = self._shared_identity_for_item(item)
                except ProductIdentityError:
                    shared_status = "local_fallback"
                if shared_identity is not None and self.shared_library is not None:
                    try:
                        probe = self.shared_library.probe(shared_identity)
                        if probe.status == "available":
                            try:
                                result = self._reuse_shared_item(
                                    item,
                                    item_root,
                                    shared_identity,
                                    probe.catalog or {},
                                )
                            except Exception:
                                shared_status = "local_fallback"
                            else:
                                results_by_sequence[item.sequence] = result
                                save_batch_results(
                                    workbook_path,
                                    output_root,
                                    list(results_by_sequence.values()),
                                    **checkpoint_options,
                                )
                                self._emit(
                                    stage="completed",
                                    total=len(items),
                                    **result,
                                    message="已复用共享素材并完成导出",
                                )
                                continue
                        elif probe.status == "locked":
                            shared_status = "locked"
                            raise SharedLibraryLockBusy(probe.lock)
                        elif probe.status == "missing":
                            try:
                                shared_lease = self.shared_library.acquire_lock(
                                    shared_identity,
                                    shared_task_id,
                                )
                            except SharedLibraryLockBusy:
                                shared_status = "locked"
                                raise
                        else:
                            shared_status = "local_fallback"
                    except SharedLibraryLockBusy:
                        raise
                    except SharedLibraryUnavailable:
                        shared_status = "local_fallback"
                recovered = restore_collected_manifest(item_root)
                reuse_message = "检测到已采集素材"
                if not recovered and self.batch_mode in {"direct_link", "direct_replace"}:
                    historical = find_prior_direct_collected_manifest(
                        item.source_url,
                        (
                            output_root.parent,
                            self.project_root / "outputs" / "store-insight",
                        ),
                        item_root,
                    )
                    if historical:
                        historical_manifest, asset_count = historical
                        reused_manifest = normalize_direct_manifest(
                            historical_manifest,
                            item_root / "collected" / "direct-manifest.json",
                        )
                        recovered = reused_manifest, asset_count
                        reuse_message = "历史链接采集素材"
                if not recovered and self.batch_mode == "image_search":
                    recovered = find_prior_collected_manifest(item.product_image, item_root, output_root)
                    reuse_message = "检测到同一商品图的历史素材"
                if recovered:
                    source_manifest, asset_count = recovered
                    self._emit(
                        stage="collection_reused",
                        sequence=item.sequence,
                        total=len(items),
                        row=item.row_number,
                        message=f"{reuse_message} {asset_count} 张，跳过重复采集",
                    )
                else:
                    if self.batch_mode in {"direct_link", "direct_replace"}:
                        self._emit(stage="collecting", sequence=item.sequence, total=len(items), row=item.row_number, message="开始直接采集指定商品")
                        raw_manifest = self.direct_collector.collect(item, item_root)
                        source_manifest = normalize_direct_manifest(
                            raw_manifest,
                            item_root / "collected" / "direct-manifest.json",
                        )
                    else:
                        self._emit(stage="collecting", sequence=item.sequence, total=len(items), row=item.row_number, message="开始淘宝搜同款")
                        source_manifest = self._collect(item, item_root)
                if self.collect_only:
                    if not _collected_manifest_is_valid(source_manifest):
                        raise RuntimeError("采集结果为空且未记录明确的缺失素材类型，不能标记为采集完成。")
                    result = {
                        "sequence": item.sequence,
                        "row": item.row_number,
                        "status": "collected",
                        "manifest": str(source_manifest),
                    }
                    results_by_sequence[item.sequence] = result
                    save_batch_results(workbook_path, output_root, list(results_by_sequence.values()), **checkpoint_options)
                    self._emit(stage="collected", total=len(items), **result, message="商品素材采集完成")
                    continue
                source_document = json.loads(source_manifest.read_text(encoding="utf-8"))
                if isinstance(item, (DirectLinkBatchItem, DirectReplaceBatchItem)) and item.manual_skus:
                    source_document = merge_manual_sku_metadata(source_document, item)
                if isinstance(item, DirectLinkBatchItem) and item.sku_screenshot_error:
                    source_document["sku_metadata_status"] = "screenshot_error"
                    source_document["sku_metadata_error"] = item.sku_screenshot_error
                source_document = upload_video_if_needed(
                    source_document,
                    self.oss_uploader,
                    item_root.name,
                )
                source_manifest.write_text(
                    json.dumps(source_document, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                top_product = (source_document.get("products") or [{}])[0]
                generation_manifest = build_generation_manifest(source_manifest, item_root / "generation-manifest.json")
                if isinstance(item, DirectReplaceBatchItem) and item.manual_skus:
                    if item.product_image is None or not item.product_image.is_file():
                        raise RuntimeError("当前商品缺少有效的我方商品图，无法建立商品身份")
                generation_document = json.loads(generation_manifest.read_text(encoding="utf-8"))
                analysis_path = item_root / "generated" / "analysis.json"
                try:
                    analysis_document = json.loads(analysis_path.read_text(encoding="utf-8"))
                except (OSError, json.JSONDecodeError):
                    analysis_document = {"records": []}
                previous_generation_records = [
                    record
                    for record in analysis_document.get("records", [])
                    if isinstance(record, dict)
                ]
                main_image = next(
                    (
                        Path(str(image.get("path"))).resolve()
                        for image in generation_document.get("images", [])
                        if image.get("type") == "main" and Path(str(image.get("path"))).is_file()
                    ),
                    None,
                )
                if main_image is None:
                    raise RuntimeError("指定商品未采集到有效主图，无法建立商品身份")
                product_image = main_image if self.batch_mode == "direct_link" else item.product_image
                if product_image is None or not product_image.is_file():
                    raise RuntimeError("当前商品缺少有效的我方商品图，无法建立商品身份")
                if isinstance(item, (DirectLinkBatchItem, DirectReplaceBatchItem)) and item.manual_skus:
                    add_manual_sku_generation_sources(generation_manifest, item, product_image)
                    generation_document = json.loads(generation_manifest.read_text(encoding="utf-8"))
                available_types = tuple(
                    asset_type
                    for asset_type in ASSET_TYPES
                    if any(image.get("type") == asset_type for image in generation_document.get("images", []))
                )
                retry_ordinals = (
                    plan_batch_retry_ordinals(
                        generation_manifest,
                        available_types,
                        previous_generation_records,
                        self.max_main_images,
                    )
                    if previous_generation_records
                    else None
                )
                self._emit(
                    stage="generating_titles",
                    sequence=item.sequence,
                    total=len(items),
                    row=item.row_number,
                    message="正在生成商品长短标题",
                )
                titles = ProductTitleClient(self.settings).generate(
                    product_image,
                    item.title,
                    str(top_product.get("title") or ""),
                    list(source_document.get("product_parameters") or []),
                )
                (item_root / "titles.json").write_text(
                    json.dumps(titles, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                generation_records = previous_generation_records
                if retry_ordinals is None or retry_ordinals:
                    retry_message = "，仅重试失败或缺失序号" if retry_ordinals is not None else ""
                    self._emit(stage="generating", sequence=item.sequence, total=len(items), row=item.row_number, message=f"开始生成 {', '.join(available_types)} 图片{retry_message}")
                    self.workflow_runner = WorkflowRunner(
                        self.settings,
                        callback=lambda event, sequence=item.sequence: self._emit(sequence=sequence, **event),
                    )
                    added_records = self.workflow_runner.run(
                        generation_manifest,
                        product_image,
                        item_root / "generated",
                        None,
                        available_types,
                        self.max_main_images,
                        generation_mode="competitor_reference" if self.batch_mode == "direct_link" else "own_product",
                        identity_image=product_image,
                        requested_ordinals=retry_ordinals,
                        existing_records=previous_generation_records,
                    )
                    added_records = upload_generation_records(added_records, self.oss_uploader)
                    generation_records = (
                        merge_generation_records(analysis_path, previous_generation_records, added_records)
                        if previous_generation_records
                        else added_records
                    )
                else:
                    self._emit(stage="generation_reused", sequence=item.sequence, total=len(items), row=item.row_number, message="已生成图片均有效，跳过重复生图")
                for record in generation_records:
                    if record.get("oss_upload_error"):
                        self._emit(
                            stage="upload_warning",
                            sequence=item.sequence,
                            total=len(items),
                            row=item.row_number,
                            message=str(record["oss_upload_error"]),
                        )
                self.workflow_runner = None
                if (
                    isinstance(item, (DirectLinkBatchItem, DirectReplaceBatchItem))
                    and item.platform == "kuaishou"
                ):
                    source_document = ensure_kuaishou_product_parameters(
                        source_document,
                        item_root / "generated" / "product-dossier.json",
                    )
                    source_manifest.write_text(
                        json.dumps(source_document, ensure_ascii=False, indent=2),
                        encoding="utf-8",
                    )
                generated_count, failed_count, missing_categories = summarize_generation_result(
                    generation_records,
                    available_types,
                )
                if generated_count == 0:
                    raise RuntimeError(
                        f"没有生成任何有效图片，本次 {len(generation_records)} 个生成任务均未成功，已停止表格导出。"
                    )
                workbook_name = _safe_name(item.title, f"商品-{item.sequence:03d}") + ".xlsx"
                exported = export_product_workbook(
                    item_root / workbook_name,
                    item,
                    source_manifest,
                    generation_records,
                    titles,
                    self.project_root,
                    include_metadata_only_skus=self.batch_mode in {"direct_link", "direct_replace"},
                )
                if (
                    shared_identity is not None
                    and shared_lease is not None
                    and self.shared_library is not None
                ):
                    try:
                        package = SharedPackageBuilder(
                            item_root / "shared-package",
                            self.shared_library.root_prefix,
                            self.shared_library.client_id,
                        ).build(
                            identity=shared_identity,
                            task_id=shared_task_id,
                            source_manifest=source_manifest,
                            generated_records=generation_records,
                            titles=titles,
                            workbook_path=exported,
                            generation_mode="competitor_reference",
                            workflows=available_types,
                            max_main_images=self.max_main_images,
                            max_sku_images=None,
                            max_detail_images=None,
                        )
                        if package is not None:
                            self.shared_library.publish(
                                package.to_publish_bundle(),
                                shared_lease,
                            )
                            shared_status = "published"
                    except Exception:
                        shared_status = "local_fallback"
                video_incomplete = bool(source_document.get("main_video_local_path")) and str(
                    source_document.get("main_video_status") or ""
                ) != "complete"
                partial = failed_count > 0 or bool(missing_categories) or video_incomplete
                result = {
                    "sequence": item.sequence,
                    "row": item.row_number,
                    "status": "failed" if partial else "completed",
                    "workbook": str(exported),
                    "generated": generated_count,
                }
                if shared_status:
                    result["shared_status"] = shared_status
                if partial:
                    errors: list[str] = []
                    if failed_count > 0 or missing_categories:
                        missing_text = f"，缺少类型：{', '.join(missing_categories)}" if missing_categories else ""
                        errors.append(f"仅生成成功 {generated_count} 张，失败 {failed_count} 张{missing_text}")
                    if video_incomplete:
                        video_error = str(source_document.get("main_video_error") or "未生成可访问的公网视频 URL")
                        errors.append(f"视频公网地址处理失败：{video_error}")
                    result["error"] = "；".join(errors)
                results_by_sequence[item.sequence] = result
                save_batch_results(workbook_path, output_root, list(results_by_sequence.values()), **checkpoint_options)
                self._emit(
                    stage="failed" if partial else "completed",
                    total=len(items),
                    **result,
                    message=(str(result.get("error")) if partial else "商品任务完成"),
                )
            except BatchCollectionPaused as error:
                result = {
                    "sequence": item.sequence,
                    "row": item.row_number,
                    "status": "stopped",
                    "error": str(error),
                }
                if shared_status:
                    result["shared_status"] = shared_status
                results_by_sequence[item.sequence] = result
                save_batch_results(workbook_path, output_root, list(results_by_sequence.values()), **checkpoint_options)
                self._emit(stage="stopped", total=len(items), **result, message=str(error))
                break
            except Exception as error:
                stopped = self.cancel_event.is_set()
                result = {
                    "sequence": item.sequence,
                    "row": item.row_number,
                    "status": "stopped" if stopped else "failed",
                    "error": str(error),
                }
                if shared_status:
                    result["shared_status"] = shared_status
                results_by_sequence[item.sequence] = result
                save_batch_results(workbook_path, output_root, list(results_by_sequence.values()), **checkpoint_options)
                self._emit(stage="stopped" if stopped else "failed", total=len(items), **result, message=str(error))
                if self.cancel_event.is_set():
                    break
            finally:
                if shared_lease is not None and self.shared_library is not None:
                    try:
                        self.shared_library.release_lock(shared_lease)
                    except Exception:
                        self._emit(
                            stage="shared_warning",
                            sequence=item.sequence,
                            total=len(items),
                            row=item.row_number,
                            message="共享素材任务锁释放失败，将由 OSS 自动过期",
                        )
        save_batch_results(workbook_path, output_root, list(results_by_sequence.values()), **checkpoint_options)
        results = sorted(results_by_sequence.values(), key=lambda item: int(item.get("sequence") or 0))
        return results
