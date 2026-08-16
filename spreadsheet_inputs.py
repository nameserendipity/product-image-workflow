from __future__ import annotations

import re
import unicodedata
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook


DISPIMG_PATTERN = re.compile(r'DISPIMG\("([^"]+)"', re.IGNORECASE)
RELATIONSHIP_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/relationships"
DRAWING_NAMESPACE = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
OFFICE_REL_NAMESPACE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def normalize_header(value: Any) -> str:
    normalized = unicodedata.normalize("NFKC", str(value or ""))
    return re.sub(r"\s+", "", normalized).lower()


def _wps_cell_image_targets(workbook_path: Path) -> dict[str, tuple[str, bytes]]:
    with zipfile.ZipFile(workbook_path) as archive:
        names = set(archive.namelist())
        image_document = "xl/cellimages.xml"
        relationship_document = "xl/_rels/cellimages.xml.rels"
        if image_document not in names or relationship_document not in names:
            return {}

        relationships = ElementTree.fromstring(archive.read(relationship_document))
        target_by_id = {
            str(node.get("Id") or ""): str(node.get("Target") or "")
            for node in relationships.findall(f"{{{RELATIONSHIP_NAMESPACE}}}Relationship")
        }
        images = ElementTree.fromstring(archive.read(image_document))
        mapped: dict[str, tuple[str, bytes]] = {}
        for cell_image in images:
            name_node = cell_image.find(f".//{{{DRAWING_NAMESPACE}}}cNvPr")
            blip = cell_image.find(".//{http://schemas.openxmlformats.org/drawingml/2006/main}blip")
            if name_node is None or blip is None:
                continue
            image_id = str(name_node.get("name") or "")
            relationship_id = str(blip.get(f"{{{OFFICE_REL_NAMESPACE}}}embed") or "")
            target = target_by_id.get(relationship_id, "")
            if not image_id or not target:
                continue
            archive_name = str(Path("xl") / target).replace("\\", "/")
            if archive_name not in names:
                continue
            mapped[image_id] = (Path(target).suffix or ".png", archive.read(archive_name))
        return mapped


def extract_embedded_images(
    workbook_path: Path,
    output_dir: Path,
) -> dict[tuple[str, int, int], list[Path]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    images: dict[tuple[str, int, int], list[Path]] = defaultdict(list)
    workbook = load_workbook(workbook_path, data_only=False)
    try:
        counter = 0
        for sheet in workbook.worksheets:
            for image in sheet._images:
                marker = getattr(getattr(image, "anchor", None), "_from", None)
                if marker is None:
                    continue
                counter += 1
                suffix = "." + str(getattr(image, "format", "png") or "png").lower().lstrip(".")
                target = output_dir / f"{counter:04d}-{sheet.title}-r{marker.row + 1}-c{marker.col + 1}{suffix}"
                target.write_bytes(image._data())
                images[(sheet.title, marker.row + 1, marker.col + 1)].append(target.resolve())

        wps_images = _wps_cell_image_targets(workbook_path)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    match = DISPIMG_PATTERN.search(str(cell.value or ""))
                    if not match or match.group(1) not in wps_images:
                        continue
                    counter += 1
                    suffix, content = wps_images[match.group(1)]
                    target = output_dir / f"{counter:04d}-{sheet.title}-r{cell.row}-c{cell.column}{suffix.lower()}"
                    target.write_bytes(content)
                    images[(sheet.title, cell.row, cell.column)].append(target.resolve())
        return dict(images)
    finally:
        workbook.close()
