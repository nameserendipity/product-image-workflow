from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import os
import re
import shutil
import socket
import subprocess
import sys
import time
import urllib.request
import zipfile
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path, PurePosixPath
from typing import Any
from urllib.parse import parse_qs, quote, unquote, urlencode, urlparse, urlunparse

from PIL import Image, ImageOps
from openpyxl import load_workbook
from playwright.sync_api import (
    Browser,
    BrowserContext,
    Page,
    Playwright,
    sync_playwright,
)
from platform_urls import is_taobao_or_tmall_host

TAOBAO_HOME = "https://www.taobao.com/"
DEFAULT_CDP_URL = "http://127.0.0.1:9223"
DEFAULT_BROWSER_PROFILE_DIR = Path.cwd() / ".taobao-main-image-profile"
COLLECTION_PAUSED_FILENAME = "collection-paused.json"
COLLECTION_FAILED_FILENAME = "collection-failed.json"
HIGH_RISK_CLICK_WAIT_MS = 1_500
NAVIGATION_SETTLE_WAIT_MS = 3_000
STORE_INSIGHT_DOWNLOAD_TIMEOUT_MS = 5_000
SUPPORTED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
WINDOWS_RESERVED_FILE_STEMS = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{index}" for index in range(1, 10)),
    *(f"LPT{index}" for index in range(1, 10)),
}
DEFAULT_STORE_INSIGHT_ASSET_TYPES = ("main", "sku", "detail")
DYNAMIC_ASSET_RULES = {
    "main": {"minimum": 6, "target": 10, "maximum": 10},
    "sku": {"minimum": 3, "maximum": 8},
    "detail": {"minimum": 6, "maximum": 15},
}
MAX_ASSET_PRODUCT_PROBES = 5
MAX_VIDEO_PRODUCT_PROBES = 3
STORE_INSIGHT_ASSET_TYPE_ALIASES = {
    "main": "main",
    "main_image": "main",
    "main-images": "main",
    "sku": "sku",
    "sku_image": "sku",
    "sku-images": "sku",
    "detail": "detail",
    "details": "detail",
    "detail_image": "detail",
    "detail-images": "detail",
    "description": "detail_long",
    "detail_long": "detail_long",
    "long": "detail_long",
    "long_detail": "detail_long",
}
STORE_INSIGHT_ASSET_DIRS = {
    "main": "main",
    "sku": "sku",
    "detail": "details",
    "detail_long": "description",
}
STORE_INSIGHT_ARCHIVE_DIR_ALIASES = {
    "main": {"1比1主图", "主图", "主图图片"},
    "display_main": {"页面展示主图", "页面图"},
    "sku": {"sku图", "sku图片"},
    "detail": {"详情图", "详情图片"},
    "detail_long": {"详情长图", "详情页长图", "长图"},
}
STORE_INSIGHT_QUICK_DOWNLOAD_LABELS = {
    "main": "1:1图",
    "sku": "仅SKU图",
    "detail": "仅详情图",
    "detail_long": "+长图",
    "main_video": "仅主图视频",
}
STORE_INSIGHT_SKU_FILENAME_PATTERN = re.compile(
    r"^SKU(?:图|image)?[\s_-]?(?P<index>\d+)[\s_-]?(?P<label>.*)$",
    re.IGNORECASE,
)
STORE_INSIGHT_SKU_BRACKET_PATTERN = re.compile(r"【([^】]+)】|\[([^\]]+)\]")
STORE_INSIGHT_SKU_COUNT_PATTERN = re.compile(
    r"[0-9一二两三四五六七八九十百]+(?:个装|件|个|杯|套|组|瓶|盒|只|支|对|包|罐|袋|装)"
)
STORE_INSIGHT_SKU_COLOR_TOKEN_PATTERN = re.compile(
    r"^(?:(?:浅|深|亮|暗|纯|高级|清新|奶油|象牙|珍珠|香槟|玫瑰|雾霾|天空|宝石|藏|墨)?"
    r"[白黑灰红橙黄绿青蓝紫粉棕米杏金银](?:色)?|卡其色|咖啡色|驼色)$"
)
STORE_INSIGHT_SKU_NON_COLOR_MARKER_PATTERN = re.compile(
    r"款|型|版|内胆|不锈钢|加厚|赠品|大号|中号|小号"
)
STORE_INSIGHT_SKU_ID_IN_FILENAME_PATTERN = re.compile(r"(?<!\d)(\d{8,})(?!\d)")
STORE_INSIGHT_SKU_SPEC_PAIR_PATTERN = re.compile(
    r"(?:尺码|规格|型号|大小|容量|净含量|size)\s*[:：=]\s*([^;；,，|/]+)",
    re.IGNORECASE,
)
STORE_INSIGHT_SKU_COLOR_PAIR_PATTERN = re.compile(
    r"(?:颜色分类|颜色|色号|color)\s*[:：=]\s*([^;；,，|/]+)",
    re.IGNORECASE,
)
STORE_INSIGHT_SKU_SIZE_TOKEN_PATTERN = re.compile(
    r"(?<![A-Za-z0-9])(?:XXXS|XXS|XS|S|M|L|XL|XXL|XXXL|XXXXL|[2-8]XL)(?![A-Za-z0-9])|"
    r"(?<!\d)\d{2,3}(?:\.\d+)?(?:码|号|cm|mm|英寸|寸)?(?!\d)|"
    r"(?<!\d)\d+(?:\.\d+)?(?:kg|g|ml|l|斤|件装|个装|套装|盒装|瓶装)(?![A-Za-z])",
    re.IGNORECASE,
)
STORE_INSIGHT_SKU_TABLE_HEADERS = {
    "sku_id": {"skuid", "sku\u7f16\u53f7"},
    "product_id": {"\u5546\u54c1id", "\u5546\u54c1\u7f16\u53f7"},
    "sku_label": {"sku\u4fe1\u606f", "sku\u540d\u79f0", "\u89c4\u683c\u4fe1\u606f"},
    "list_price": {"\u4ef7\u683c", "sku\u4ef7\u683c"},
    "after_coupon_price": {"\u5238\u540e\u4ef7\u683c", "\u4f18\u60e0\u540e\u4ef7\u683c"},
    "net_content": {"\u51c0\u542b\u91cf", "\u89c4\u683c"},
    "stock": {"\u5e93\u5b58", "\u5e93\u5b58\u6570\u91cf"},
}
STOP_MARKERS = (
    "符合镜像",
    "图形验证",
    "\u9a8c\u8bc1\u7801",
    "\u6ed1\u5757",
    "\u5b89\u5168\u9a8c\u8bc1",
    "\u5feb\u901f\u8eab\u4efd\u9a8c\u8bc1",
    "\u8eab\u4efd\u9a8c\u8bc1",
    "\u6211\u662f\u4eba\u7c7b",
    "\u6211\u662f AI",
    "\u6211\u662fAI",
    "\u8bbf\u95ee\u53d7\u9650",
    "\u8bbf\u95ee\u88ab\u62d2\u7edd",
    "\u8bf7\u6c42\u592a\u9891\u7e41",
    "\u64cd\u4f5c\u8fc7\u4e8e\u9891\u7e41",
    "I am human",
    "I'm human",
    "I am not a robot",
    "punish",
    "sec.taobao.com",
)
VERIFICATION_WAIT_MARKERS = (
    "符合镜像",
    "图形验证",
    "\u9a8c\u8bc1\u7801",
    "\u6ed1\u5757",
    "\u5b89\u5168\u9a8c\u8bc1",
    "\u5feb\u901f\u8eab\u4efd\u9a8c\u8bc1",
    "\u8eab\u4efd\u9a8c\u8bc1",
    "\u6211\u662f\u4eba\u7c7b",
    "\u6211\u662f AI",
    "\u6211\u662fAI",
    "I am human",
    "I'm human",
    "I am not a robot",
)
CAPTCHA_SELECTORS = (
    "#baxia-dialog-content",
    "[class*='baxia']",
    "[id*='nocaptcha' i]",
    "[class*='nocaptcha' i]",
    "[class*='nc-container']",
    "iframe[src*='captcha' i]",
    "iframe[src*='baxia' i]",
    "iframe[src*='nvc' i]",
)
LOGIN_MARKERS = (
    "\u4eb2\uff0c\u8bf7\u767b\u5f55",
    "\u8bf7\u767b\u5f55",
    "\u767b\u5f55/\u6ce8\u518c",
    "\u626b\u7801\u767b\u5f55",
    "\u5bc6\u7801\u767b\u5f55",
    "\u8d26\u53f7\u767b\u5f55",
    "\u6dd8\u5b9d\u8d26\u53f7\u767b\u5f55",
)
LOGGED_IN_ORDER_MARKERS = (
    "\u5f85\u6536\u8d27",
    "\u5f85\u4ed8\u6b3e",
    "\u5f85\u8bc4\u4ef7",
    "\u5f85\u53d1\u8d27",
)
CLOSE_POPUP_TEXTS = (
    "\u5173\u95ed",
    "\u5173\u95ed\u5f39\u7a97",
    "\u5173\u95ed\u5e7f\u544a",
    "\u6211\u77e5\u9053\u4e86",
    "\u77e5\u9053\u4e86",
    "\u6682\u4e0d",
    "\u7a0d\u540e\u518d\u8bf4",
    "\u53d6\u6d88",
    "\u00d7",
    "\u2715",
    "\u2716",
)
STORE_INSIGHT_DOWNLOAD_COMPLETION_MARKERS = (
    "\u56fe\u7247\u5df2\u4fdd\u5b58\u81f3\u672c\u5730",
    "\u4e0b\u8f7d\u5b8c\u6210",
)

@dataclass(slots=True)
class CandidateRow:
    source_order: int
    href: str
    raw_text: str
    title_hint: str
    image_url: str


@dataclass(slots=True)
class ProductRecord:
    rank: int
    source_order: int
    product_id: str
    title: str
    sales_text: str
    sales_floor: int | None
    item_url: str
    source_card_url: str
    source_image_url: str
    main_image_local_path: str
    main_image_public_url: str
    sha256: str
    width: int
    height: int
    download_source: str
    status: str
    error: str = ""


@dataclass(slots=True)
class CollectorConfig:
    reference_image: Path | None
    output_dir: Path
    cdp_url: str = DEFAULT_CDP_URL
    max_count: int = 10
    candidate_limit: int = 8
    min_sales_floor: int = 0
    use_current_page: bool = False
    public_url_prefix: str = ""
    store_insight_download: bool = False
    collect_store_insight_assets: bool = False
    store_insight_asset_types: tuple[str, ...] = DEFAULT_STORE_INSIGHT_ASSET_TYPES
    collect_main_image_assets: bool = False
    main_image_asset_count: int = 0
    collect_sku_image_assets: bool = False
    sku_image_asset_count: int = 0
    collect_detail_image_assets: bool = False
    detail_image_asset_count: int = 0
    collect_detail_long_image_assets: bool = False
    detail_long_image_asset_count: int = 0
    asset_count_mode: str = "fixed"
    fixed_asset_counts: dict[str, int] = field(default_factory=dict)
    collect_main_video: bool = False
    head_wait_seconds: float = 2.0
    scroll_rounds: int = 1
    detail_timeout_ms: int = 45_000
    auto_launch_browser: bool = True
    reuse_existing_cdp: bool = True
    browser_executable: str = ""
    browser_profile_dir: Path = DEFAULT_BROWSER_PROFILE_DIR
    login_timeout_seconds: int = 600
    login_poll_seconds: int = 5
    top_product_only: bool = False


@dataclass(frozen=True, slots=True)
class InspectedAssetPackage:
    asset_type: str
    product_id: str
    package_path: Path | None
    extract_dir: Path
    files: tuple[Path, ...]
    error: str = ""

    @property
    def valid_count(self) -> int:
        return len(self.files)


def normalize_store_insight_sku_table_cell(value: Any) -> str:
    text = str(value or "").strip()
    return "" if text in {"-", "--", "\u2014", "\u6682\u65e0"} else text


def normalize_store_insight_sku_table_header(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_store_insight_sku_table_cell(value)).lower()


def sku_table_field_indexes(headers: list[Any]) -> dict[str, int]:
    normalized = [normalize_store_insight_sku_table_header(value) for value in headers]
    indexes: dict[str, int] = {}
    for field, aliases in STORE_INSIGHT_SKU_TABLE_HEADERS.items():
        for index, value in enumerate(normalized):
            if value in aliases:
                indexes[field] = index
                break
    return indexes


def parse_store_insight_sku_table(path: Path) -> list[dict[str, str]]:
    rows: list[list[Any]] = []
    if path.suffix.lower() == ".csv":
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                with path.open("r", encoding=encoding, newline="") as source:
                    rows = list(csv.reader(source))
                break
            except UnicodeDecodeError:
                continue
    else:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            rows = [list(row) for row in workbook.active.iter_rows(values_only=True)]
        finally:
            workbook.close()
    if not rows:
        return []
    indexes = sku_table_field_indexes(rows[0])
    if "sku_label" not in indexes and "sku_id" not in indexes:
        return []
    parsed: list[dict[str, str]] = []
    for row_index, values in enumerate(rows[1:], start=1):
        record = {
            field: normalize_store_insight_sku_table_cell(values[index]) if index < len(values) else ""
            for field, index in indexes.items()
        }
        if not record.get("sku_label") and not record.get("sku_id"):
            continue
        record["source_index"] = str(row_index)
        parsed.append(record)
    return parsed


class CollectorPaused(RuntimeError):
    """Raised when login, verification, or page risk requires manual handling."""


class StoreInsightAssetUnavailable(RuntimeError):
    """Raised when the current product lacks the requested asset quick action."""


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def product_id_from_url(url: str) -> str:
    try:
        parsed = urlparse(url)
    except ValueError:
        return ""
    values = parse_qs(parsed.query)
    for key in ("id", "itemId"):
        value = values.get(key)
        if value and str(value[0]).isdigit():
            return str(value[0])
    match = re.search(r"/item/(\d+)", parsed.path)
    return match.group(1) if match else ""


def is_login_url(url: str) -> bool:
    hostname = (urlparse(url).hostname or "").lower()
    return hostname == "login.taobao.com" or hostname.endswith(".login.taobao.com")


def is_taobao_result_url(url: str) -> bool:
    try:
        parsed = urlparse(url)
    except ValueError:
        return False
    hostname = (parsed.hostname or "").lower()
    lowered = url.lower()
    if hostname == "s.taobao.com" and parsed.path.startswith("/search"):
        return True
    return "imgsearch" in lowered or "xxc=taobaosearch" in lowered


def is_taobao_or_tmall_url(url: str) -> bool:
    try:
        hostname = (urlparse(url).hostname or "").lower()
    except ValueError:
        return False
    return is_taobao_or_tmall_host(hostname)


def page_priority(url: str) -> int:
    hostname = (urlparse(url).hostname or "").lower()
    if is_login_url(url):
        return 9
    if product_id_from_url(url) and is_taobao_or_tmall_host(hostname):
        return 0
    if hostname == "s.taobao.com":
        return 1
    if is_taobao_or_tmall_host(hostname):
        return 2
    return 5


def normalize_image_url(url: str) -> str:
    value = (url or "").strip()
    if value.startswith("//"):
        value = "https:" + value
    if not value:
        return ""
    value = re.sub(
        r"(?i)(\.(?:jpg|jpeg|png|webp))(?:_[^/?#]+)(?=$|[?#])",
        r"\1",
        value,
    )
    return value


SALES_UNIT_PATTERN = r"\u4e07|\u5343|w|W|k|K"
SALES_LABEL_PATTERN = (
    r"\u4eba\u4ed8\u6b3e|\u4ed8\u6b3e\u4eba\u6570|\u4ed8\u6b3e|"
    r"\u4eba\u6536\u8d27|\u6536\u8d27|\u5df2\u552e|\u9500\u91cf|\u4ef6"
)
SALES_NUMBER_FIRST_PATTERN = re.compile(
    rf"(?<![\d.])(?P<num>\d+(?:\.\d+)?)\s*(?P<unit>{SALES_UNIT_PATTERN})?\s*(?P<plus>\+?)\s*"
    rf"(?P<label>{SALES_LABEL_PATTERN})"
)
SALES_LABEL_FIRST_PATTERN = re.compile(
    rf"(?P<label>{SALES_LABEL_PATTERN})\s*[:\uff1a]?\s*"
    rf"(?P<num>\d+(?:\.\d+)?)\s*(?P<unit>{SALES_UNIT_PATTERN})?\s*(?P<plus>\+?)"
)


def _sales_text_variants(text: str) -> tuple[str, ...]:
    spaced = re.sub(r"\s+", " ", text.replace(",", "")).strip()
    compact = re.sub(r"\s+", "", spaced)
    if compact == spaced:
        return (spaced,)
    return (spaced, compact)


def _sales_floor_from_match(match: re.Match[str]) -> int | None:
    number_text = match.group("num")
    unit = match.group("unit")
    if "." in number_text and unit is None:
        return None
    multiplier = {
        None: 1,
        "\u5343": 1_000,
        "k": 1_000,
        "K": 1_000,
        "\u4e07": 10_000,
        "w": 10_000,
        "W": 10_000,
    }[unit]
    return int(float(number_text) * multiplier)


def _find_sales_match(text: str) -> re.Match[str] | None:
    for variant in _sales_text_variants(text):
        for pattern in (SALES_NUMBER_FIRST_PATTERN, SALES_LABEL_FIRST_PATTERN):
            for match in pattern.finditer(variant):
                if _sales_floor_from_match(match) is not None:
                    return match
    return None


def parse_sales_floor(text: str) -> int | None:
    match = _find_sales_match(text)
    return _sales_floor_from_match(match) if match else None


def parse_sales_text(text: str) -> str:
    match = _find_sales_match(text)
    if not match:
        return ""
    return re.sub(r"\s+", "", match.group(0))


def normalize_store_insight_asset_types(value: str | list[str] | tuple[str, ...] | None) -> tuple[str, ...]:
    if value is None:
        raw_items = DEFAULT_STORE_INSIGHT_ASSET_TYPES
    elif isinstance(value, str):
        raw_items = tuple(item.strip() for item in re.split(r"[,;\s]+", value) if item.strip())
    else:
        raw_items = tuple(str(item).strip() for item in value if str(item).strip())

    normalized: list[str] = []
    for item in raw_items:
        key = item.strip().lower().replace(" ", "_")
        asset_type = STORE_INSIGHT_ASSET_TYPE_ALIASES.get(key)
        if asset_type is None:
            raise ValueError(f"unsupported Store Insight asset type: {item}")
        if asset_type not in normalized:
            normalized.append(asset_type)
    return tuple(normalized or DEFAULT_STORE_INSIGHT_ASSET_TYPES)


def store_insight_download_filename(
    product_id: str,
    suggested_filename: str,
    asset_type: str = "",
) -> str:
    def safe_identifier(value: str, fallback: str, limit: int) -> str:
        sanitized = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value)).strip("._-")
        return (sanitized or fallback)[:limit]

    product = safe_identifier(product_id, "product", 20)
    components = [product]
    if asset_type:
        components.append(safe_identifier(asset_type, "asset", 12))
    suffix = Path(suggested_filename or "").suffix.lower()
    if not re.fullmatch(r"\.[a-z0-9]{1,8}", suffix):
        suffix = ".zip"
    return "_".join(components) + suffix


def build_store_insight_asset_targets(config: CollectorConfig) -> dict[str, int | None]:
    if config.asset_count_mode == "dynamic":
        return {}

    if config.fixed_asset_counts:
        return {
            asset_type: max(0, int(count))
            for asset_type, count in config.fixed_asset_counts.items()
            if asset_type in DEFAULT_STORE_INSIGHT_ASSET_TYPES and int(count) > 0
        }

    targets: dict[str, int | None] = {}

    if config.collect_store_insight_assets:
        for asset_type in normalize_store_insight_asset_types(config.store_insight_asset_types):
            targets[asset_type] = None

    requested = (
        ("main", config.collect_main_image_assets, config.main_image_asset_count),
        ("sku", config.collect_sku_image_assets, config.sku_image_asset_count),
        ("detail", config.collect_detail_image_assets, config.detail_image_asset_count),
        ("detail_long", config.collect_detail_long_image_assets, config.detail_long_image_asset_count),
    )
    for asset_type, enabled, count in requested:
        normalized_count = max(0, int(count or 0))
        if enabled or normalized_count > 0:
            targets[asset_type] = normalized_count or None

    return targets


def ranked_products_for_asset_collection(
    config: CollectorConfig,
    ranked_products: list[dict[str, Any]],
    targets: dict[str, int | None],
) -> list[dict[str, Any]]:
    has_quantity_target = any(target is not None for target in targets.values())
    has_bounded_target = any(
        asset_type in targets and targets[asset_type] is None
        for asset_type in ("sku", "detail")
    )
    if config.asset_count_mode == "dynamic" or has_quantity_target or has_bounded_target:
        return ranked_products
    return ranked_products[:1] if config.top_product_only else ranked_products


def remaining_store_insight_asset_targets(
    targets: dict[str, int | None],
    collected: dict[str, int],
    include_unbounded: bool = True,
) -> dict[str, int | None]:
    remaining: dict[str, int | None] = {}
    for asset_type, target in targets.items():
        if target is None:
            if include_unbounded:
                remaining[asset_type] = None
            continue
        missing = max(0, target - collected.get(asset_type, 0))
        if missing > 0:
            remaining[asset_type] = missing
    return remaining


def remaining_bounded_store_insight_asset_targets(
    targets: dict[str, int | None],
    collected: dict[str, int],
    rank: int,
) -> dict[str, int | None]:
    """Use the first product up to each bound, then backfill only shortfalls."""
    remaining = remaining_store_insight_asset_targets(targets, collected, include_unbounded=False)
    for asset_type in ("sku", "detail"):
        if targets.get(asset_type, "__missing__") is not None:
            continue
        rule = DYNAMIC_ASSET_RULES[asset_type]
        count = collected.get(asset_type, 0)
        if rank == 1 or count < int(rule["minimum"]):
            capacity = max(0, int(rule["maximum"]) - count)
            if capacity:
                remaining[asset_type] = capacity
    return remaining


def store_insight_asset_shortfalls(
    targets: dict[str, int | None],
    collected: dict[str, int],
) -> dict[str, int]:
    return {
        asset_type: max(0, target - collected.get(asset_type, 0))
        for asset_type, target in targets.items()
        if target is not None and collected.get(asset_type, 0) < target
    }


def cdp_url_with_port(cdp_url: str, port: int) -> str:
    parsed = urlparse(cdp_url)
    scheme = parsed.scheme or "http"
    host = parsed.hostname or "127.0.0.1"
    return f"{scheme}://{host}:{port}"


def cdp_port(cdp_url: str) -> int:
    return urlparse(cdp_url).port or 9223


def is_port_open(host: str, port: int) -> bool:
    try:
        with socket.create_connection((host, port), timeout=0.5):
            return True
    except OSError:
        return False


def is_cdp_alive(cdp_url: str) -> bool:
    try:
        with urllib.request.urlopen(f"{cdp_url.rstrip('/')}/json/version", timeout=2) as response:
            return response.status == 200
    except Exception:
        return False


def find_browser_executable(custom_path: str = "") -> Path:
    candidates: list[str | Path] = []
    if custom_path:
        candidates.append(custom_path)
    for name in ("msedge.exe", "chrome.exe"):
        resolved = shutil.which(name)
        if resolved:
            candidates.append(resolved)
    candidates.extend(
        [
            r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
            r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
            r"C:\Users\COLORFUL\AppData\Local\Microsoft\Edge\Application\msedge.exe",
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        ]
    )
    for candidate in candidates:
        path = Path(candidate)
        if path.is_file():
            return path
    raise CollectorPaused("No Edge/Chrome executable found. Pass --browser-executable explicitly.")


def close_project_browser_for_profile(profile_dir: Path) -> int:
    if os.name != "nt":
        return 0
    environment = os.environ.copy()
    environment["PRODUCT_IMAGE_PROFILE_DIR"] = str(profile_dir)
    script = """
$profile = [Environment]::GetEnvironmentVariable('PRODUCT_IMAGE_PROFILE_DIR')
Get-CimInstance Win32_Process | Where-Object {
  $_.CommandLine -and
  $_.Name -in @('waxiang.exe', 'msedge.exe', 'chrome.exe') -and
  $_.CommandLine -notmatch '--type=' -and
  $_.CommandLine.IndexOf("--user-data-dir=$profile", [System.StringComparison]::OrdinalIgnoreCase) -ge 0
} | ForEach-Object {
  Stop-Process -Id $_.ProcessId -Force -ErrorAction SilentlyContinue
  Write-Output $_.ProcessId
}
"""
    try:
        result = subprocess.run(
            ["powershell.exe", "-NoProfile", "-NonInteractive", "-Command", script],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            env=environment,
            timeout=10,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
    except (OSError, subprocess.TimeoutExpired):
        return 0
    return sum(line.strip().isdigit() for line in result.stdout.splitlines())


def ensure_cdp_browser(config: CollectorConfig) -> str:
    fixed_cdp_alive = is_cdp_alive(config.cdp_url)
    if fixed_cdp_alive:
        if config.reuse_existing_cdp:
            print(f"Using fixed CDP browser: {config.cdp_url}", flush=True)
            return config.cdp_url
        raise CollectorPaused(
            "Fixed CDP browser is already running but reuse is disabled. "
            "Enable --reuse-existing-cdp before starting collection."
        )
    if not config.auto_launch_browser:
        raise CollectorPaused(
            f"Fixed CDP browser is not available and auto launch is disabled: {config.cdp_url}"
        )

    parsed = urlparse(config.cdp_url)
    host = parsed.hostname or "127.0.0.1"
    if host not in {"127.0.0.1", "localhost"}:
        raise CollectorPaused("Auto browser launch only supports a local CDP URL.")

    port = cdp_port(config.cdp_url)
    if is_port_open(host, port):
        raise CollectorPaused(
            f"Fixed CDP port {port} is occupied but is not a CDP browser. "
            "Close the conflicting process before starting collection."
        )

    browser_path = find_browser_executable(config.browser_executable)
    profile_dir = config.browser_profile_dir.resolve()
    profile_dir.mkdir(parents=True, exist_ok=True)
    closed_count = close_project_browser_for_profile(profile_dir)
    if closed_count:
        print("Closed an existing collection browser; restarting it on the fixed CDP port.", flush=True)
        time.sleep(1)
    print(f"Launching visible browser on fixed CDP port {port}: {profile_dir}", flush=True)
    process = subprocess.Popen(
        [
            str(browser_path),
            f"--remote-debugging-port={port}",
            f"--user-data-dir={profile_dir}",
            "--no-first-run",
            "--no-default-browser-check",
            TAOBAO_HOME,
        ],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    deadline = time.monotonic() + 20
    process_exited = False
    while time.monotonic() < deadline:
        if is_cdp_alive(config.cdp_url):
            config.browser_profile_dir = profile_dir
            print(f"Browser is ready: {config.cdp_url}", flush=True)
            return config.cdp_url
        if process.poll() is not None:
            process_exited = True
        time.sleep(0.5)
    status = "exited" if process_exited else "did not expose CDP"
    raise CollectorPaused(
        f"Browser was launched on fixed CDP port {port}, but it {status} before CDP became ready. "
        "The collector will not try another port."
    )


def current_stop_marker(page: Page) -> str:
    frames = [page, *(getattr(page, "frames", []) or [])]
    for frame in frames:
        try:
            text = frame.locator("body").first.inner_text(timeout=5_000)[:50_000]
        except Exception:
            text = ""
        url = str(getattr(frame, "url", "") or "")
        marker = next((item for item in STOP_MARKERS if item in text or item in url), "")
        if marker:
            return marker
        for selector in CAPTCHA_SELECTORS:
            try:
                locator = frame.locator(selector)
                if locator.count() and locator.first.is_visible(timeout=300):
                    return "图形验证"
            except Exception:
                continue
    return ""


def detect_stop(page: Page, verification_timeout_seconds: int = 600, poll_seconds: int = 5) -> None:
    marker = current_stop_marker(page)
    if not marker:
        return

    if marker not in VERIFICATION_WAIT_MARKERS:
        raise CollectorPaused(f"Page shows risk or verification marker: {marker}")

    print(
        "Taobao verification is required. Waiting up to "
        f"{verification_timeout_seconds} seconds in the current browser tab. "
        "Complete the visible verification, then the collector will continue.",
        flush=True,
    )
    deadline = time.monotonic() + verification_timeout_seconds
    while marker:
        if time.monotonic() >= deadline:
            raise CollectorPaused(f"Verification was not completed within {verification_timeout_seconds} seconds.")
        page.wait_for_timeout(max(1, poll_seconds) * 1000)
        marker = current_stop_marker(page)
        if marker and marker not in VERIFICATION_WAIT_MARKERS:
            raise CollectorPaused(f"Page shows risk or verification marker: {marker}")

    if marker:
        raise CollectorPaused(f"Page shows risk or verification marker: {marker}")
    print("Taobao verification cleared. Continuing collection.", flush=True)


def wait_before_high_risk_click(page: Page) -> None:
    detect_stop(page)
    page.wait_for_timeout(HIGH_RISK_CLICK_WAIT_MS)
    detect_stop(page)


def close_common_popups(page: Page) -> int:
    detect_stop(page)
    try:
        return int(
            page.evaluate(
                """
                ({ closeTexts, loginTexts }) => {
                  const visible = el => {
                    const r = el.getBoundingClientRect();
                    const s = getComputedStyle(el);
                    return r.width > 0 && r.height > 0
                      && s.display !== 'none'
                      && s.visibility !== 'hidden'
                      && s.pointerEvents !== 'none';
                  };
                  const textOf = el => [
                    el?.innerText || '',
                    el?.textContent || '',
                    el?.getAttribute?.('aria-label') || '',
                    el?.getAttribute?.('title') || '',
                    el?.className || ''
                  ].join(' ').replace(/\\s+/g, ' ').trim();
                  const controls = Array.from(document.querySelectorAll(
                    'button,a,[role="button"],[aria-label],[title],i,span,div'
                  )).filter(visible);
                  let closed = 0;
                  for (const el of controls) {
                    if (closed >= 3) break;
                    const text = textOf(el);
                    const rect = el.getBoundingClientRect();
                    const tagName = (el.tagName || '').toLowerCase();
                    const smallControl = rect.width > 0 && rect.height > 0 && rect.width <= 120 && rect.height <= 120;
                    const glyphClose = /^(x|X|\u00d7|\u2715|\u2716)$/.test(text);
                    const explicitTextClose = closeTexts
                      .filter(value => !/^(x|X|\u00d7|\u2715|\u2716)$/.test(value))
                      .some(value => text.includes(value));
                    const classClose = /(^|\\s)(close|btn-close|modal-close|dialog-close|icon-close)(\\s|$)/i.test(text);
                    const shouldClose = explicitTextClose || classClose || (glyphClose && smallControl);
                    if (!shouldClose) continue;
                    if (!smallControl && !/^(button|a|i|span)$/i.test(tagName)) continue;
                    const href = el.href || el.getAttribute?.('href') || '';
                    if (tagName === 'a' && href && !/^javascript:/i.test(href)) continue;
                    const panel = el.closest('[role="dialog"],.modal,.dialog,[class*="popup"],[class*="modal"]')
                      || el.parentElement;
                    const panelText = textOf(panel || el);
                    if (loginTexts.some(value => panelText.includes(value))) continue;
                    if (panelText.includes('\u6309\u56fe\u7247\u641c\u7d22')
                      || panelText.includes('\u56fe\u7247\u641c\u7d22')
                      || panelText.includes('\u641c\u540c\u6b3e')) continue;
                    try {
                      el.click();
                      closed += 1;
                    } catch (_) {}
                  }
                  return closed;
                }
                """,
                {"closeTexts": list(CLOSE_POPUP_TEXTS), "loginTexts": list(LOGIN_MARKERS)},
            )
        )
    except Exception:
        return 0


def store_insight_download_completion_visible(page: Page) -> bool:
    try:
        return bool(
            page.evaluate(
                """
                markers => {
                  const visible = el => {
                    const rect = el.getBoundingClientRect();
                    const style = getComputedStyle(el);
                    return rect.width > 0 && rect.height > 0
                      && style.display !== 'none'
                      && style.visibility !== 'hidden'
                      && style.pointerEvents !== 'none';
                  };
                  const textOf = el => [
                    el?.innerText || '',
                    el?.textContent || '',
                    el?.getAttribute?.('aria-label') || '',
                    el?.getAttribute?.('title') || ''
                  ].join(' ').replace(/\\s+/g, ' ').trim();
                  return Array.from(document.querySelectorAll('body *')).some(
                    el => visible(el) && markers.some(marker => textOf(el).includes(marker))
                  );
                }
                """,
                list(STORE_INSIGHT_DOWNLOAD_COMPLETION_MARKERS),
            )
        )
    except Exception:
        return False


def dismiss_store_insight_download_completion(page: Page, timeout_ms: int = 5_000) -> bool:
    if not store_insight_download_completion_visible(page):
        return True

    deadline = time.monotonic() + max(0, timeout_ms) / 1000
    while True:
        close_common_popups(page)
        if not store_insight_download_completion_visible(page):
            return True
        if time.monotonic() >= deadline:
            return False
        page.wait_for_timeout(min(250, max(1, int((deadline - time.monotonic()) * 1000))))


def has_logged_in_marker(page: Page) -> bool:
    if is_login_url(page.url):
        return False
    try:
        text = page.locator("body").first.inner_text(timeout=5_000)[:50_000]
    except Exception:
        return False
    if re.search(r"\btb\d{4,}\b", text, re.IGNORECASE):
        return True
    if "\u6536\u8d27\u5730\u5740" in text and any(marker in text for marker in LOGGED_IN_ORDER_MARKERS):
        return True
    try:
        return bool(
            page.evaluate(
                """
                ({ orderMarkers }) => {
                  const text = document.body?.innerText || '';
                  if (/\\btb\\d{4,}\\b/i.test(text)) return true;
                  if (!text.includes('\u6536\u8d27\u5730\u5740')) return false;
                  return orderMarkers.some(value => text.includes(value));
                }
                """,
                {"orderMarkers": list(LOGGED_IN_ORDER_MARKERS)},
            )
        )
    except Exception:
        return False


def has_login_marker(page: Page) -> bool:
    if is_login_url(page.url):
        return True
    if has_logged_in_marker(page):
        return False
    try:
        text = page.locator("body").first.inner_text(timeout=5_000)[:50_000]
    except Exception:
        return False
    if "\u4eb2\uff0c\u8bf7\u767b\u5f55" in text:
        return True
    try:
        return bool(
            page.evaluate(
                """
                ({ loginMarkers }) => {
                  const visible = el => {
                    const r = el.getBoundingClientRect();
                    const s = getComputedStyle(el);
                    return r.width > 0 && r.height > 0
                      && s.display !== 'none'
                      && s.visibility !== 'hidden';
                  };
                  const textOf = el => [
                    el?.innerText || '',
                    el?.textContent || '',
                    el?.getAttribute?.('aria-label') || '',
                    el?.getAttribute?.('title') || ''
                  ].join(' ').replace(/\\s+/g, ' ').trim();
                  const controls = Array.from(document.querySelectorAll(
                    'a,button,[role="button"],[aria-label],[title]'
                  ));
                  return controls.some(el => {
                    if (!visible(el)) return false;
                    const text = textOf(el);
                    const href = el.href || el.getAttribute?.('href') || '';
                    const loginHref = /login\\.(taobao|tmall)\\.com/i.test(href);
                    return loginMarkers.some(value => text.includes(value)) || loginHref;
                  });
                }
                """,
                {"loginMarkers": list(LOGIN_MARKERS)},
            )
        )
    except Exception:
        return False


def page_selection_score(page: Page) -> tuple[int, int]:
    url = page.url
    hostname = (urlparse(url).hostname or "").lower()
    if is_taobao_result_url(url):
        return (0, 0)
    if product_id_from_url(url) and is_taobao_or_tmall_host(hostname):
        return (1, 0)
    if has_logged_in_marker(page):
        return (2, 0)
    if is_login_url(url):
        return (9, 0)
    if is_taobao_or_tmall_host(hostname):
        return (4, 0)
    return (5, 0)


def select_best_page(pages: list[Page]) -> Page | None:
    open_pages = [page for page in pages if not page.is_closed()]
    if not open_pages:
        return None
    return min(reversed(open_pages), key=page_selection_score)


def connect_over_cdp(playwright: Playwright, cdp_url: str) -> tuple[Browser, BrowserContext, Page]:
    try:
        browser = playwright.chromium.connect_over_cdp(cdp_url, timeout=15_000)
    except Exception as error:
        raise CollectorPaused(
            "The collection browser CDP session is unresponsive. Close the collection browser window and retry; "
            "your Store Insight login profile will be kept."
        ) from error
    if not browser.contexts:
        raise CollectorPaused("CDP browser has no available context. Start Edge with remote debugging first.")
    context = browser.contexts[0]
    page = select_best_page(list(context.pages))
    if page is None:
        raise CollectorPaused("No browser tabs are available in the CDP context.")
    return browser, context, page


def select_active_page(context: BrowserContext, fallback: Page) -> Page:
    pages = [page for page in context.pages if not page.is_closed()]
    if not pages:
        raise CollectorPaused(
            "No browser tabs are available in the current CDP session. "
            "Please reopen the Taobao browser tab and rerun the task."
        )
    return select_best_page(pages) or fallback


def wait_for_login(
    context: BrowserContext,
    page: Page,
    timeout_seconds: int,
    poll_seconds: int,
) -> Page:
    deadline = time.monotonic() + timeout_seconds
    current = page
    first_notice = True
    while True:
        if current.is_closed():
            current = select_active_page(context, current)
        close_common_popups(current)
        detect_stop(current)
        if not has_login_marker(current):
            return current
        if time.monotonic() >= deadline:
            raise CollectorPaused(f"Login was not completed within {timeout_seconds} seconds.")
        if first_notice:
            print(
                "Taobao login is required. Waiting up to "
                f"{timeout_seconds} seconds in the current browser tab. "
                "Do not close this tab.",
                flush=True,
            )
            first_notice = False
        time.sleep(max(1, poll_seconds))


def find_result_page(context: BrowserContext, excluded_page_ids: set[int] | None = None) -> Page | None:
    excluded_page_ids = excluded_page_ids or set()
    pages = [page for page in context.pages if not page.is_closed()]
    for page in reversed(pages):
        if id(page) in excluded_page_ids:
            continue
        if is_taobao_result_url(page.url):
            return page
    return None


def wait_for_result_page(
    context: BrowserContext,
    page: Page,
    timeout_seconds: int,
    poll_seconds: int,
) -> Page:
    deadline = time.monotonic() + timeout_seconds
    current = page
    first_notice = True
    while True:
        result_page = find_result_page(context)
        if result_page is not None:
            detect_stop(result_page)
            return result_page

        if current.is_closed():
            current = select_active_page(context, current)
        close_common_popups(current)
        detect_stop(current)
        if is_taobao_result_url(current.url):
            return current

        if time.monotonic() >= deadline:
            raise CollectorPaused(
                "Taobao same-item/search result page was not opened within "
                f"{timeout_seconds} seconds. Current URL: {current.url}"
            )
        if first_notice:
            print(
                "Open a Taobao same-item/search result page in the browser. "
                "The collector is waiting.",
                flush=True,
            )
            first_notice = False
        time.sleep(max(1, poll_seconds))
        current = select_active_page(context, current)


def click_image_search_submit(page: Page, timeout_ms: int = 15_000) -> None:
    deadline = time.monotonic() + timeout_ms / 1000
    selectors = (
        "#image-search-upload-button",
        "[class*='image-search'] button:has-text('\u641c\u7d22')",
        "[class*='ImageSearch'] button:has-text('\u641c\u7d22')",
        "button:has-text('\u641c\u7d22')",
        "[role='button']:has-text('\u641c\u7d22')",
    )
    while time.monotonic() < deadline:
        for selector in selectors:
            locator = page.locator(selector)
            try:
                count = min(locator.count(), 8)
            except Exception:
                continue
            for index in range(count):
                button = locator.nth(index)
                try:
                    if button.is_visible(timeout=500) and button.is_enabled(timeout=500):
                        panel_text = button.evaluate(
                            """
                            el => {
                              const textOf = node => (node?.innerText || node?.textContent || '').trim();
                              const panel = el.closest(
                                '[role="dialog"],[class*="modal"],[class*="popup"],[class*="image-search"],[class*="ImageSearch"]'
                              ) || el.parentElement;
                              return textOf(panel);
                            }
                            """
                        )
                        if "\u6309\u56fe\u7247\u641c\u7d22" in panel_text or selector == "#image-search-upload-button":
                            button.click(timeout=5_000)
                            return
                except Exception:
                    continue

        try:
            clicked = bool(
                page.evaluate(
                    """
                    () => {
                      const visible = el => {
                        const r = el.getBoundingClientRect();
                        const s = getComputedStyle(el);
                        return r.width > 0 && r.height > 0
                          && s.display !== 'none'
                          && s.visibility !== 'hidden'
                          && s.pointerEvents !== 'none';
                      };
                      const textOf = el => [
                        el?.innerText || '',
                        el?.textContent || '',
                        el?.getAttribute?.('aria-label') || '',
                        el?.getAttribute?.('title') || ''
                      ].join(' ').replace(/\\s+/g, ' ').trim();
                      const controls = Array.from(document.querySelectorAll(
                        'button,a,[role="button"],div,span'
                      )).filter(visible);
                      const candidates = controls.filter(el => {
                        const text = textOf(el);
                        if (!/^搜索$/.test(text)) return false;
                        const disabled = el.disabled || el.getAttribute('aria-disabled') === 'true';
                        if (disabled) return false;
                        const panel = el.closest(
                          '[role="dialog"],[class*="modal"],[class*="popup"],[class*="image-search"],[class*="ImageSearch"]'
                        ) || el.parentElement;
                        const panelText = textOf(panel);
                        return panelText.includes('按图片搜索')
                          || Boolean(panel?.querySelector?.('input[type="file"]'))
                          || panelText.includes('搜同款');
                      });
                      const target = candidates[candidates.length - 1];
                      if (!target) return false;
                      target.click();
                      return true;
                    }
                    """
                )
            )
            if clicked:
                return
        except Exception:
            pass
        page.wait_for_timeout(500)

    raise CollectorPaused("Image search submit button was not found or was not clickable.")


def wait_for_submitted_image_search_result(
    context: BrowserContext,
    page: Page,
    excluded_page_ids: set[int] | None = None,
    timeout_ms: int = 45_000,
) -> Page:
    deadline = time.monotonic() + timeout_ms / 1000
    current = page
    while time.monotonic() < deadline:
        result_page = find_result_page(context, excluded_page_ids)
        if result_page is not None:
            result_page.wait_for_load_state("domcontentloaded", timeout=30_000)
            detect_stop(result_page)
            return result_page
        if current.is_closed():
            current = select_active_page(context, current)
        close_common_popups(current)
        if is_taobao_result_url(current.url):
            current.wait_for_load_state("domcontentloaded", timeout=30_000)
            detect_stop(current)
            return current
        page.wait_for_timeout(500)
    raise CollectorPaused(f"Image search did not open a Taobao result page. Current URL: {current.url}")


def sales_sort_url(url: str) -> str:
    parsed = urlparse(url)
    query = parse_qs(parsed.query, keep_blank_values=True)
    query["sort"] = ["sale-desc"]
    return urlunparse(parsed._replace(query=urlencode(query, doseq=True)))


def read_sales_sort_state(page: Page) -> dict[str, Any]:
    try:
        state = page.evaluate(
            """
            () => {
              const TEXT_ALL = '\u7efc\u5408';
              const TEXT_SALES = '\u9500\u91cf';
              const TEXT_PRICE = '\u4ef7\u683c';
              const TEXT_HAS_SALES = '\u6709\u9500\u91cf';
              const visible = el => {
                const r = el.getBoundingClientRect();
                const s = getComputedStyle(el);
                return r.width > 0 && r.height > 0
                  && s.display !== 'none'
                  && s.visibility !== 'hidden'
                  && s.pointerEvents !== 'none';
              };
              const normalize = value => (value || '').replace(/\\s+/g, '').trim();
              const rawText = el => [
                el?.innerText || '',
                el?.textContent || '',
                el?.getAttribute?.('aria-label') || '',
                el?.getAttribute?.('title') || ''
              ].join(' ');
              const textOf = el => normalize(rawText(el));
              const isSortLabel = (text, label) => {
                if (text === label) return true;
                if (text === `${label}${label}`) return true;
                return text.includes(label)
                  && text.length <= label.length * 2 + 2
                  && !text.includes(TEXT_HAS_SALES);
              };
              const closestSortRow = el => {
                let node = el;
                for (let i = 0; i < 10 && node; i += 1) {
                  const text = textOf(node);
                  if (text.includes(TEXT_ALL)
                    && text.includes(TEXT_SALES)
                    && text.includes(TEXT_PRICE)
                    && !text.includes(TEXT_HAS_SALES)) {
                    return node;
                  }
                  node = node.parentElement;
                }
                return null;
              };
              const activeScore = el => {
                let score = 0;
                let node = el;
                for (let i = 0; i < 4 && node; i += 1) {
                  const classText = String(node.className || '');
                  const ariaSelected = node.getAttribute?.('aria-selected');
                  const ariaCurrent = node.getAttribute?.('aria-current');
                  const style = getComputedStyle(node);
                  const color = style.color || '';
                  if (/active|selected|current|checked|on/i.test(classText)) score += 2;
                  if (ariaSelected === 'true' || ariaCurrent === 'true' || ariaCurrent === 'page') score += 2;
                  if (/rgb\\(\\s*(2[0-5]{2}|1[8-9]\\d)\\s*,\\s*([0-9]{1,2}|1[0-2]\\d)\\s*,\\s*0\\s*\\)/.test(color)) score += 3;
                  if (/#[fF][fF](5|6|7|8|9|a|A|b|B|c|C)/.test(style.color || '')) score += 3;
                  node = node.parentElement;
                }
                return score;
              };
              const allControls = Array.from(document.querySelectorAll(
                'a,button,[role="button"],[role="tab"],span,div,li'
              )).filter(visible);
              const salesCandidates = allControls
                .filter(el => isSortLabel(textOf(el), TEXT_SALES))
                .map(el => {
                  const row = closestSortRow(el);
                  const rect = el.getBoundingClientRect();
                  return { el, row, rect };
                })
                .filter(item => item.row)
                .sort((a, b) => (a.rect.top - b.rect.top) || (a.rect.left - b.rect.left));
              const sales = salesCandidates[0];
              if (!sales) return { found: false, active: '', rowText: '', salesCenter: null };
              const rowControls = Array.from(sales.row.querySelectorAll(
                'a,button,[role="button"],[role="tab"],span,div,li'
              )).filter(visible);
              const tabs = [TEXT_ALL, TEXT_SALES, TEXT_PRICE].map(label => {
                const candidates = rowControls.filter(el => isSortLabel(textOf(el), label));
                const item = candidates.find(el => el.getAttribute('role') === 'tab') || candidates[0];
                if (!item) return null;
                const rect = item.getBoundingClientRect();
                return {
                  label,
                  score: activeScore(item),
                  center: { x: rect.left + rect.width / 2, y: rect.top + rect.height / 2 }
                };
              }).filter(Boolean);
              const activeTab = [...tabs].sort((a, b) => b.score - a.score)[0];
              const salesRect = sales.rect;
              return {
                found: true,
                active: activeTab && activeTab.score > 0 ? activeTab.label : '',
                rowText: textOf(sales.row).slice(0, 120),
                salesCenter: { x: salesRect.left + salesRect.width / 2, y: salesRect.top + salesRect.height / 2 },
                tabs
              };
            }
            """
        )
        return dict(state) if isinstance(state, dict) else {"found": False}
    except Exception:
        return {"found": False}


def close_stale_result_pages(context: BrowserContext) -> None:
    for item in list(context.pages):
        if item.is_closed():
            continue
        try:
            if is_taobao_result_url(item.url):
                item.close()
        except Exception:
            continue


def switch_to_sales_sort(page: Page, timeout_ms: int = 20_000) -> None:
    close_common_popups(page)
    detect_stop(page)
    state = read_sales_sort_state(page)
    if state.get("active") == "\u9500\u91cf":
        print("[collector] Taobao sales sort is active", flush=True)
        return

    center = state.get("salesCenter") if state.get("found") else None
    if isinstance(center, dict):
        print(f"[collector] clicking Taobao sales sort: {state.get('rowText', '')}", flush=True)
        wait_before_high_risk_click(page)
        page.mouse.click(float(center["x"]), float(center["y"]))
        page.wait_for_load_state("domcontentloaded", timeout=30_000)
        page.wait_for_timeout(NAVIGATION_SETTLE_WAIT_MS)
        close_common_popups(page)
        detect_stop(page)
        if read_sales_sort_state(page).get("active") == "\u9500\u91cf":
            print("[collector] Taobao sales sort is active", flush=True)
            return

    if is_taobao_result_url(page.url):
        target_url = sales_sort_url(page.url)
        if target_url != page.url:
            print(f"[collector] opening Taobao sales sort URL: {target_url}", flush=True)
            page.goto(target_url, wait_until="domcontentloaded", timeout=30_000)
            page.wait_for_timeout(NAVIGATION_SETTLE_WAIT_MS)
            close_common_popups(page)
            detect_stop(page)
            if read_sales_sort_state(page).get("active") == "\u9500\u91cf":
                print("[collector] Taobao sales sort is active", flush=True)
                return

    raise CollectorPaused("Could not verify that the Taobao result page switched to the native sales sort tab.")


def perform_image_search(
    page: Page,
    context: BrowserContext,
    reference_image: Path,
    login_timeout_seconds: int,
    login_poll_seconds: int,
) -> Page:
    if not reference_image or not reference_image.is_file():
        raise ValueError("reference image is required unless --use-current-page is set")
    print("[collector] checking Taobao login", flush=True)
    page = wait_for_login(context, page, login_timeout_seconds, login_poll_seconds)

    print("[collector] opening Taobao home", flush=True)
    page.goto(TAOBAO_HOME, wait_until="domcontentloaded", timeout=30_000)
    page.wait_for_timeout(NAVIGATION_SETTLE_WAIT_MS)
    close_common_popups(page)
    if not is_taobao_or_tmall_url(page.url):
        page.goto(TAOBAO_HOME, wait_until="domcontentloaded", timeout=30_000)
        page.wait_for_timeout(NAVIGATION_SETTLE_WAIT_MS)
    page = wait_for_login(context, page, login_timeout_seconds, login_poll_seconds)
    detect_stop(page)

    print("[collector] opening image search entry", flush=True)
    wait_before_high_risk_click(page)
    for text in (
        "\u641c\u540c\u6b3e",
        "\u62cd\u7acb\u6dd8",
        "\u56fe\u7247\u641c\u7d22",
    ):
        if click_first_visible_text(page, text, timeout_ms=2_000):
            page.wait_for_timeout(500)
            break
    else:
        for selector in (
            ".image-search-icon-wrapper",
            "[class*='image-search']",
            "[class*='camera']",
            "[aria-label*='\u641c\u540c\u6b3e']",
            "[title*='\u641c\u540c\u6b3e']",
        ):
            locator = page.locator(selector)
            try:
                if locator.count() > 0 and locator.first.is_visible():
                    locator.first.click()
                    page.wait_for_timeout(500)
                    break
            except Exception:
                continue

    existing_other_page_ids = {id(item) for item in context.pages if item != page and not item.is_closed()}
    file_input = page.locator("#image-search-custom-file-input")
    if file_input.count() == 0:
        file_input = page.locator('input[type="file"]').first
    file_input.wait_for(state="attached", timeout=15_000)
    print("[collector] uploading reference image", flush=True)
    file_input.set_input_files(str(reference_image))
    page.wait_for_timeout(NAVIGATION_SETTLE_WAIT_MS)

    auto_result_page = find_result_page(context, existing_other_page_ids)
    if auto_result_page is not None:
        auto_result_page.wait_for_load_state("domcontentloaded", timeout=30_000)
        detect_stop(auto_result_page)
        print(f"[collector] image search result page: {auto_result_page.url}", flush=True)
        return auto_result_page

    print("[collector] submitting image search", flush=True)
    wait_before_high_risk_click(page)
    click_image_search_submit(page)
    result_page = wait_for_submitted_image_search_result(context, page, existing_other_page_ids)

    result_page.wait_for_load_state("domcontentloaded", timeout=30_000)
    result_page.wait_for_timeout(NAVIGATION_SETTLE_WAIT_MS)
    detect_stop(result_page)
    print(f"[collector] image search result page: {result_page.url}", flush=True)
    return result_page


def collect_candidate_rows(page: Page, candidate_limit: int, scroll_rounds: int) -> list[CandidateRow]:
    rows: list[CandidateRow] = []
    seen_keys: set[str] = set()

    for round_index in range(max(1, scroll_rounds)):
        close_common_popups(page)
        batch = page.locator('a[href]').evaluate_all(
            """
            (links, limit) => {
              const visible = el => {
                const r = el.getBoundingClientRect();
                const s = getComputedStyle(el);
                return r.width > 0 && r.height > 0 && s.display !== 'none'
                  && s.visibility !== 'hidden';
              };
              const normalizeText = value => (value || '').replace(/\\s+/g, ' ').trim();
              const productId = href => {
                try {
                  const u = new URL(href, location.href);
                  return u.searchParams.get('id') || u.searchParams.get('itemId') || '';
                } catch (_) {
                  return '';
                }
              };
              const cardOf = link => link.closest(
                'div[class*="item"],div[class*="Item"],li,article,[data-index]'
              ) || link;
              return links.filter(visible).map((link, index) => {
                const href = link.href || '';
                const pid = productId(href);
                if (!pid && !/imgsearch\\.item|xxc=taobaoSearch/i.test(href)) return null;
                const card = cardOf(link);
                const text = normalizeText(card.innerText || link.innerText || '');
                const imgs = Array.from(card.querySelectorAll('img')).filter(img => {
                  if (!visible(img)) return false;
                  const r = img.getBoundingClientRect();
                  return r.width >= 80 && r.height >= 80;
                });
                const img = imgs[0];
                const titled = Array.from(card.querySelectorAll('[title]'))
                  .map(el => el.getAttribute('title') || '')
                  .find(value => value.length >= 6) || '';
                return {
                  source_order: index + 1,
                  href,
                  raw_text: text,
                  title_hint: titled,
                  image_url: img ? (img.currentSrc || img.src || img.getAttribute('data-src') || '') : ''
                };
              }).filter(row => row && row.href && row.image_url).slice(0, limit);
            }
            """,
            candidate_limit,
        )
        for item in batch:
            href = str(item.get("href") or "")
            image_url = normalize_image_url(str(item.get("image_url") or ""))
            key = f"{product_id_from_url(href)}|{image_url}"
            if key in seen_keys:
                continue
            seen_keys.add(key)
            rows.append(
                CandidateRow(
                    source_order=int(item.get("source_order") or len(rows) + 1),
                    href=href,
                    raw_text=str(item.get("raw_text") or ""),
                    title_hint=str(item.get("title_hint") or ""),
                    image_url=image_url,
                )
            )
        if len(rows) >= candidate_limit:
            break
        if round_index < scroll_rounds - 1:
            page.mouse.wheel(0, 1400)
            page.wait_for_timeout(1_000)
            close_common_popups(page)
    return rows


def rows_to_products(rows: list[CandidateRow], min_sales_floor: int) -> list[dict[str, Any]]:
    products: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows:
        product_id = product_id_from_url(row.href)
        if not product_id or product_id in seen:
            continue
        sales_floor = parse_sales_floor(row.raw_text)
        if sales_floor is None or sales_floor < min_sales_floor:
            continue
        title = row.title_hint.strip()
        if not title:
            title = re.sub(r"\s+", " ", row.raw_text).strip()[:180]
        products.append(
            {
                "product_id": product_id,
                "source_order": row.source_order,
                "title": title,
                "sales_text": parse_sales_text(row.raw_text),
                "sales_floor": sales_floor,
                "item_url": row.href,
                "source_card_url": row.href,
                "source_image_url": row.image_url,
            }
        )
        seen.add(product_id)
    return sorted(products, key=lambda item: (-(item["sales_floor"] or 0), item["source_order"]))


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def public_url_for(path: Path, root: Path, prefix: str) -> str:
    if not prefix:
        return ""
    relative = path.resolve().relative_to(root.resolve()).as_posix()
    return f"{prefix.rstrip('/')}/{quote(relative)}"


def close_result_page_after_collection(context: BrowserContext, result_page: Page) -> None:
    if result_page.is_closed():
        return
    try:
        open_pages = [item for item in context.pages if not item.is_closed()]
        other_pages = [item for item in open_pages if item != result_page]
        if other_pages:
            result_page.close()
        elif is_taobao_result_url(result_page.url):
            placeholder = context.new_page()
            try:
                placeholder.goto(TAOBAO_HOME, wait_until="domcontentloaded", timeout=15_000)
            except Exception:
                pass
            result_page.close()
    except Exception:
        pass


def download_image_from_url(context: BrowserContext, url: str, destination: Path) -> tuple[str, int, int]:
    candidates = [normalize_image_url(url), url]
    last_error = ""
    for candidate in dict.fromkeys(item for item in candidates if item):
        try:
            response = context.request.get(candidate, timeout=30_000)
            if not response.ok:
                last_error = f"HTTP {response.status}"
                continue
            destination.parent.mkdir(parents=True, exist_ok=True)
            tmp = destination.with_suffix(".download")
            tmp.write_bytes(response.body())
            with Image.open(tmp) as source:
                image = ImageOps.exif_transpose(source).convert("RGB")
                width, height = image.size
                image.thumbnail((1800, 1800), Image.Resampling.LANCZOS)
                image.save(destination, "JPEG", quality=92, optimize=True, exif=b"")
            tmp.unlink(missing_ok=True)
            return sha256_file(destination), width, height
        except Exception as error:  # noqa: BLE001 - return clear per-item failure
            last_error = f"{type(error).__name__}: {error}"
            destination.unlink(missing_ok=True)
    raise RuntimeError(last_error or "download failed")


def files_from_download(
    download_path: Path,
    extract_dir: Path,
    supported_extensions: set[str],
) -> list[Path]:
    if zipfile.is_zipfile(download_path):
        extract_dir.mkdir(parents=True, exist_ok=True)
        resolved_extract_dir = extract_dir.resolve()
        with zipfile.ZipFile(download_path) as archive:
            for member in archive.infolist():
                if member.is_dir():
                    continue
                target = safe_archive_member_path(extract_dir, member.filename)
                if target is None:
                    continue
                target = target.resolve()
                try:
                    target.relative_to(resolved_extract_dir)
                except ValueError:
                    continue
                target.parent.mkdir(parents=True, exist_ok=True)
                with archive.open(member) as source, target.open("wb") as output:
                    shutil.copyfileobj(source, output)
        return [
            item
            for item in sorted(extract_dir.rglob("*"))
            if item.is_file() and item.suffix.lower() in supported_extensions
        ]
    if download_path.suffix.lower() in supported_extensions:
        return [download_path]
    return []


def safe_archive_member_path(extract_dir: Path, member_name: str) -> Path | None:
    archive_path = PurePosixPath(member_name.replace("\\", "/"))
    if archive_path.is_absolute() or ".." in archive_path.parts:
        return None

    parts = [part for part in archive_path.parts if part not in ("", ".")]
    if not parts:
        return None

    safe_parts: list[str] = []
    for part in parts:
        safe_part = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", part).rstrip(". ") or "_"
        if Path(safe_part).stem.upper() in WINDOWS_RESERVED_FILE_STEMS:
            safe_part = f"_{safe_part}"
        safe_parts.append(safe_part)
    return extract_dir.joinpath(*safe_parts)


def image_files_from_download(download_path: Path, extract_dir: Path) -> list[Path]:
    return files_from_download(download_path, extract_dir, SUPPORTED_IMAGE_EXTENSIONS)


def classify_store_insight_asset_file(source: Path, extract_dir: Path) -> str:
    try:
        relative = source.relative_to(extract_dir)
    except ValueError:
        relative = source

    directory_parts = {
        re.sub(r"[\s:_-]+", "", part).lower()
        for part in relative.parts[:-1]
    }
    for asset_type in ("detail_long", "detail", "sku", "display_main", "main"):
        if directory_parts & STORE_INSIGHT_ARCHIVE_DIR_ALIASES[asset_type]:
            return asset_type
    return ""


def filter_store_insight_asset_files(
    extracted: list[Path],
    extract_dir: Path,
    asset_type: str,
) -> list[Path]:
    classified = [
        (source, classify_store_insight_asset_file(source, extract_dir))
        for source in extracted
    ]
    matching = [source for source, detected_type in classified if detected_type == asset_type]
    if matching:
        return matching
    if any(detected_type for _source, detected_type in classified):
        return []
    return extracted


def store_insight_sku_source_index(filename: str) -> int | None:
    match = STORE_INSIGHT_SKU_FILENAME_PATTERN.match(Path(filename).stem.strip())
    return int(match.group("index")) if match else None


def parse_store_insight_sku_filename(filename: str, fallback_index: int) -> dict[str, Any]:
    stem = Path(filename).stem.strip()
    match = STORE_INSIGHT_SKU_FILENAME_PATTERN.match(stem)
    index = int(match.group("index")) if match else fallback_index
    sku_label = (match.group("label") if match else stem).strip()
    if not match:
        return {
            "index": index,
            "sku_label": sku_label,
            "spec_text": "",
            "color_text": "",
            "parse_status": "unparsed",
        }

    bracket_values = [
        next(value for value in groups if value).strip()
        for groups in STORE_INSIGHT_SKU_BRACKET_PATTERN.findall(sku_label)
    ]
    spec_pair = STORE_INSIGHT_SKU_SPEC_PAIR_PATTERN.search(sku_label)
    spec_match = next(
        (
            match
            for value in bracket_values
            if (match := STORE_INSIGHT_SKU_COUNT_PATTERN.search(value))
        ),
        None,
    )
    spec_text = spec_pair.group(1).strip() if spec_pair else spec_match.group(0) if spec_match else ""
    if not spec_text:
        for value in bracket_values:
            size_match = STORE_INSIGHT_SKU_SIZE_TOKEN_PATTERN.search(value)
            if size_match:
                spec_text = size_match.group(0).strip()
                break
    if not spec_text:
        size_match = STORE_INSIGHT_SKU_SIZE_TOKEN_PATTERN.search(sku_label)
        spec_text = size_match.group(0).strip() if size_match else ""

    unbracketed_color_source = STORE_INSIGHT_SKU_BRACKET_PATTERN.sub("", sku_label)
    color_sources = []
    for value in bracket_values:
        count_match = STORE_INSIGHT_SKU_COUNT_PATTERN.search(value)
        if count_match:
            value = value[: count_match.start()] + value[count_match.end() :]
        if value:
            color_sources.append(value)
    color_sources.append(unbracketed_color_source)
    color_pair = STORE_INSIGHT_SKU_COLOR_PAIR_PATTERN.search(sku_label)
    color_text = color_pair.group(1).strip() if color_pair else ""
    for color_source in color_sources:
        if color_text:
            break
        color_source = STORE_INSIGHT_SKU_SPEC_PAIR_PATTERN.sub("", color_source)
        color_source = STORE_INSIGHT_SKU_COLOR_PAIR_PATTERN.sub("", color_source)
        color_source = STORE_INSIGHT_SKU_SIZE_TOKEN_PATTERN.sub("", color_source)
        candidate = re.sub(r"（[^）]*）|\([^)]*\)", "", color_source).strip(" _-")
        if not candidate or STORE_INSIGHT_SKU_NON_COLOR_MARKER_PATTERN.search(candidate):
            continue
        for segment in re.split(r"[;；,，|/\s]+", candidate):
            color_tokens = [token.strip() for token in re.split(r"[+＋]", segment) if token.strip()]
            if color_tokens and all(
                STORE_INSIGHT_SKU_COLOR_TOKEN_PATTERN.fullmatch(token) for token in color_tokens
            ):
                color_text = "+".join(color_tokens)
                break
    parse_status = (
        "parsed"
        if spec_text and color_text
        else "partial"
        if spec_text or color_text
        else "unparsed"
    )
    return {
        "index": index,
        "sku_label": sku_label,
        "spec_text": spec_text,
        "color_text": color_text,
        "parse_status": parse_status,
    }


def sku_table_metadata_for_image(
    source: Path,
    source_index: int | None,
    fallback_index: int,
    rows: list[dict[str, str]],
    used_source_indexes: set[str],
) -> dict[str, str] | None:
    sku_ids = STORE_INSIGHT_SKU_ID_IN_FILENAME_PATTERN.findall(source.stem)
    for sku_id in sku_ids:
        match = next(
            (
                row
                for row in rows
                if row.get("sku_id") == sku_id
                and row.get("source_index", "") not in used_source_indexes
            ),
            None,
        )
        if match is not None:
            return match
    if source_index is not None:
        match = next(
            (
                row
                for row in rows
                if row.get("source_index") == str(source_index)
                and row.get("source_index", "") not in used_source_indexes
            ),
            None,
        )
        if match is not None:
            return match
    if source_index is None and 0 < fallback_index <= len(rows):
        candidate = rows[fallback_index - 1]
        if candidate.get("source_index", "") not in used_source_indexes:
            return candidate
    return None


def sku_table_fields(row: dict[str, str] | None) -> dict[str, str]:
    row = row or {}
    return {
        "sku_id": row.get("sku_id", ""),
        "list_price": row.get("list_price", ""),
        "after_coupon_price": row.get("after_coupon_price", ""),
        "net_content": row.get("net_content", ""),
        "stock": row.get("stock", ""),
    }


def normalize_local_image(source: Path, destination: Path) -> tuple[str, int, int]:
    with Image.open(source) as opened:
        image = ImageOps.exif_transpose(opened).convert("RGB")
        width, height = image.size
        image.thumbnail((1800, 1800), Image.Resampling.LANCZOS)
        destination.parent.mkdir(parents=True, exist_ok=True)
        image.save(destination, "JPEG", quality=92, optimize=True, exif=b"")
    return sha256_file(destination), width, height


def click_first_visible_text(page: Page, text: str, timeout_ms: int = 8_000) -> bool:
    detect_stop(page)
    locator = page.get_by_text(text, exact=False)
    deadline = time.monotonic() + timeout_ms / 1000
    while time.monotonic() < deadline:
        for index in range(locator.count()):
            item = locator.nth(index)
            try:
                if item.is_visible():
                    detect_stop(page)
                    item.click()
                    return True
            except Exception:
                continue
        page.wait_for_timeout(250)
    return False


def click_first_visible_exact_text(page: Page, text: str, timeout_ms: int = 8_000) -> bool:
    detect_stop(page)
    locator = page.get_by_text(text, exact=True)
    deadline = time.monotonic() + timeout_ms / 1000
    while time.monotonic() < deadline:
        for index in range(locator.count()):
            item = locator.nth(index)
            try:
                if item.is_visible():
                    detect_stop(page)
                    item.click()
                    return True
            except Exception:
                continue
        page.wait_for_timeout(250)
    return False


def store_insight_asset_result_path(asset_type: str, product_id: str, filename: str) -> str:
    directory = STORE_INSIGHT_ASSET_DIRS[asset_type]
    return f"_work/competitor-assets/{directory}/{product_id}/{filename}"


def write_extended_asset_manifest(
    path: Path,
    config: CollectorConfig,
    records: list[dict[str, Any]],
    metadata: dict[str, Any] | None = None,
) -> None:
    image_count = 0
    collected: dict[str, int] = {}
    for product in records:
        for asset_type, type_entry in product.get("types", {}).items():
            count = sum(1 for image in (type_entry.get("images") or []) if image.get("status") == "ok")
            image_count += count
            collected[asset_type] = collected.get(asset_type, 0) + count
    targets = build_store_insight_asset_targets(config)
    shortfalls = store_insight_asset_shortfalls(targets, collected)
    payload = {
        "schema_version": 1,
        "status": "complete" if records and not shortfalls else "partial" if records else "empty",
        "created_at": utc_now_iso(),
        "reference_image": str(config.reference_image) if config.reference_image else "",
        "asset_targets": targets,
        "asset_collected": collected,
        "asset_shortfalls": shortfalls,
        "product_count": len(records),
        "image_count": image_count,
        "products": records,
    }
    if metadata:
        payload.update(metadata)
        if metadata.get("asset_count_mode") == "dynamic" and any(
                not bool(metadata.get(f"{asset_type}_meets_minimum"))
                for asset_type in DYNAMIC_ASSET_RULES
        ):
            payload["status"] = "partial"
        if metadata.get("main_video_requested") and metadata.get("main_video_status") != "complete":
            payload["status"] = "partial"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def first_asset_image(asset_entry: dict[str, Any] | None, asset_type: str) -> dict[str, Any] | None:
    if not asset_entry:
        return None
    type_entry = dict(asset_entry.get("types") or {}).get(asset_type) or {}
    images = type_entry.get("images") or []
    return images[0] if images else None


def try_store_insight_download(
    context: BrowserContext,
    product: dict[str, Any],
    downloads_dir: Path,
    timeout_ms: int,
) -> Path | None:
    page = context.new_page()
    try:
        page.goto(product["item_url"], wait_until="domcontentloaded", timeout=timeout_ms)
        page.wait_for_timeout(3_000)
        detect_stop(page)
        if not click_first_visible_text(page, "\u5546\u54c1\u56fe\u4e0b\u8f7d", timeout_ms=10_000):
            return None
        page.wait_for_timeout(800)
        if not click_first_visible_text(page, "\u81ea\u5b9a\u4e49\u4e0b\u8f7d", timeout_ms=10_000):
            return None
        page.wait_for_timeout(1_000)

        # Prefer page-display main images. If the checkbox is already selected,
        # this click may toggle it in some UIs, so we only click when there is
        # no visible selected marker near the label.
        label = page.get_by_text("\u9875\u9762\u5c55\u793a\u4e3b\u56fe", exact=False)
        if label.count() > 0:
            try:
                label.first.click()
            except Exception:
                pass

        downloads_dir.mkdir(parents=True, exist_ok=True)
        with page.expect_download(timeout=min(timeout_ms, STORE_INSIGHT_DOWNLOAD_TIMEOUT_MS)) as download_info:
            if not click_first_visible_text(page, "\u6309\u7c7b\u578b\u4e0b\u8f7d", timeout_ms=10_000):
                return None
        download = download_info.value
        filename = store_insight_download_filename(
            str(product["product_id"]),
            download.suggested_filename or "images.zip",
        )
        save_path = downloads_dir / filename
        download.save_as(str(save_path))
        return save_path
    except CollectorPaused:
        raise
    except Exception:
        return None
    finally:
        page.close()


def download_store_insight_asset_package_on_page(
    page: Page,
    product: dict[str, Any],
    asset_type: str,
    downloads_dir: Path,
    timeout_ms: int,
    *,
    strict_download: bool = False,
) -> Path | None:
    try:
        detect_stop(page)
        wait_before_high_risk_click(page)
        if not click_first_visible_text(page, "\u5546\u54c1\u56fe\u4e0b\u8f7d", timeout_ms=10_000):
            if strict_download:
                raise StoreInsightAssetUnavailable(
                    f"Store Insight download entry was unavailable for "
                    f"{product['product_id']} {asset_type}."
                )
            return None
        page.wait_for_timeout(800)

        downloads_dir.mkdir(parents=True, exist_ok=True)
        wait_before_high_risk_click(page)
        with page.expect_download(timeout=min(timeout_ms, STORE_INSIGHT_DOWNLOAD_TIMEOUT_MS)) as download_info:
            if not click_first_visible_exact_text(
                page,
                STORE_INSIGHT_QUICK_DOWNLOAD_LABELS[asset_type],
                timeout_ms=10_000,
            ):
                if strict_download:
                    raise StoreInsightAssetUnavailable(
                        f"Store Insight quick action was unavailable for "
                        f"{product['product_id']} {asset_type}."
                    )
                return None
        download = download_info.value
        filename = store_insight_download_filename(
            str(product["product_id"]),
            download.suggested_filename or "images.zip",
            asset_type,
        )
        save_path = downloads_dir / filename
        download.save_as(str(save_path))
        if strict_download and (not save_path.is_file() or save_path.stat().st_size <= 0):
            raise StoreInsightAssetUnavailable(
                f"Store Insight download was not saved for {product['product_id']} {asset_type}."
            )
        if not dismiss_store_insight_download_completion(page):
            if strict_download:
                raise StoreInsightAssetUnavailable(
                    f"Store Insight download completion popup could not be closed for "
                    f"{product['product_id']} {asset_type}."
                )
        return save_path
    except StoreInsightAssetUnavailable:
        raise
    except CollectorPaused:
        raise
    except Exception as error:
        detect_stop(page)
        if strict_download:
            raise StoreInsightAssetUnavailable(
                f"Store Insight download failed for {product['product_id']} {asset_type}: "
                f"{type(error).__name__}: {error}"
            ) from error
        return None


def download_store_insight_asset_package(
    context: BrowserContext,
    product: dict[str, Any],
    asset_type: str,
    downloads_dir: Path,
    timeout_ms: int,
) -> Path | None:
    page = context.new_page()
    try:
        page.goto(product["item_url"], wait_until="domcontentloaded", timeout=timeout_ms)
        page.wait_for_timeout(3_000)
        return download_store_insight_asset_package_on_page(
            page,
            product,
            asset_type,
            downloads_dir,
            timeout_ms,
        )
    except CollectorPaused:
        raise
    except Exception:
        return None
    finally:
        page.close()


def download_store_insight_sku_export_on_page(
    page: Page,
    product: dict[str, Any],
    label: str,
    downloads_dir: Path,
    timeout_ms: int,
    fallback_filename: str,
) -> Path | None:
    """Download one visible SKU-preview export without using internal APIs."""
    try:
        radio = page.locator("label[role='radio']").filter(has_text=label)
        for index in range(radio.count()):
            item = radio.nth(index)
            if item.is_visible() and item.get_attribute("aria-checked") != "true":
                item.click()
                page.wait_for_timeout(300)
                break
        locator = page.locator("button").filter(has_text=label)
        download_timeout_ms = min(timeout_ms, STORE_INSIGHT_DOWNLOAD_TIMEOUT_MS)
        deadline = time.monotonic() + download_timeout_ms / 1000
        while time.monotonic() < deadline:
            for index in range(locator.count()):
                item = locator.nth(index)
                if not item.is_visible():
                    continue
                downloads_dir.mkdir(parents=True, exist_ok=True)
                with page.expect_download(timeout=download_timeout_ms) as download_info:
                    item.click()
                download = download_info.value
                suggested = re.sub(
                    r'[<>:"/\\|?*]+',
                    "_",
                    download.suggested_filename or fallback_filename,
                )
                save_path = downloads_dir / f"{product['product_id']}_{suggested}"
                download.save_as(str(save_path))
                return save_path if save_path.is_file() and save_path.stat().st_size > 0 else None
            page.wait_for_timeout(250)
    except CollectorPaused:
        raise
    except Exception:
        return None
    return None


def download_store_insight_sku_image_archive_on_page(
    page: Page,
    product: dict[str, Any],
    downloads_dir: Path,
    timeout_ms: int,
) -> Path | None:
    """Download only the first eight visible SKU preview images."""
    try:
        page.wait_for_timeout(1_000)
        image_buttons = page.locator("button").filter(has_text="下载图片")
        downloaded: list[Path] = []
        limit = int(DYNAMIC_ASSET_RULES["sku"]["maximum"])
        for index in range(min(image_buttons.count(), limit)):
            button = image_buttons.nth(index)
            try:
                button.scroll_into_view_if_needed(timeout=min(timeout_ms, 3_000))
                if not button.is_visible():
                    continue
                with page.expect_download(timeout=min(timeout_ms, 3_000)) as download_info:
                    button.click()
                download = download_info.value
                downloads_dir.mkdir(parents=True, exist_ok=True)
                suggested = re.sub(
                    r'[<>:"/\\|?*]+',
                    "_",
                    download.suggested_filename or f"SKU图_{index + 1:02d}.jpg",
                )
                save_path = downloads_dir / f"{product['product_id']}_image_{index + 1:02d}_{suggested}"
                download.save_as(str(save_path))
                if save_path.is_file() and save_path.stat().st_size > 0:
                    downloaded.append(save_path)
            except Exception:
                no_image_message = page.locator(".el-message").filter(has_text="暂无图片")
                if any(
                    no_image_message.nth(message_index).is_visible()
                    for message_index in range(no_image_message.count())
                ):
                    break
                continue
        if not downloaded:
            return None
        archive_path = downloads_dir / f"{product['product_id']}_sku-images.zip"
        with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for image_path in downloaded:
                archive.write(image_path, f"SKU图/{image_path.name}")
        return archive_path
    except CollectorPaused:
        raise
    except Exception:
        return None


def download_store_insight_sku_exports_on_page(
    page: Page,
    product: dict[str, Any],
    downloads_dir: Path,
    timeout_ms: int,
) -> tuple[Path | None, Path | None]:
    """Download the SKU table and image package from Store Insight's visible SKU preview."""
    try:
        detect_stop(page)
        context = page.context
        original_pages = set(context.pages)
        if not click_first_visible_text(page, "SKU预览", timeout_ms=10_000):
            return None, None
        page.wait_for_timeout(800)
        preview_page = next(
            (candidate for candidate in context.pages if candidate not in original_pages),
            page,
        )
        if preview_page is not page:
            preview_page.wait_for_load_state("domcontentloaded", timeout=timeout_ms)
        table_path = download_store_insight_sku_export_on_page(
            preview_page,
            product,
            "导出表格",
            downloads_dir,
            timeout_ms,
            "sku-table.xlsx",
        )
        image_path = download_store_insight_sku_image_archive_on_page(
            preview_page,
            product,
            downloads_dir,
            timeout_ms,
        )
        if preview_page is not page:
            preview_page.close()
        return table_path, image_path
    except CollectorPaused:
        raise
    except Exception:
        return None, None


def inspected_store_insight_sku_exports(
    product: dict[str, Any],
    downloads_dir: Path,
    table_path: Path | None,
    image_path: Path | None,
) -> tuple[InspectedAssetPackage, list[dict[str, str]]]:
    table_rows: list[dict[str, str]] = []
    if table_path is not None:
        try:
            table_rows = parse_store_insight_sku_table(table_path)
        except Exception:
            table_rows = []
    extract_dir = downloads_dir / f"{product['product_id']}_sku_exports_extracted"
    files = (
        filter_store_insight_asset_files(
            image_files_from_download(image_path, extract_dir),
            extract_dir,
            "sku",
        )
        if image_path is not None
        else []
    )
    return (
        InspectedAssetPackage(
            asset_type="sku",
            product_id=str(product["product_id"]),
            package_path=image_path,
            extract_dir=extract_dir,
            files=tuple(files),
            error="" if files or table_rows else "SKU preview exports were unavailable",
        ),
        table_rows,
    )


def inspect_store_insight_sku_exports_on_page(
    page: Page,
    product: dict[str, Any],
    downloads_dir: Path,
    timeout_ms: int,
) -> tuple[InspectedAssetPackage, list[dict[str, str]]]:
    table_path, image_path = download_store_insight_sku_exports_on_page(
        page,
        product,
        downloads_dir,
        timeout_ms,
    )
    return inspected_store_insight_sku_exports(
        product,
        downloads_dir,
        table_path,
        image_path,
    )


def inspect_store_insight_sku_exports(
    context: BrowserContext,
    product: dict[str, Any],
    downloads_dir: Path,
    timeout_ms: int,
) -> tuple[InspectedAssetPackage, list[dict[str, str]]]:
    page = context.new_page()
    try:
        page.goto(product["item_url"], wait_until="domcontentloaded", timeout=timeout_ms)
        page.wait_for_timeout(3_000)
        return inspect_store_insight_sku_exports_on_page(
            page,
            product,
            downloads_dir,
            timeout_ms,
        )
    finally:
        page.close()


def collect_first_store_insight_main_video(
    context: BrowserContext,
    ranked_products: list[dict[str, Any]],
    downloads_dir: Path,
    output_root: Path,
    timeout_ms: int,
) -> dict[str, Any]:
    del downloads_dir, output_root
    metadata: dict[str, Any] = {
        "main_video_requested": True,
        "main_video_url": "",
        "main_video_result_path": "",
        "main_video_local_path": "",
        "main_video_count": 0,
        "main_video_source_product_id": "",
        "main_video_status": "not_found",
    }
    candidates = ranked_products[:MAX_VIDEO_PRODUCT_PROBES]
    for probe_index, product in enumerate(candidates, start=1):
        print(
            f"[collector] checking main video {probe_index}/{len(candidates)}: "
            f"{product.get('product_id', '')}",
            flush=True,
        )
        product_id = str(product.get("product_id") or "")
        video_url = collect_product_video_url(context, product, timeout_ms)
        if not video_url:
            continue
        metadata.update(
            {
                "main_video_url": video_url,
                "main_video_count": 1,
                "main_video_source_product_id": product_id,
                "main_video_status": "complete",
            }
        )
        break
    if metadata["main_video_status"] != "complete":
        print(
            f"[collector] main video not found after {len(candidates)} probes",
            flush=True,
        )
    return metadata


def normalize_video_url(value: str) -> str:
    candidate = html.unescape(str(value or "").strip().strip('"\''))
    candidate = candidate.replace("\\u002F", "/").replace("\\/", "/").replace("\\u0026", "&")
    if "%2F" in candidate.upper():
        candidate = unquote(candidate)
    if candidate.startswith("//"):
        candidate = "https:" + candidate
    if not candidate.startswith(("http://", "https://")):
        return ""
    return candidate


def select_video_url(values: list[str]) -> str:
    candidates: list[tuple[int, int, str]] = []
    seen: set[str] = set()
    for order, value in enumerate(values):
        candidate = normalize_video_url(value)
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        lowered = candidate.lower()
        if lowered.endswith(".ts") or ".ts?" in lowered:
            continue
        is_direct_video = (
            ".mp4" in lowered
            or ".m3u8" in lowered
            or "/play/" in lowered
            or "cloud.video.taobao.com" in lowered
        )
        if not is_direct_video:
            continue
        score = 0
        if ".mp4" in lowered:
            score += 100
        elif ".m3u8" in lowered:
            score += 80
        if "cloud.video.taobao.com" in lowered:
            score += 50
        elif "video" in lowered or "alicdn" in lowered:
            score += 20
        if "/play/" in lowered:
            score += 20
        if score:
            candidates.append((score, -order, candidate))
    return max(candidates, default=(0, 0, ""))[2]


def extract_product_video_urls(page: Page, captured_urls: list[str] | None = None) -> list[str]:
    values = list(captured_urls or [])
    values.extend(
        page.evaluate(
            """
            () => {
                const values = [];
                const add = value => {
                    if (typeof value === 'string' && value.trim()) values.push(value.trim());
                };
                for (const element of document.querySelectorAll('video, video source')) {
                    add(element.currentSrc);
                    add(element.src);
                    for (const name of ['src', 'data-src', 'data-video-url', 'data-video-src']) {
                        add(element.getAttribute(name));
                    }
                }
                for (const element of document.querySelectorAll(
                    '[data-video-url], [data-video-src], [data-video], [video-url], [video-src]'
                )) {
                    for (const name of ['data-video-url', 'data-video-src', 'data-video', 'video-url', 'video-src']) {
                        add(element.getAttribute(name));
                    }
                }
                for (const entry of performance.getEntriesByType('resource')) add(entry.name);
                return values;
            }
            """
        )
    )
    page_source = html.unescape(page.content())
    page_source = page_source.replace("\\u002F", "/").replace("\\/", "/").replace("\\u0026", "&")
    values.extend(
        re.findall(
            r"(?:https?:)?//[^\"'<>\s\\]+?(?:\.mp4|\.m3u8)(?:\?[^\"'<>\s\\]*)?",
            page_source,
            flags=re.IGNORECASE,
        )
    )
    values.extend(
        re.findall(
            r"(?:https?:)?//cloud\.video\.taobao\.com/[^\"'<>\s\\]+",
            page_source,
            flags=re.IGNORECASE,
        )
    )
    return values


def collect_product_video_url(
    context: BrowserContext,
    product: dict[str, Any],
    timeout_ms: int,
) -> str:
    item_url = str(product.get("item_url") or "")
    if not item_url:
        return ""
    captured_urls: list[str] = []
    page = context.new_page()

    def record_response(response: Any) -> None:
        try:
            content_type = str(response.headers.get("content-type") or "").lower()
            url = str(response.url or "")
            if "video/" in content_type or "mpegurl" in content_type or select_video_url([url]):
                captured_urls.append(url)
        except Exception:
            return

    page.on("response", record_response)
    try:
        page.goto(item_url, wait_until="domcontentloaded", timeout=timeout_ms)
        page.wait_for_timeout(3_000)
        detect_stop(page)
        try:
            video = page.locator("video").first
            if video.count() and video.is_visible(timeout=1_000):
                video.evaluate("element => { element.muted = true; element.play().catch(() => {}); }")
                page.wait_for_timeout(1_500)
        except Exception:
            pass
        return select_video_url(extract_product_video_urls(page, captured_urls))
    finally:
        page.close()


def inspect_store_insight_asset_package(
    context: BrowserContext,
    product: dict[str, Any],
    asset_type: str,
    downloads_dir: Path,
    timeout_ms: int,
) -> InspectedAssetPackage:
    package_path = download_store_insight_asset_package(
        context,
        product,
        asset_type,
        downloads_dir,
        timeout_ms,
    )
    extract_dir = downloads_dir / f"{product['product_id']}_{asset_type}_extracted"
    if package_path is None:
        return InspectedAssetPackage(
            asset_type=asset_type,
            product_id=str(product["product_id"]),
            package_path=None,
            extract_dir=extract_dir,
            files=(),
            error="Store Insight download did not produce a file",
        )

    extracted = image_files_from_download(package_path, extract_dir)
    extracted = filter_store_insight_asset_files(extracted, extract_dir, asset_type)
    return InspectedAssetPackage(
        asset_type=asset_type,
        product_id=str(product["product_id"]),
        package_path=package_path,
        extract_dir=extract_dir,
        files=tuple(extracted),
        error="" if extracted else "download file contained no usable images",
    )


def inspect_store_insight_asset_package_on_page(
    page: Page,
    product: dict[str, Any],
    asset_type: str,
    downloads_dir: Path,
    timeout_ms: int,
) -> InspectedAssetPackage:
    unavailable_error = ""
    try:
        package_path = download_store_insight_asset_package_on_page(
            page,
            product,
            asset_type,
            downloads_dir,
            timeout_ms,
            strict_download=True,
        )
    except StoreInsightAssetUnavailable as error:
        package_path = None
        unavailable_error = str(error)
    if package_path is None:
        return InspectedAssetPackage(
            asset_type=asset_type,
            product_id=str(product["product_id"]),
            package_path=None,
            extract_dir=downloads_dir / f"{product['product_id']}_{asset_type}_extracted",
            files=(),
            error=unavailable_error or "Store Insight download did not produce a file",
        )

    extract_dir = downloads_dir / f"{product['product_id']}_{asset_type}_extracted"
    extracted = image_files_from_download(package_path, extract_dir)
    extracted = filter_store_insight_asset_files(extracted, extract_dir, asset_type)
    return InspectedAssetPackage(
        asset_type=asset_type,
        product_id=str(product["product_id"]),
        package_path=package_path,
        extract_dir=extract_dir,
        files=tuple(extracted),
        error="" if extracted else "download file contained no usable images",
    )


def materialize_store_insight_asset_package(
    inspected: InspectedAssetPackage,
    output_root: Path,
    public_url_prefix: str,
    max_images: int | None = None,
    sku_table_rows: list[dict[str, str]] | None = None,
) -> dict[str, Any]:
    images: list[dict[str, Any]] = []
    ok_count = 0
    sku_table_rows = list(sku_table_rows or [])
    matched_table_rows: set[str] = set()
    sources = list(inspected.files)
    if inspected.asset_type == "sku":
        sources.sort(
            key=lambda source: (
                store_insight_sku_source_index(source.name) is None,
                store_insight_sku_source_index(source.name) or 0,
                source.name,
            )
        )
    for index, source in enumerate(sources, start=1):
        if max_images is not None and ok_count >= max_images:
            break
        sku_metadata = (
            parse_store_insight_sku_filename(source.name, index)
            if inspected.asset_type == "sku"
            else {}
        )
        sku_metadata.pop("index", None)
        source_index = (
            store_insight_sku_source_index(source.name)
            if inspected.asset_type == "sku"
            else None
        )
        filename = f"{index:02d}.jpg"
        result_path = store_insight_asset_result_path(
            inspected.asset_type,
            inspected.product_id,
            filename,
        )
        destination = output_root / result_path
        try:
            sha256, width, height = normalize_local_image(source, destination)
        except Exception as error:  # noqa: BLE001 - keep per-file diagnostics
            if inspected.asset_type == "sku":
                destination.unlink(missing_ok=True)
            images.append(
                {
                    "index": index,
                    "status": "failed",
                    "source_path": str(source),
                    "local_path": str(destination),
                    "result_path": result_path,
                    "error": f"{type(error).__name__}: {error}"[:300],
                }
            )
            continue
        image_entry = {
            "index": index,
            "status": "ok",
            "source_path": str(source),
            "local_path": str(destination),
            "result_path": result_path,
            "public_url": public_url_for(destination, output_root, public_url_prefix),
            "sha256": sha256,
            "width": width,
            "height": height,
        }
        if inspected.asset_type == "sku":
            table_row = sku_table_metadata_for_image(
                source,
                source_index,
                index,
                sku_table_rows,
                matched_table_rows,
            )
            if table_row is not None:
                matched_table_rows.add(table_row.get("source_index", ""))
                if table_row.get("sku_label"):
                    sku_metadata["sku_label"] = table_row["sku_label"]
                    table_metadata = parse_store_insight_sku_filename(
                        f"SKU图_{source_index or index}_{table_row['sku_label']}.jpg",
                        index,
                    )
                    for field in ("spec_text", "color_text", "parse_status"):
                        if table_metadata.get(field):
                            sku_metadata[field] = table_metadata[field]
            image_entry.update(
                {
                    "product_id": table_row.get("product_id", inspected.product_id)
                    if table_row is not None
                    else inspected.product_id,
                    "source_index": source_index,
                    **sku_metadata,
                    **sku_table_fields(table_row),
                    "metadata_status": "table_matched" if table_row is not None else "image_only",
                }
            )
        images.append(image_entry)
        ok_count += 1
    if inspected.asset_type == "sku":
        for table_row in sku_table_rows:
            source_row = table_row.get("source_index", "")
            if source_row in matched_table_rows:
                continue
            source_index = int(source_row) if source_row.isdigit() else None
            parsed = parse_store_insight_sku_filename(
                f"SKU\u56fe_{source_row}_{table_row.get('sku_label', '')}.jpg",
                source_index or len(images) + 1,
            )
            images.append(
                {
                    "index": len(images) + 1,
                    "status": "metadata_only",
                    "source_path": "",
                    "local_path": "",
                    "result_path": "",
                    "public_url": "",
                    "product_id": table_row.get("product_id", inspected.product_id),
                    "source_index": source_index,
                    **parsed,
                    **sku_table_fields(table_row),
                    "metadata_status": "metadata_only",
                }
            )

    return {
        "status": "ok" if ok_count else "failed",
        "asset_type": inspected.asset_type,
        "image_count": ok_count,
        "download_path": str(inspected.package_path or ""),
        "images": images,
        "error": "" if ok_count else inspected.error or "download file contained no usable images",
    }


def collect_store_insight_asset_type(
    context: BrowserContext,
    product: dict[str, Any],
    asset_type: str,
    downloads_dir: Path,
    assets_root: Path,
    output_root: Path,
    public_url_prefix: str,
    timeout_ms: int,
    max_images: int | None = None,
) -> dict[str, Any]:
    sku_table_rows: list[dict[str, str]] = []
    if asset_type == "sku" and context is not None:
        inspected, sku_table_rows = inspect_store_insight_sku_exports(
            context,
            product,
            downloads_dir,
            timeout_ms,
        )
        if not inspected.files:
            fallback = inspect_store_insight_asset_package(
                context,
                product,
                asset_type,
                downloads_dir,
                timeout_ms,
            )
            if fallback.files:
                inspected = fallback
    else:
        inspected = inspect_store_insight_asset_package(
            context,
            product,
            asset_type,
            downloads_dir,
            timeout_ms,
        )
    materialize_kwargs: dict[str, Any] = {"max_images": max_images}
    if sku_table_rows:
        materialize_kwargs["sku_table_rows"] = sku_table_rows
    return materialize_store_insight_asset_package(
        inspected,
        output_root,
        public_url_prefix,
        **materialize_kwargs,
    )


def collect_store_insight_assets_for_product(
    context: BrowserContext,
    product: dict[str, Any],
    asset_targets: dict[str, int | None],
    downloads_dir: Path,
    assets_root: Path,
    output_root: Path,
    public_url_prefix: str,
    timeout_ms: int,
) -> dict[str, Any]:
    type_results: dict[str, Any] = {}
    for asset_type, max_images in asset_targets.items():
        type_results[asset_type] = collect_store_insight_asset_type(
            context,
            product,
            asset_type,
            downloads_dir / asset_type,
            assets_root,
            output_root,
            public_url_prefix,
            timeout_ms,
            max_images=max_images,
        )
    ok_types = sum(1 for item in type_results.values() if item.get("status") == "ok")
    return {
        "product_id": str(product["product_id"]),
        "rank": int(product.get("rank") or 0),
        "title": str(product.get("title") or ""),
        "item_url": str(product.get("item_url") or ""),
        "sales_text": str(product.get("sales_text") or ""),
        "status": "ok" if ok_types == len(asset_targets) else "partial" if ok_types else "failed",
        "types": type_results,
    }


def merge_asset_record(records: list[dict[str, Any]], incoming: dict[str, Any]) -> dict[str, Any]:
    product_id = str(incoming.get("product_id") or "")
    existing = next(
        (record for record in records if str(record.get("product_id") or "") == product_id),
        None,
    )
    if existing is None:
        records.append(incoming)
        existing = incoming
    else:
        existing.setdefault("types", {}).update(incoming.get("types") or {})

    type_results = list((existing.get("types") or {}).values())
    ok_types = sum(1 for item in type_results if item.get("status") == "ok")
    existing["status"] = (
        "ok" if type_results and ok_types == len(type_results) else "partial" if ok_types else "failed"
    )
    return existing


def asset_counts_from_records(records: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for product in records:
        for asset_type, type_entry in (product.get("types") or {}).items():
            count = sum(
                1 for image in (type_entry.get("images") or []) if image.get("status") == "ok"
            )
            counts[asset_type] = counts.get(asset_type, 0) + count
    return counts


def collect_dynamic_store_insight_assets(
    context: BrowserContext,
    ranked_products: list[dict[str, Any]],
    downloads_dir: Path,
    output_root: Path,
    public_url_prefix: str,
    timeout_ms: int,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    counts = {asset_type: 0 for asset_type in DYNAMIC_ASSET_RULES}
    source_ids: dict[str, list[str]] = {asset_type: [] for asset_type in DYNAMIC_ASSET_RULES}
    records: list[dict[str, Any]] = []
    for rank, product in enumerate(ranked_products, start=1):
        pending_types = [
            asset_type
            for asset_type, rule in DYNAMIC_ASSET_RULES.items()
            if counts[asset_type] < int(rule.get("target", rule["minimum"]))
        ]
        if not pending_types:
            break
        product_for_assets = {**product, "rank": rank}
        print(
            f"[collector] checking dynamic assets for {product['product_id']}: "
            f"{','.join(pending_types)}",
            flush=True,
        )
        page = context.new_page()
        try:
            page.goto(product["item_url"], wait_until="domcontentloaded", timeout=timeout_ms)
            page.wait_for_timeout(3_000)
            detect_stop(page)
            for asset_type in pending_types:
                rule = DYNAMIC_ASSET_RULES[asset_type]
                sku_table_rows: list[dict[str, str]] = []
                if asset_type == "sku" and hasattr(page, "context"):
                    inspected, sku_table_rows = inspect_store_insight_sku_exports_on_page(
                        page,
                        product_for_assets,
                        downloads_dir / asset_type,
                        timeout_ms,
                    )
                    if not inspected.files:
                        inspected = inspect_store_insight_asset_package_on_page(
                            page,
                            product_for_assets,
                            asset_type,
                            downloads_dir / asset_type,
                            timeout_ms,
                        )
                else:
                    inspected = inspect_store_insight_asset_package_on_page(
                        page,
                        product_for_assets,
                        asset_type,
                        downloads_dir / asset_type,
                        timeout_ms,
                    )
                remaining_capacity = int(rule["maximum"]) - counts[asset_type]
                target_count = int(rule.get("target", rule["minimum"]))
                materialize_kwargs: dict[str, Any] = {"max_images": remaining_capacity}
                if sku_table_rows:
                    materialize_kwargs["sku_table_rows"] = sku_table_rows
                type_entry = materialize_store_insight_asset_package(
                    inspected,
                    output_root,
                    public_url_prefix,
                    **materialize_kwargs,
                )
                added_count = int(type_entry.get("image_count") or 0)
                counts[asset_type] += added_count
                product_id = str(product["product_id"])
                if added_count > 0 and product_id not in source_ids[asset_type]:
                    source_ids[asset_type].append(product_id)
                type_entry["minimum"] = int(rule["minimum"])
                type_entry["maximum"] = int(rule["maximum"])
                type_entry["target"] = target_count
                type_entry["source_valid_count"] = inspected.valid_count
                type_entry["meets_minimum"] = counts[asset_type] >= int(rule["minimum"])
                merge_asset_record(
                    records,
                    {
                        "product_id": product_id,
                        "rank": rank,
                        "title": str(product.get("title") or ""),
                        "item_url": str(product.get("item_url") or ""),
                        "sales_text": str(product.get("sales_text") or ""),
                        "status": type_entry.get("status", "failed"),
                        "types": {asset_type: type_entry},
                    },
                )
                print(
                    f"[collector] dynamic assets {product_id} {asset_type}: "
                    f"package={inspected.valid_count} added={added_count} "
                    f"total={counts[asset_type]}/{target_count} "
                    f"cap={rule['maximum']}",
                    flush=True,
                )
        except CollectorPaused:
            raise
        except Exception as error:
            detect_stop(page)
            raise CollectorPaused(
                f"Dynamic asset inspection failed for {product['product_id']}: "
                f"{type(error).__name__}: {error}"
            ) from error
        finally:
            page.close()

    metadata: dict[str, Any] = {
        "asset_count_mode": "dynamic",
        "dynamic_asset_minimums": {
            asset_type: int(rule["minimum"])
            for asset_type, rule in DYNAMIC_ASSET_RULES.items()
        },
        "dynamic_asset_maximums": {
            asset_type: int(rule["maximum"])
            for asset_type, rule in DYNAMIC_ASSET_RULES.items()
        },
        "dynamic_asset_targets": {
            asset_type: int(rule.get("target", rule["minimum"]))
            for asset_type, rule in DYNAMIC_ASSET_RULES.items()
        },
        "missing_asset_types": [],
    }
    for asset_type, rule in DYNAMIC_ASSET_RULES.items():
        images = [
            image
            for record in records
            for image in (((record.get("types") or {}).get(asset_type) or {}).get("images") or [])
            if image.get("status") == "ok"
            or (asset_type == "sku" and image.get("status") == "metadata_only")
        ]
        image_count = sum(1 for image in images if image.get("status") == "ok")
        meets_minimum = image_count >= int(rule["minimum"])
        metadata[f"{asset_type}_assets"] = images
        metadata[f"{asset_type}_image_count"] = image_count
        metadata[f"{asset_type}_source_product_id"] = (
            source_ids[asset_type][0] if source_ids[asset_type] else ""
        )
        metadata[f"{asset_type}_source_product_ids"] = source_ids[asset_type]
        metadata[f"{asset_type}_meets_minimum"] = meets_minimum
        if not images:
            metadata["missing_asset_types"].append(asset_type)

    pending_types = [
        asset_type
        for asset_type, rule in DYNAMIC_ASSET_RULES.items()
        if counts[asset_type] < int(rule.get("target", rule["minimum"]))
    ]
    print(
        f"[collector] dynamic asset collection finished: "
        f"main={counts['main']} sku={counts['sku']} detail={counts['detail']} "
        f"pending={','.join(pending_types) or 'none'}",
        flush=True,
    )

    return records, metadata


def run_collection(config: CollectorConfig) -> dict[str, Any]:
    config.output_dir.mkdir(parents=True, exist_ok=True)
    work_dir = config.output_dir / "_work"
    images_dir = work_dir / "main-images"
    downloads_dir = work_dir / "store-insight-downloads"
    assets_root = work_dir / "competitor-assets"
    asset_manifest_path = work_dir / "extended-competitor-assets.json"
    manifest_path = config.output_dir / "main-image-manifest.json"
    store_insight_asset_targets = build_store_insight_asset_targets(config)

    cdp_url = ensure_cdp_browser(config)
    with sync_playwright() as playwright:
        _browser, context, page = connect_over_cdp(playwright, cdp_url)
        owned_start_page: Page | None = None
        try:
            if config.use_current_page and config.reference_image is None:
                close_common_popups(page)
                page = wait_for_login(
                    context,
                    page,
                    config.login_timeout_seconds,
                    config.login_poll_seconds,
                )
                result_page = wait_for_result_page(
                    context,
                    page,
                    config.login_timeout_seconds,
                    config.login_poll_seconds,
                )
            else:
                close_stale_result_pages(context)
                page = context.new_page()
                owned_start_page = page
                page.goto(TAOBAO_HOME, wait_until="domcontentloaded", timeout=30_000)
                result_page = perform_image_search(
                    page,
                    context,
                    config.reference_image,
                    config.login_timeout_seconds,
                    config.login_poll_seconds,
                )

            print("[collector] switching to sales sort", flush=True)
            switch_to_sales_sort(result_page)
            sort_state = read_sales_sort_state(result_page)
            if config.head_wait_seconds > 0:
                result_page.wait_for_timeout(int(config.head_wait_seconds * 1000))
            close_common_popups(result_page)
            print("[collector] collecting result cards", flush=True)
            rows = collect_candidate_rows(result_page, config.candidate_limit, config.scroll_rounds)
            ranked_products = rows_to_products(rows, config.min_sales_floor)
            ranked_for_assets = ranked_products_for_asset_collection(
                config,
                ranked_products,
                store_insight_asset_targets,
            )
            asset_products = ranked_for_assets[:MAX_ASSET_PRODUCT_PROBES]
            products_to_process = (
                asset_products if store_insight_asset_targets else ranked_products[: config.max_count]
            )
            print(
                f"[collector] collected rows={len(rows)} products={len(ranked_products)}",
                flush=True,
            )

            records: list[ProductRecord] = []
            asset_records: list[dict[str, Any]] = []
            asset_metadata: dict[str, Any] = {"asset_count_mode": config.asset_count_mode}

            # 采集第一款商品的参数
            from parameter_collector import collect_first_product_parameters
            parameter_metadata = collect_first_product_parameters(
                context,
                asset_products,
                config.detail_timeout_ms,
            )
            asset_metadata.update(parameter_metadata)

            if config.asset_count_mode == "dynamic":
                asset_records, dynamic_metadata = collect_dynamic_store_insight_assets(
                    context,
                    asset_products,
                    downloads_dir,
                    config.output_dir,
                    config.public_url_prefix,
                    config.detail_timeout_ms,
                )
                asset_metadata.update(dynamic_metadata)
            if config.collect_main_video:
                asset_metadata.update(
                    collect_first_store_insight_main_video(
                        context,
                        ranked_products,
                        downloads_dir / "main_video",
                        config.output_dir,
                        config.detail_timeout_ms,
                    )
                )
            else:
                asset_metadata.update(
                    {
                        "main_video_requested": False,
                        "main_video_url": "",
                        "main_video_result_path": "",
                        "main_video_count": 0,
                        "main_video_source_product_id": "",
                        "main_video_status": "disabled",
                    }
                )
            if asset_records or config.collect_main_video or config.asset_count_mode == "dynamic":
                write_extended_asset_manifest(
                    asset_manifest_path,
                    config,
                    asset_records,
                    asset_metadata,
                )
            asset_collected = asset_counts_from_records(asset_records)
            for asset_type in store_insight_asset_targets:
                asset_collected.setdefault(asset_type, 0)
            for rank, product in enumerate(products_to_process, start=1):
                collect_product_record = rank <= config.max_count
                remaining_asset_targets = remaining_bounded_store_insight_asset_targets(
                    store_insight_asset_targets,
                    asset_collected,
                    rank,
                )
                if not collect_product_record and not remaining_asset_targets:
                    break

                destination = images_dir / f"{rank:02d}_{product['product_id']}.jpg"
                status = "ok"
                error = ""
                sha256 = ""
                width = 0
                height = 0
                source = "search_card"
                product_for_assets = {**product, "rank": rank}
                asset_entry: dict[str, Any] | None = None

                try:
                    print(
                        f"[collector] processing product {rank}/{len(products_to_process)}: "
                        f"{product['product_id']} {product['sales_text']}",
                        flush=True,
                    )
                    if remaining_asset_targets:
                        print(
                            f"[collector] collecting Store Insight assets for {product['product_id']}: "
                            f"{','.join(remaining_asset_targets)}",
                            flush=True,
                        )
                        asset_entry = collect_store_insight_assets_for_product(
                            context,
                            product_for_assets,
                            remaining_asset_targets,
                            downloads_dir,
                            assets_root,
                            config.output_dir,
                            config.public_url_prefix,
                            config.detail_timeout_ms,
                        )
                        merge_asset_record(asset_records, asset_entry)
                        for asset_type, type_entry in (asset_entry.get("types") or {}).items():
                            asset_collected[asset_type] = asset_collected.get(asset_type, 0) + int(
                                type_entry.get("image_count") or 0
                            )
                        write_extended_asset_manifest(
                            asset_manifest_path,
                            config,
                            asset_records,
                            asset_metadata,
                        )
                        if collect_product_record:
                            main_asset = first_asset_image(asset_entry, "main")
                            if main_asset and main_asset.get("status") == "ok":
                                sha256, width, height = normalize_local_image(
                                    Path(str(main_asset["local_path"])),
                                    destination,
                                )
                                source = "store_insight_asset_main"
                    if (
                        collect_product_record
                        and config.store_insight_download
                        and config.asset_count_mode != "dynamic"
                    ):
                        if not sha256:
                            downloaded = try_store_insight_download(
                                context,
                                product,
                                downloads_dir,
                                config.detail_timeout_ms,
                            )
                            if downloaded:
                                extracted = image_files_from_download(
                                    downloaded,
                                    downloads_dir / f"{product['product_id']}_extracted",
                                )
                                if extracted:
                                    sha256, width, height = normalize_local_image(extracted[0], destination)
                                    source = "store_insight_product_image_download"
                    if collect_product_record and not sha256:
                        sha256, width, height = download_image_from_url(
                            context,
                            product["source_image_url"],
                            destination,
                        )
                except CollectorPaused:
                    raise
                except Exception as exc:  # noqa: BLE001 - manifest records individual failures
                    status = "failed"
                    error = str(exc)[:300]

                if collect_product_record:
                    records.append(
                        ProductRecord(
                            rank=rank,
                            source_order=int(product["source_order"]),
                            product_id=str(product["product_id"]),
                            title=str(product["title"]),
                            sales_text=str(product["sales_text"]),
                            sales_floor=product["sales_floor"],
                            item_url=str(product["item_url"]),
                            source_card_url=str(product["source_card_url"]),
                            source_image_url=str(product["source_image_url"]),
                            main_image_local_path=str(destination),
                            main_image_public_url=public_url_for(
                                destination,
                                config.output_dir,
                                config.public_url_prefix,
                            )
                            if status == "ok"
                            else "",
                            sha256=sha256,
                            width=width,
                            height=height,
                            download_source=source,
                            status=status,
                            error=error,
                        )
                    )

            ok_count = sum(1 for item in records if item.status == "ok")
            title_analysis_input = [
                {
                    "rank": item.rank,
                    "title": item.title,
                    "sales_text": item.sales_text,
                    "sales_floor": item.sales_floor,
                    "item_url": item.item_url,
                    "image_status": item.status,
                }
                for item in records
                if item.title
            ]
            title_records = [
                {
                    "rank": item.rank,
                    "title": item.title,
                }
                for item in records
                if item.title
            ]
            titles_txt_path = config.output_dir / "titles.txt"
            titles_json_path = config.output_dir / "titles.json"
            titles_txt_path.write_text(
                "\n".join(f"{item['rank']:02d}. {item['title']}" for item in title_records) + "\n",
                encoding="utf-8-sig",
            )
            titles_json_path.write_text(
                json.dumps(title_records, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            asset_collected = asset_counts_from_records(asset_records)
            asset_shortfalls = store_insight_asset_shortfalls(
                store_insight_asset_targets,
                asset_collected,
            )
            dynamic_partial = config.asset_count_mode == "dynamic" and any(
                not bool(asset_metadata.get(f"{asset_type}_meets_minimum"))
                for asset_type in DYNAMIC_ASSET_RULES
            )
            video_partial = config.collect_main_video and asset_metadata.get("main_video_status") != "complete"
            manifest = {
                "schema_version": 1,
                "status": (
                    "complete"
                    if (
                        ok_count >= config.max_count
                        and not asset_shortfalls
                        and not dynamic_partial
                        and not video_partial
                    )
                    else "partial"
                ),
                "created_at": utc_now_iso(),
                "reference_image": str(config.reference_image) if config.reference_image else "",
                "cdp_url": config.cdp_url,
                "result_page_url": result_page.url,
                "sort_mode": "native_sales_verified",
                "sort_state": sort_state,
                "requested_count": config.max_count,
                "candidate_count": len(ranked_products),
                "image_count": ok_count,
                "title_count": len(title_records),
                "titles_txt_path": str(titles_txt_path),
                "titles_json_path": str(titles_json_path),
                "products": [asdict(item) for item in records],
                "title_analysis_input": title_analysis_input,
                "asset_count_mode": config.asset_count_mode,
            }
            manifest.update(asset_metadata)
            if store_insight_asset_targets:
                manifest["asset_targets"] = store_insight_asset_targets
                manifest["asset_collected"] = asset_collected
                manifest["asset_shortfalls"] = asset_shortfalls
            if asset_records or config.collect_main_video or config.asset_count_mode == "dynamic":
                asset_count = sum(asset_collected.values())
                manifest["asset_count"] = asset_count
                manifest["asset_manifest_path"] = str(asset_manifest_path)
                manifest["extended_assets"] = asset_records
                write_extended_asset_manifest(
                    asset_manifest_path,
                    config,
                    asset_records,
                    asset_metadata,
                )
            manifest_path.write_text(
                json.dumps(manifest, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            if owned_start_page and owned_start_page != result_page and not owned_start_page.is_closed():
                try:
                    owned_start_page.close()
                except Exception:
                    pass
            close_result_page_after_collection(context, result_page)
            return manifest
        finally:
            # Leaving the sync_playwright context stops the client transport.
            # Do not close the remote browser; it belongs to the user/session.
            pass


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Collect top Taobao same-item main images from a visible CDP browser.",
    )
    parser.add_argument("--reference-image", type=Path, help="Local image used for Taobao image search.")
    parser.add_argument("--output-dir", type=Path, required=True, help="Directory for manifest and images.")
    parser.add_argument("--cdp-url", default=DEFAULT_CDP_URL, help="Edge/Chrome CDP URL.")
    parser.add_argument("--max-count", type=int, default=10, help="Number of products to keep.")
    parser.add_argument("--candidate-limit", type=int, default=8, help="Visible card scan limit.")
    parser.add_argument("--min-sales-floor", type=int, default=0, help="Drop cards below this sales floor.")
    parser.add_argument(
        "--use-current-page",
        action="store_true",
        help="Without --reference-image, wait for and parse an already-open result page.",
    )
    parser.add_argument("--public-url-prefix", default="", help="Optional public URL prefix for output-dir.")
    parser.add_argument(
        "--store-insight-download",
        action="store_true",
        help="Try Store Insight product-image custom download before card image fallback.",
    )
    parser.add_argument(
        "--collect-store-insight-assets",
        action="store_true",
        help="Collect Store Insight main/SKU/detail assets into _work/competitor-assets.",
    )
    parser.add_argument(
        "--store-insight-asset-types",
        default=",".join(DEFAULT_STORE_INSIGHT_ASSET_TYPES),
        help="Comma-separated asset types: main,sku,detail,detail_long.",
    )
    parser.add_argument("--collect-main-image-assets", action="store_true")
    parser.add_argument("--main-image-asset-count", type=int, default=0)
    parser.add_argument("--collect-sku-image-assets", action="store_true")
    parser.add_argument("--sku-image-asset-count", type=int, default=0)
    parser.add_argument("--collect-detail-image-assets", action="store_true")
    parser.add_argument("--detail-image-asset-count", type=int, default=0)
    parser.add_argument("--collect-detail-long-image-assets", action="store_true")
    parser.add_argument("--detail-long-image-asset-count", type=int, default=0)
    parser.add_argument(
        "--no-auto-launch-browser",
        action="store_true",
        help="Do not start Edge/Chrome automatically when the CDP port is unavailable.",
    )
    parser.add_argument(
        "--reuse-existing-cdp",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Reuse an already-running browser on the configured CDP endpoint.",
    )
    parser.add_argument("--browser-executable", default="", help="Optional Edge/Chrome executable path.")
    parser.add_argument(
        "--browser-profile-dir",
        type=Path,
        default=DEFAULT_BROWSER_PROFILE_DIR,
        help="Profile directory used when launching a browser automatically.",
    )
    parser.add_argument(
        "--login-timeout-seconds",
        type=int,
        default=600,
        help="Maximum time to wait for manual Taobao login.",
    )
    parser.add_argument(
        "--login-poll-seconds",
        type=int,
        default=5,
        help="Login status polling interval.",
    )
    parser.add_argument("--scroll-rounds", type=int, default=1, help="Scroll rounds for loading cards.")
    parser.add_argument(
        "--collect-main-video",
        action="store_true",
        help="Collect the first available original main video from sales-ranked products.",
    )
    parser.add_argument(
        "--top-product-only",
        action="store_true",
        help="Use only the highest-sales result for unbounded assets; explicit targets and video may fall through.",
    )
    parser.add_argument("--json-stdout", action="store_true", help="Print the full manifest JSON to stdout.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv if argv is not None else sys.argv[1:])
    config = CollectorConfig(
        reference_image=args.reference_image,
        output_dir=args.output_dir,
        cdp_url=args.cdp_url,
        max_count=args.max_count,
        candidate_limit=args.candidate_limit,
        min_sales_floor=args.min_sales_floor,
        use_current_page=args.use_current_page,
        public_url_prefix=args.public_url_prefix,
        store_insight_download=args.store_insight_download,
        collect_store_insight_assets=args.collect_store_insight_assets,
        store_insight_asset_types=normalize_store_insight_asset_types(args.store_insight_asset_types),
        collect_main_image_assets=args.collect_main_image_assets,
        main_image_asset_count=args.main_image_asset_count,
        collect_sku_image_assets=args.collect_sku_image_assets,
        sku_image_asset_count=args.sku_image_asset_count,
        collect_detail_image_assets=args.collect_detail_image_assets,
        detail_image_asset_count=args.detail_image_asset_count,
        collect_detail_long_image_assets=args.collect_detail_long_image_assets,
        detail_long_image_asset_count=args.detail_long_image_asset_count,
        collect_main_video=args.collect_main_video,
        scroll_rounds=args.scroll_rounds,
        auto_launch_browser=not args.no_auto_launch_browser,
        reuse_existing_cdp=args.reuse_existing_cdp,
        browser_executable=args.browser_executable,
        browser_profile_dir=args.browser_profile_dir,
        login_timeout_seconds=args.login_timeout_seconds,
        login_poll_seconds=args.login_poll_seconds,
        top_product_only=args.top_product_only,
    )
    try:
        manifest = run_collection(config)
    except CollectorPaused as exc:
        payload = {
            "schema_version": 1,
            "status": "paused",
            "created_at": utc_now_iso(),
            "message": str(exc),
        }
        config.output_dir.mkdir(parents=True, exist_ok=True)
        (config.output_dir / COLLECTION_PAUSED_FILENAME).write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(json.dumps(payload, ensure_ascii=False) if args.json_stdout else str(exc))
        return 2
    except Exception as exc:  # noqa: BLE001 - CLI must report controlled JSON
        payload = {
            "schema_version": 1,
            "status": "failed",
            "created_at": utc_now_iso(),
            "message": f"{type(exc).__name__}: {exc}",
        }
        config.output_dir.mkdir(parents=True, exist_ok=True)
        (config.output_dir / COLLECTION_FAILED_FILENAME).write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(json.dumps(payload, ensure_ascii=False) if args.json_stdout else payload["message"])
        return 1

    if args.json_stdout:
        print(json.dumps(manifest, ensure_ascii=False))
    else:
        print(f"Done: {manifest['image_count']} images -> {args.output_dir / 'main-image-manifest.json'}")
    return 0


if __name__ == "__main__":
    # Parameter collection imports this module by name; keep exception classes identical in script/EXE mode.
    sys.modules.setdefault("same_item_collector", sys.modules[__name__])
    raise SystemExit(main())
