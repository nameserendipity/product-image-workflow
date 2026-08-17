import json
import tempfile
import unittest
import shutil
from types import SimpleNamespace
from pathlib import Path
from unittest.mock import Mock, patch

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image

from batch_workflow import (
    BatchCollectionPaused,
    BatchItem,
    BatchRunner,
    DirectLinkBatchItem,
    DirectReplaceBatchItem,
    DirectLinkCollector,
    resolve_supplement_workbook,
    ensure_workbook_available,
    build_generation_manifest,
    export_product_workbook,
    plan_supplement_ordinals,
    plan_all_supplement_ordinals,
    extract_batch_items,
    extract_direct_link_items,
    extract_direct_replace_items,
    find_prior_collected_manifest,
    load_batch_results,
    normalize_direct_manifest,
    restore_collected_manifest,
    resolve_direct_item_url,
    save_batch_results,
    summarize_generation_result,
    merge_manual_sku_metadata,
    add_manual_sku_generation_sources,
    _direct_link_platform,
    _collected_manifest_is_valid,
    _sku_rows,
)
from image_workflows import ApiSettings, ProductTitleClient, validate_listing_titles
from shared_library_client import LockLease, SharedLibraryUnavailable, SharedProbe
from shared_package_builder import ReusedPackage


class BatchWorkflowTests(unittest.TestCase):
    def test_direct_link_shared_hit_skips_collector_and_generator(self) -> None:
        workbook = self.root / "shared-hit.xlsx"
        workbook.write_bytes(b"workbook")
        output = self.root / "shared-hit-output"
        item = DirectLinkBatchItem(
            1,
            2,
            "https://item.taobao.com/item.htm?id=123",
            "taobao",
            "共享商品",
        )
        reused_root = self.root / "materialized"
        source_manifest = reused_root / "source" / "manifest.json"
        generated = reused_root / "generated" / "main" / "001.png"
        packaged_workbook = reused_root / "result.xlsx"
        source_manifest.parent.mkdir(parents=True)
        generated.parent.mkdir(parents=True)
        source_manifest.write_text(
            json.dumps({"images": [{"type": "main", "path": str(self.image)}]}),
            encoding="utf-8",
        )
        generated.write_bytes(b"generated")
        packaged_workbook.write_bytes(b"xlsx")
        reused = ReusedPackage(
            reused_root,
            source_manifest,
            [{"category": "main", "ordinal": 1, "status": "completed", "output_path": str(generated)}],
            {"long_title": "共享商品长标题", "short_title": "共享商品短标题"},
            packaged_workbook,
        )
        package_zip = self.root / "complete-package.zip"
        package_zip.write_bytes(b"package")
        client = Mock()
        client.probe.return_value = SharedProbe(
            "available",
            {
                "product_key": "taobao-123",
                "package_object": "product-workflow/shared-library/packages/taobao/123/complete-package.zip",
                "package_size": 7,
                "package_sha256": "a" * 64,
            },
            None,
        )
        client.download.return_value = package_zip
        cache = Mock()
        runner = BatchRunner(
            None,
            self.root,
            self.root,
            batch_mode="direct_link",
            shared_library=client,
            shared_cache=cache,
        )
        runner.direct_collector.collect = Mock()

        def export(path, *_args, **_kwargs):
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(b"xlsx")
            return path

        with (
            patch("batch_workflow.extract_direct_link_items", return_value=[item]),
            patch("batch_workflow.materialize_reused_package", return_value=reused),
            patch("batch_workflow.export_product_workbook", side_effect=export),
            patch("batch_workflow.WorkflowRunner") as workflow_runner,
        ):
            results = runner.run(workbook, output)

        self.assertEqual(results[0]["status"], "completed")
        self.assertEqual(results[0]["shared_status"], "reused")
        self.assertTrue(Path(results[0]["workbook"]).is_file())
        runner.direct_collector.collect.assert_not_called()
        workflow_runner.assert_not_called()
        cache.record_download.assert_called_once()

    def test_direct_link_shared_lock_fails_only_locked_row(self) -> None:
        workbook = self.root / "shared-lock.xlsx"
        workbook.write_bytes(b"workbook")
        output = self.root / "shared-lock-output"
        first = DirectLinkBatchItem(1, 2, "https://item.taobao.com/item.htm?id=123", "taobao")
        second = DirectLinkBatchItem(
            2,
            3,
            "https://example.com/not-shared",
            "unsupported",
            validation_error="不支持的平台",
        )
        client = Mock()
        client.probe.return_value = SharedProbe(
            "locked",
            None,
            {"product_key": "taobao-123", "expires_at": "2026-08-15T12:00:00+00:00"},
        )
        runner = BatchRunner(
            None,
            self.root,
            self.root,
            batch_mode="direct_link",
            shared_library=client,
            shared_cache=Mock(),
        )

        with patch("batch_workflow.extract_direct_link_items", return_value=[first, second]):
            results = runner.run(workbook, output)

        self.assertEqual([result["status"] for result in results], ["failed", "failed"])
        self.assertEqual(results[0]["shared_status"], "locked")
        self.assertIn("其他用户正在生成", results[0]["error"])
        self.assertEqual(results[1]["error"], "不支持的平台")

    def test_direct_link_shared_unavailable_runs_local_row(self) -> None:
        workbook = self.root / "shared-fallback.xlsx"
        workbook.write_bytes(b"workbook")
        output = self.root / "shared-fallback-output"
        item = DirectLinkBatchItem(1, 2, "https://item.taobao.com/item.htm?id=123", "taobao", "本地商品")
        source_manifest = self._shared_local_manifest(output / "001-row-0002")
        records = self._shared_generation_records(1, 1, 1)
        client = Mock()
        client.probe.side_effect = SharedLibraryUnavailable("共享素材库暂时不可用")
        runner = BatchRunner(
            ApiSettings("https://api.example", "vision", "image"),
            self.root,
            self.root,
            batch_mode="direct_link",
            shared_library=client,
            shared_cache=Mock(),
        )
        exported = output / "001-row-0002" / "本地商品.xlsx"

        with (
            patch("batch_workflow.extract_direct_link_items", return_value=[item]),
            patch("batch_workflow.restore_collected_manifest", return_value=(source_manifest, 3)),
            patch("batch_workflow.ProductTitleClient.generate", return_value={"long_title": "本地商品长标题", "short_title": "本地商品短标题"}),
            patch("batch_workflow.WorkflowRunner.run", return_value=records),
            patch("batch_workflow.upload_generation_records", side_effect=lambda value, _uploader: value),
            patch("batch_workflow.export_product_workbook", return_value=exported),
        ):
            results = runner.run(workbook, output)

        self.assertEqual(results[0]["status"], "completed")
        self.assertEqual(results[0]["shared_status"], "local_fallback")
        client.acquire_lock.assert_not_called()
        client.publish.assert_not_called()

    def test_kuaishou_batch_persists_visual_parameters_before_export(self) -> None:
        workbook = self.root / "kuaishou.xlsx"
        workbook.write_bytes(b"workbook")
        output = self.root / "output"
        item = DirectLinkBatchItem(
            1,
            1,
            "https://app.kwaixiaodian.com/web/kwaishop-goods-detail-page-app?id=26065497098904",
            "kuaishou",
            "快手商品",
        )
        item_root = output / "001-row-0001"
        source_manifest = self._shared_local_manifest(item_root)

        def run_workflow(*_args, **_kwargs):
            dossier = item_root / "generated" / "product-dossier.json"
            dossier.parent.mkdir(parents=True, exist_ok=True)
            dossier.write_text(
                json.dumps(
                    {
                        "observations": [],
                        "dossier": {
                            "anchor_identity": {
                                "source_index": 1,
                                "object": "one pump shampoo bottle",
                                "visible_product_labeling": ["800ml"],
                            }
                        },
                    }
                ),
                encoding="utf-8",
            )
            return self._shared_generation_records(1, 1, 1)

        exported_source: dict[str, object] = {}

        def export(path, _item, manifest, *_args, **_kwargs):
            exported_source.update(json.loads(manifest.read_text(encoding="utf-8")))
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(b"xlsx")
            return path

        runner = BatchRunner(
            ApiSettings("https://api.example", "vision", "image"),
            self.root,
            self.root,
            batch_mode="direct_link",
        )
        with (
            patch("batch_workflow.extract_direct_link_items", return_value=[item]),
            patch("batch_workflow.restore_collected_manifest", return_value=(source_manifest, 3)),
            patch(
                "batch_workflow.ProductTitleClient.generate",
                return_value={"long_title": "快手商品长标题", "short_title": "快手商品"},
            ),
            patch("batch_workflow.WorkflowRunner.run", side_effect=run_workflow),
            patch("batch_workflow.upload_generation_records", side_effect=lambda rows, _uploader: rows),
            patch("batch_workflow.export_product_workbook", side_effect=export),
        ):
            result = runner.run(workbook, output)

        persisted_source = json.loads(source_manifest.read_text(encoding="utf-8"))
        self.assertEqual(result[0]["status"], "completed")
        self.assertEqual(exported_source["parameter_status"], "inferred")
        self.assertTrue(exported_source["product_parameters"])
        self.assertEqual(
            persisted_source["product_parameters"],
            exported_source["product_parameters"],
        )

    def test_direct_link_shared_miss_publishes_after_export_and_releases(self) -> None:
        workbook = self.root / "shared-publish.xlsx"
        workbook.write_bytes(b"workbook")
        output = self.root / "shared-publish-output"
        item = DirectLinkBatchItem(1, 2, "https://item.taobao.com/item.htm?id=123", "taobao", "发布商品")
        item_root = output / "001-row-0002"
        source_manifest = self._shared_local_manifest(item_root)
        records = self._shared_generation_records(10, 3, 6)
        lease = LockLease(
            "taobao-123",
            "task-1",
            "client-1",
            "etag-1",
            "2026-08-15T10:00:00+00:00",
            "2026-08-15T12:00:00+00:00",
        )
        client = Mock(root_prefix="product-workflow/shared-library", client_id="client-1")
        client.probe.return_value = SharedProbe("missing", None, None)
        client.acquire_lock.return_value = lease
        client.publish.return_value = {"product_key": "taobao-123"}
        package = Mock()
        package.to_publish_bundle.return_value = Mock()
        builder = Mock()
        builder.build.return_value = package
        runner = BatchRunner(
            ApiSettings("https://api.example", "vision", "image"),
            self.root,
            self.root,
            batch_mode="direct_link",
            shared_library=client,
            shared_cache=Mock(),
        )
        exported = item_root / "发布商品.xlsx"

        with (
            patch("batch_workflow.extract_direct_link_items", return_value=[item]),
            patch("batch_workflow.restore_collected_manifest", return_value=(source_manifest, 3)),
            patch("batch_workflow.ProductTitleClient.generate", return_value={"long_title": "发布商品长标题", "short_title": "发布商品短标题"}),
            patch("batch_workflow.WorkflowRunner.run", return_value=records),
            patch("batch_workflow.upload_generation_records", side_effect=lambda value, _uploader: value),
            patch("batch_workflow.export_product_workbook", return_value=exported) as export,
            patch("batch_workflow.SharedPackageBuilder", return_value=builder),
        ):
            results = runner.run(workbook, output)

        export.assert_called_once()
        client.publish.assert_called_once()
        client.release_lock.assert_called_once_with(lease)
        self.assertEqual(results[0]["shared_status"], "published")

    def test_non_direct_link_modes_do_not_query_shared_library(self) -> None:
        client = Mock()
        for mode, item in (
            ("direct_replace", DirectReplaceBatchItem(1, "商品", 2, self.image, "https://item.taobao.com/item.htm?id=123", "taobao")),
            ("image_search", BatchItem(1, 2, "商品", self.image, "", "", "")),
        ):
            with self.subTest(mode=mode):
                runner = BatchRunner(
                    None,
                    self.root,
                    self.root,
                    batch_mode=mode,
                    shared_library=client,
                    shared_cache=Mock(),
                )
                runner.cancel_event.set()
                extractor = (
                    "batch_workflow.extract_direct_replace_items"
                    if mode == "direct_replace"
                    else "batch_workflow.extract_batch_items"
                )
                with patch(extractor, return_value=[item]):
                    runner.run(self.root / "input.xlsx", self.root / f"output-{mode}")

        client.probe.assert_not_called()

    def test_collect_only_manifest_requires_assets_or_explicit_missing_types(self) -> None:
        empty = self.root / "empty-manifest.json"
        empty.write_text(json.dumps({"images": []}), encoding="utf-8")
        explicit_missing = self.root / "explicit-missing-manifest.json"
        explicit_missing.write_text(
            json.dumps(
                {
                    "images": [],
                    "requested_asset_types": ["main", "sku"],
                    "missing_asset_types": ["main", "sku"],
                }
            ),
            encoding="utf-8",
        )

        self.assertFalse(_collected_manifest_is_valid(empty))
        self.assertTrue(_collected_manifest_is_valid(explicit_missing))

    def test_all_supplement_ordinals_cover_only_missing_slots_in_each_category(self) -> None:
        main = self.root / "main.jpg"
        sku = self.root / "sku.jpg"
        detail = self.root / "detail.jpg"
        main.write_bytes(b"main")
        sku.write_bytes(b"sku")
        detail.write_bytes(b"detail")
        manifest = self.root / "manifest.json"
        manifest.write_text(
            json.dumps(
                {
                    "images": [
                        {"type": "main", "path": str(main)},
                        {"type": "sku", "path": str(sku)},
                        {"type": "detail", "path": str(detail)},
                    ]
                }
            ),
            encoding="utf-8",
        )
        existing = [
            {"category": "main", "ordinal": 1, "status": "completed", "output_path": str(main)},
            {"category": "main", "ordinal": 2, "status": "failed"},
            {"category": "sku", "ordinal": 1, "status": "completed", "output_path": str(sku)},
            {"category": "sku", "ordinal": 2, "status": "failed"},
            {"category": "detail", "ordinal": 1, "status": "completed", "output_path": str(detail)},
            {"category": "detail", "ordinal": 2, "status": "failed"},
        ]

        planned = plan_all_supplement_ordinals(existing, manifest)

        self.assertEqual(planned["main"], [2])
        self.assertEqual(planned["sku"], [2])
        self.assertEqual(planned["detail"], [2])

    def test_resolves_direct_link_supplement_from_exported_workbook(self) -> None:
        item_root = self.root / "store-insight" / "7005" / "run"
        main = item_root / "main" / "001.jpg"
        generated = item_root / "generated" / "main" / "001.png"
        main.parent.mkdir(parents=True)
        generated.parent.mkdir(parents=True)
        main.write_bytes(b"main")
        generated.write_bytes(b"generated")
        manifest = item_root / "manifest.json"
        manifest.write_text(
            json.dumps({"product_id": "7005", "images": [{"type": "main", "path": str(main)}]}),
            encoding="utf-8",
        )
        (item_root / "generated" / "analysis.json").write_text(
            json.dumps({"records": [{"category": "main", "ordinal": 1, "status": "completed", "source_path": str(main), "output_path": str(generated)}]}),
            encoding="utf-8",
        )
        workbook = Workbook()
        overview = workbook.active
        overview.title = "总览"
        overview.append(["字段", "值"])
        overview.append(["source_row", 1])
        overview.append(["source_title", "测试商品"])
        overview.append(["来源平台", "京东"])
        overview.append(["来源商品链接", "https://item.jd.com/7005.html"])
        main_sheet = workbook.create_sheet("主图")
        main_sheet.append(["序号", "采集图缩略图", "采集图路径", "生成图缩略图", "生成图路径"])
        main_sheet.append([1, "", "main/001.jpg", "", "generated/main/001.png"])
        exported = item_root / "测试商品.xlsx"
        workbook.save(exported)

        context = resolve_supplement_workbook(exported, self.root)

        self.assertIsInstance(context.item, DirectLinkBatchItem)
        self.assertEqual(context.source_manifest.resolve(), manifest.resolve())
        self.assertEqual(context.generated_root.resolve(), (item_root / "generated").resolve())
        self.assertEqual(context.generation_mode, "competitor_reference")

    def test_resolves_direct_replace_supplement_with_own_product_identity(self) -> None:
        item_root = self.root / "batch" / "001-row-0002"
        main = item_root / "collected" / "main" / "001.jpg"
        product = self.root / "batch" / "source-images" / "001-row-0002.png"
        main.parent.mkdir(parents=True)
        product.parent.mkdir(parents=True)
        main.write_bytes(b"main")
        product.write_bytes(b"product")
        manifest = item_root / "collected" / "direct-manifest.json"
        manifest.write_text(
            json.dumps(
                {
                    "product_id": "9001",
                    "images": [{"type": "main", "path": str(main)}],
                    "sku_variants": [
                        {
                            "sku_label": "白色大号",
                            "color_text": "白色",
                            "spec_text": "大号",
                            "list_price": "39.9",
                            "source_status": "text_conditioned",
                        }
                    ],
                }
            ),
            encoding="utf-8",
        )
        workbook = Workbook()
        overview = workbook.active
        overview.title = "总览"
        overview.append(["字段", "值"])
        overview.append(["source_row", 2])
        overview.append(["source_title", "替换商品"])
        overview.append(["来源平台", "抖音"])
        overview.append(["来源商品链接", "https://haohuo.jinritemai.com/views/product/item2.html?id=9001"])
        overview.append(["我方商品图", str(product)])
        exported = item_root / "替换商品.xlsx"
        workbook.save(exported)

        context = resolve_supplement_workbook(exported, self.root)

        self.assertIsInstance(context.item, DirectReplaceBatchItem)
        self.assertEqual(context.item.product_image.resolve(), product.resolve())
        self.assertEqual(context.item.platform, "douyin")
        self.assertEqual(context.item.manual_skus[0]["spec"], "大号")
        self.assertEqual(context.generation_mode, "own_product")

    def test_batch_with_no_generated_image_is_failed_and_not_exported(self) -> None:
        workbook = self.root / "input.xlsx"
        workbook.write_bytes(b"workbook")
        output = self.root / "output"
        item_root = output / "001-row-0002"
        manifest = item_root / "collected" / "main-image-manifest.json"
        manifest.parent.mkdir(parents=True)
        manifest.write_text(
            json.dumps(
                {
                    "products": [{"title": "测试商品"}],
                    "extended_assets": [
                        {"types": {"main": {"images": [{"status": "ok", "local_path": str(self.image)}]}}}
                    ],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        item = BatchItem(1, 2, "测试商品", self.image, "https://detail.1688.com/offer/1.html", 0, 0)
        runner = BatchRunner(None, self.root, self.root)

        with (
            patch("batch_workflow.extract_batch_items", return_value=[item]),
            patch("batch_workflow.ProductTitleClient.generate", return_value={"long_title": "标题", "short_title": "短标题"}),
            patch("batch_workflow.WorkflowRunner.run", return_value=[
                {"category": "main", "ordinal": 1, "status": "failed", "error": "视觉接口失败"}
            ]),
            patch("batch_workflow.export_product_workbook") as export,
        ):
            results = runner.run(workbook, output)

        self.assertEqual(results[0]["status"], "failed")
        self.assertIn("没有生成任何有效图片", results[0]["error"])
        export.assert_not_called()

    def test_generation_summary_marks_missing_or_failed_outputs_as_partial(self) -> None:
        records = [
            {"category": "main", "status": "completed", "output_path": str(self.image)},
            {"category": "sku", "status": "failed"},
            {"category": "detail", "status": "completed", "output_path": str(self.root / "missing.png")},
        ]

        generated, failed, missing = summarize_generation_result(
            records,
            ("main", "sku", "detail"),
        )

        self.assertEqual((generated, failed), (1, 2))
        self.assertEqual(missing, ("sku", "detail"))

    def test_resolves_read_only_unsized_overview_without_max_row(self) -> None:
        item_root = self.root / "store-insight" / "7006" / "run"
        main = item_root / "main" / "001.jpg"
        main.parent.mkdir(parents=True)
        main.write_bytes(b"main")
        manifest = item_root / "manifest.json"
        manifest.write_text(
            json.dumps({"product_id": "7006", "images": [{"type": "main", "path": str(main)}]}),
            encoding="utf-8",
        )
        workbook_path = item_root / "测试商品.xlsx"
        workbook_path.write_bytes(b"xlsx")

        class UnsizedOverview:
            max_row = None

            def iter_rows(self, *, min_row: int, values_only: bool):
                return iter(
                    [
                        ("source_title", "测试商品"),
                        ("来源平台", "京东"),
                        ("来源商品链接", "https://item.jd.com/7006.html"),
                    ]
                )

        class FakeWorkbook:
            sheetnames = ["总览"]

            def __getitem__(self, name: str):
                assert name == "总览"
                return UnsizedOverview()

            def close(self):
                return None

        workbook = FakeWorkbook()

        with patch("batch_workflow.load_workbook", return_value=workbook):
            context = resolve_supplement_workbook(workbook_path, self.root)

        self.assertIsInstance(context.item, DirectLinkBatchItem)
        self.assertEqual(context.item.source_url, "https://item.jd.com/7006.html")

    def test_generation_manifest_rebases_legacy_absolute_asset_paths(self) -> None:
        item_root = self.root / "001-row-0002"
        source_image = item_root / "collected" / "990604877206" / "run" / "sku" / "001.jpg"
        source_image.parent.mkdir(parents=True)
        source_image.write_bytes(b"sku")
        source_manifest = item_root / "collected" / "direct-manifest.json"
        source_manifest.write_text(
            json.dumps(
                {
                    "images": [
                        {
                            "type": "sku",
                            "path": r"C:\old-machine\outputs\batches\run\001-row-0002\collected\990604877206\run\sku\001.jpg",
                        }
                    ]
                }
            ),
            encoding="utf-8",
        )
        target = item_root / "generation-manifest.json"

        build_generation_manifest(source_manifest, target)

        generated = json.loads(target.read_text(encoding="utf-8"))
        self.assertEqual(generated["images"], [{"type": "sku", "path": str(source_image.resolve())}])

    def test_supplement_rejects_workbook_locked_by_excel_before_generation(self) -> None:
        workbook = self.root / "商品-001.xlsx"
        workbook.write_bytes(b"workbook")
        lock_file = workbook.with_name(f"~${workbook.name}")
        lock_file.write_bytes(b"lock")

        with self.assertRaisesRegex(RuntimeError, "Excel/WPS"):
            ensure_workbook_available(workbook)

    def test_supplement_ordinals_fill_failed_and_missing_outputs_before_appending(self) -> None:
        existing = [
            {"category": "main", "ordinal": 1, "status": "completed", "output_path": str(self.image)},
            {"category": "main", "ordinal": 2, "status": "failed"},
            {"category": "main", "ordinal": 3, "status": "completed", "output_path": str(self.root / "missing.png")},
            {"category": "main", "ordinal": 4, "status": "completed", "output_path": str(self.second_image)},
        ]

        ordinals = plan_supplement_ordinals(existing, "main", 4)

        self.assertEqual(ordinals, [2, 3, 5, 6])

    def test_supplement_ordinals_respect_sku_and_detail_limits(self) -> None:
        sku = [
            {"category": "sku", "ordinal": index, "status": "completed", "output_path": str(self.image)}
            for index in range(1, 8)
        ]
        detail = [
            {"category": "detail", "ordinal": index, "status": "completed", "output_path": str(self.image)}
            for index in range(1, 15)
        ]

        self.assertEqual(plan_supplement_ordinals(sku, "sku", 1), [8])
        self.assertEqual(plan_supplement_ordinals(detail, "detail", 1), [15])

    def test_batch_supplement_replaces_failed_slot_and_keeps_existing_success(self) -> None:
        workbook = self.root / "links.xlsx"
        workbook.write_bytes(b"workbook")
        output = self.root / "output"
        item = DirectLinkBatchItem(1, 2, "https://item.jd.com/7003.html", "jd", "测试商品")
        item_root = output / "001-row-0002"
        collected = item_root / "collected"
        collected.mkdir(parents=True)
        manifest = collected / "direct-manifest.json"
        manifest.write_text(
            json.dumps(
                {
                    "products": [{"product_id": "7003"}],
                    "extended_assets": [
                        {"types": {"main": {"images": [
                            {"status": "ok", "local_path": str(self.image)},
                            {"status": "ok", "local_path": str(self.second_image)},
                        ]}}}
                    ],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        generated = item_root / "generated"
        generated.mkdir()
        existing_output = generated / "main" / "002.png"
        existing_output.parent.mkdir()
        existing_output.write_bytes(b"existing")
        (generated / "analysis.json").write_text(
            json.dumps(
                {"records": [
                    {"category": "main", "ordinal": 1, "status": "failed", "source_path": str(self.image)},
                    {"category": "main", "ordinal": 2, "status": "completed", "source_path": str(self.second_image), "output_path": str(existing_output)},
                ]},
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        runner = BatchRunner(ApiSettings("https://api.example", "vision", "image"), self.root, self.root, batch_mode="direct_link")
        added = {"category": "main", "ordinal": 1, "status": "completed", "source_path": str(self.image), "output_path": str(generated / "main" / "001.png")}
        Path(added["output_path"]).write_bytes(b"supplement")

        with (
            patch("batch_workflow.extract_direct_link_items", return_value=[item]),
            patch("batch_workflow.WorkflowRunner.run", return_value=[added]) as run,
            patch("batch_workflow.upload_generation_records", side_effect=lambda records, _uploader: records),
            patch("batch_workflow.export_product_workbook", return_value=item_root / "测试商品.xlsx") as export,
        ):
            result = runner.supplement(workbook, output, 1, "main", 1)

        self.assertEqual(run.call_args.kwargs["requested_ordinals"], {"main": [1]})
        merged = json.loads((generated / "analysis.json").read_text(encoding="utf-8"))["records"]
        self.assertEqual([(record["ordinal"], record["status"]) for record in merged], [(1, "completed"), (2, "completed")])
        self.assertEqual(result["supplemented"], 1)
        export.assert_called_once()

    def test_direct_replace_sku_supplement_keeps_product_identity_and_manual_sku(self) -> None:
        item_root = self.root / "001-row-0002"
        generated = item_root / "generated"
        generated.mkdir(parents=True)
        manifest = item_root / "collected" / "direct-manifest.json"
        manifest.parent.mkdir(parents=True)
        manifest.write_text(
            json.dumps({"images": [{"type": "main", "path": str(self.second_image)}]}),
            encoding="utf-8",
        )
        item = DirectReplaceBatchItem(
            1,
            "商品",
            2,
            self.image,
            "https://item.jd.com/7009.html",
            "jd",
            manual_skus=(
                {
                    "sku_name": "白色大号",
                    "color": "白色",
                    "spec": "大号",
                    "price": "39.9",
                    "reference_image": "",
                    "source_status": "text_conditioned",
                },
            ),
        )
        added = {
            "category": "sku",
            "ordinal": 1,
            "status": "completed",
            "source_path": str(self.image),
            "output_path": str(self.second_image),
        }
        runner = BatchRunner(ApiSettings("https://api.example", "vision", "image"), self.root, self.root, batch_mode="direct_replace")

        with (
            patch("batch_workflow.WorkflowRunner.run", return_value=[added]) as run,
            patch("batch_workflow.upload_generation_records", side_effect=lambda records, _uploader: records),
            patch("batch_workflow.export_product_workbook", return_value=item_root / "商品.xlsx"),
        ):
            result = runner._supplement_item(
                item=item,
                item_root=item_root,
                source_manifest=manifest,
                generated_root=generated,
                generation_mode="own_product",
                workbook_path=item_root / "商品.xlsx",
                category="sku",
                count=1,
            )

        generation = json.loads((item_root / "generation-manifest.json").read_text(encoding="utf-8"))
        self.assertEqual(result["supplemented"], 1)
        self.assertEqual(run.call_args.args[1].resolve(), self.image.resolve())
        self.assertEqual(run.call_args.kwargs["identity_image"].resolve(), self.image.resolve())
        sku_source = next(image for image in generation["images"] if image["type"] == "sku")
        self.assertEqual(sku_source["manual_sku"]["spec"], "大号")

    def test_selected_workbook_supplement_updates_the_selected_excel(self) -> None:
        workbook = self.root / "selected.xlsx"
        workbook.write_bytes(b"xlsx")
        manifest = self.root / "manifest.json"
        manifest.write_text(json.dumps({"images": []}), encoding="utf-8")
        generated = self.root / "generated"
        generated.mkdir()
        item = DirectLinkBatchItem(1, 1, "https://item.jd.com/7006.html", "jd", "测试商品")
        context = SimpleNamespace(
            workbook_path=workbook,
            item=item,
            item_root=self.root,
            source_manifest=manifest,
            generated_root=generated,
            generation_mode="competitor_reference",
        )
        runner = BatchRunner(ApiSettings("https://api.example", "vision", "image"), self.root, self.root, batch_mode="direct_link")

        with (
            patch("batch_workflow.resolve_supplement_workbook", return_value=context),
            patch.object(runner, "_supplement_item", return_value={"workbook": str(workbook)}) as execute,
        ):
            result = runner.supplement_exported_workbook(workbook, "detail", 2)

        self.assertEqual(result["workbook"], str(workbook))
        self.assertEqual(execute.call_args.kwargs["workbook_path"], workbook)
        self.assertEqual(execute.call_args.kwargs["category"], "detail")
        self.assertEqual(execute.call_args.kwargs["count"], 2)
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.image = self.root / "product.png"
        Image.new("RGB", (120, 120), "white").save(self.image)
        self.second_image = self.root / "product-2.png"
        Image.new("RGB", (120, 120), "gray").save(self.second_image)

    def tearDown(self) -> None:
        self.temp.cleanup()

    def _shared_local_manifest(self, item_root: Path) -> Path:
        manifest = item_root / "collected" / "direct-manifest.json"
        manifest.parent.mkdir(parents=True, exist_ok=True)
        manifest.write_text(
            json.dumps(
                {
                    "images": [
                        {"type": "main", "path": str(self.image)},
                        {"type": "sku", "path": str(self.image)},
                        {"type": "detail", "path": str(self.image)},
                    ],
                    "products": [{"title": "测试商品", "sku_variants": []}],
                    "product_parameters": [],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        return manifest

    def _shared_generation_records(
        self,
        main_count: int,
        sku_count: int,
        detail_count: int,
    ) -> list[dict[str, object]]:
        records: list[dict[str, object]] = []
        for category, count in (("main", main_count), ("sku", sku_count), ("detail", detail_count)):
            for ordinal in range(1, count + 1):
                output_path = self.root / "generated" / category / f"{ordinal:03d}.png"
                output_path.parent.mkdir(parents=True, exist_ok=True)
                if not output_path.exists():
                    Image.new("RGB", (120, 120), "white").save(output_path)
                records.append(
                    {
                        "category": category,
                        "ordinal": ordinal,
                        "status": "completed",
                        "output_path": str(output_path),
                        "source_path": str(self.image),
                    }
                )
        return records

    def test_extracts_only_column_e_images_in_row_order(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(["标题", "原图", "单价", "视频", "1688商品图", "拿货价", "运费", "链接"])
        sheet.cell(2, 1, "第二行")
        sheet.cell(2, 6, 12.5)
        sheet.cell(2, 7, 3)
        sheet.cell(2, 8, "打开1688商品").hyperlink = "https://detail.1688.com/offer/2.html"
        sheet.cell(5, 1, "第五行")
        sheet.cell(5, 8, "https://detail.1688.com/offer/5.html")
        sheet.add_image(ExcelImage(str(self.image)), "E5")
        sheet.add_image(ExcelImage(str(self.image)), "B3")
        sheet.add_image(ExcelImage(str(self.image)), "E2")
        source = self.root / "input.xlsx"
        workbook.save(source)

        items = extract_batch_items(source, self.root / "output")

        self.assertEqual([item.row_number for item in items], [2, 5])
        self.assertEqual(items[0].source_url, "https://detail.1688.com/offer/2.html")
        self.assertEqual(items[0].purchase_price, 12.5)
        self.assertTrue(all(item.product_image.is_file() for item in items))

    def test_uses_all_embedded_images_when_no_1688_image_header_exists(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "photo-list"
        sheet.append(["sequence", "name", "product image"])
        sheet.add_image(ExcelImage(str(self.second_image)), "C5")
        sheet.add_image(ExcelImage(str(self.image)), "C3")
        source = self.root / "photos.xlsx"
        workbook.save(source)

        items = extract_batch_items(source, self.root / "output")

        self.assertEqual([item.row_number for item in items], [3, 5])
        self.assertEqual(len(items), 2)
        self.assertTrue(all(item.product_image.is_file() for item in items))

    def test_prefers_images_under_the_1688_product_image_header(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "products"
        sheet.cell(1, 2, "other image")
        sheet.cell(1, 5, "1688商品图")
        sheet.add_image(ExcelImage(str(self.image)), "B2")
        sheet.add_image(ExcelImage(str(self.second_image)), "E4")
        source = self.root / "prioritized.xlsx"
        workbook.save(source)

        items = extract_batch_items(source, self.root / "output")

        self.assertEqual([item.row_number for item in items], [4])
        self.assertEqual(len(items), 1)

    def test_extracts_named_direct_link_column_in_row_order(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["标题", "商品链接"])
        sheet.append(["淘宝商品", "https://item.taobao.com/item.htm?id=1001"])
        sheet.append(["京东商品", "https://item.jd.com/1002.html"])
        sheet.append(["无效商品", "https://example.com/item/1003"])
        source = self.root / "direct-links.xlsx"
        workbook.save(source)

        items = extract_direct_link_items(source)

        self.assertEqual([item.row_number for item in items], [2, 3, 4])
        self.assertEqual([item.platform for item in items], ["taobao", "jd", "unsupported"])
        self.assertEqual(items[0].title, "淘宝商品")
        self.assertEqual(items[2].validation_error, "不支持的商品链接平台")

    def test_direct_link_extracts_sku_screenshot(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(1, 1, "https://haohuo.jinritemai.com/views/product/item2.html?id=3827632284920578347")
        screenshot = self.root / "sku-screenshot.png"
        Image.new("RGB", (180, 120), "white").save(screenshot)
        sheet.cell(1, 2, screenshot.as_uri())
        source = self.root / "douyin-link-with-sku-screenshot.xlsx"
        workbook.save(source)

        items = extract_direct_link_items(source, self.root / "staged")

        self.assertEqual(len(items), 1)
        self.assertEqual(items[0].platform, "douyin")
        self.assertIsNotNone(items[0].sku_screenshot)
        self.assertTrue(items[0].sku_screenshot.is_file())
        self.assertNotEqual(items[0].sku_screenshot.resolve(), screenshot.resolve())

    def test_screenshot_skus_create_generation_sources_without_collected_sku_images(self) -> None:
        reference = self.root / "sku-reference.png"
        Image.new("RGB", (80, 80), "blue").save(reference)
        source_manifest = self.root / "source.json"
        source_manifest.write_text(
            json.dumps(
                {
                    "images": [{"type": "main", "path": str(self.image)}],
                    "sku_variants": [],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        item = DirectLinkBatchItem(
            1,
            1,
            "https://haohuo.jinritemai.com/views/product/item2.html?id=1",
            "douyin",
            manual_skus=(
                {
                    "sku_name": "蓝色 500ml",
                    "color": "蓝色",
                    "spec": "500ml",
                    "price": "39.90",
                    "reference_image": str(reference),
                    "source_status": "screenshot_thumbnail",
                    "visual_confidence": 0.95,
                },
                {
                    "sku_name": "红色 500ml",
                    "color": "红色",
                    "spec": "500ml",
                    "price": "39.90",
                    "reference_image": "",
                    "source_status": "low_visual_confidence",
                    "visual_confidence": 0.4,
                },
            ),
        )

        merged = merge_manual_sku_metadata(
            json.loads(source_manifest.read_text(encoding="utf-8")),
            item,
        )
        source_manifest.write_text(json.dumps(merged, ensure_ascii=False), encoding="utf-8")
        generation_manifest = build_generation_manifest(source_manifest, self.root / "generation.json")
        add_manual_sku_generation_sources(generation_manifest, item, self.image)
        generation = json.loads(generation_manifest.read_text(encoding="utf-8"))

        sku_images = [image for image in generation["images"] if image["type"] == "sku"]
        self.assertEqual(len(merged["sku_variants"]), 2)
        self.assertEqual(merged["sku_metadata_status"], "screenshot")
        self.assertEqual(len(sku_images), 2)
        self.assertEqual(Path(sku_images[0]["path"]).resolve(), reference.resolve())
        self.assertEqual(Path(sku_images[1]["path"]).resolve(), self.image.resolve())

    def test_direct_links_fall_back_to_first_column_without_header(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(1, 1, "https://detail.tmall.com/item.htm?id=2001")
        sheet.cell(2, 1, "https://item.jd.com/2002.html")
        sheet.cell(3, 1, "not-a-url")
        source = self.root / "first-column-links.xlsx"
        workbook.save(source)

        items = extract_direct_link_items(source)

        self.assertEqual([item.row_number for item in items], [1, 2, 3])
        self.assertEqual([item.platform for item in items], ["tmall", "jd", "invalid"])
        self.assertEqual(items[2].validation_error, "不是有效的 HTTP 商品链接")

    def test_direct_link_parser_accepts_known_taobao_and_jd_short_links(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["商品链接"])
        sheet.append(["https://m.tb.cn/h.abc123"])
        sheet.append(["https://u.jd.com/example"])
        source = self.root / "short-links.xlsx"
        workbook.save(source)

        items = extract_direct_link_items(source)

        self.assertEqual([item.platform for item in items], ["taobao", "jd"])
        self.assertEqual([item.validation_error for item in items], ["", ""])

    def test_direct_link_platform_accepts_tmall_global_item_url(self) -> None:
        self.assertEqual(
            _direct_link_platform("https://detail.tmall.hk/hk/item.htm?id=2003"),
            ("tmall", ""),
        )

    def test_direct_link_platform_accepts_e_tb_share_url(self) -> None:
        self.assertEqual(
            _direct_link_platform("https://e.tb.cn/h.abc123"),
            ("taobao", ""),
        )

    def test_direct_item_resolver_follows_e_tb_share_url(self) -> None:
        class FakeResponse:
            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return None

            @staticmethod
            def geturl() -> str:
                return "https://item.taobao.com/item.htm?id=2004"

        with patch("batch_workflow.urlopen", return_value=FakeResponse()):
            resolved = resolve_direct_item_url("https://e.tb.cn/h.abc123")

        self.assertEqual(resolved, "https://item.taobao.com/item.htm?id=2004")

    def test_direct_link_parser_accepts_platform_specific_link_headers(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["商品名称", "抖音链接"])
        sheet.append(["测试商品", "https://v.douyin.com/abc123"])
        source = self.root / "douyin-header.xlsx"
        workbook.save(source)

        items = extract_direct_link_items(source)

        self.assertEqual(len(items), 1)
        self.assertEqual(items[0].platform, "douyin")
        self.assertEqual(items[0].source_url, "https://v.douyin.com/abc123")

    def test_direct_link_platform_accepts_douyin_short_and_detail_urls(self) -> None:
        self.assertEqual(_direct_link_platform("https://v.douyin.com/abc123"), ("douyin", ""))
        self.assertEqual(
            _direct_link_platform(
                "https://haohuo.jinritemai.com/views/product/item2.html?id=3827632284920578347"
            ),
            ("douyin", ""),
        )

    def test_direct_link_platform_rejects_douyin_detail_without_product_id(self) -> None:
        platform, error = _direct_link_platform(
            "https://haohuo.jinritemai.com/views/product/item2.html"
        )

        self.assertEqual(platform, "invalid")
        self.assertIn("商品 ID", error)

    def test_direct_link_parser_does_not_read_other_columns_in_fallback_mode(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["商品名称", "备注"])
        sheet.append(["普通文本", "https://item.taobao.com/item.htm?id=3001"])
        source = self.root / "strict-first-column.xlsx"
        workbook.save(source)

        items = extract_direct_link_items(source)

        self.assertEqual(items, [])

    def test_direct_replace_extracts_wps_dispimg_rows_without_fixed_columns(self) -> None:
        fixture_dir = Path(
            r"C:\Users\Administrator\Documents\xwechat_files\wxid_ex358te1357c22_6c76\msg\file\2026-08"
        )
        fixture = next(fixture_dir.glob("*.xlsx"))

        items = extract_direct_replace_items(fixture, self.root / "direct-replace")

        self.assertEqual(len(items), 4)
        self.assertEqual([item.row_number for item in items], [1, 10, 19, 28])
        self.assertEqual([item.sheet_name for item in items], ["Sheet1"] * 4)
        self.assertTrue(all(item.product_image and item.product_image.is_file() for item in items))
        self.assertTrue(all(item.platform == "tmall" for item in items))
        self.assertTrue(all(not item.validation_error for item in items))

    def test_direct_replace_uses_named_image_and_link_columns(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "商品"
        sheet.append(["我方商品图", "标题", "抖音链接", "SKU名称", "颜色", "规格", "价格"])
        sheet.add_image(ExcelImage(self.image), "A2")
        sheet.cell(2, 2, "保温杯")
        sheet.cell(
            2,
            3,
            "https://haohuo.jinritemai.com/ecommerce/trade/detail/index.html?id=3827632284920578347",
        )
        sheet.cell(2, 4, "银色 500ml")
        sheet.cell(2, 5, "银色")
        sheet.cell(2, 6, "500ml")
        sheet.cell(2, 7, "39.90")
        source = self.root / "direct-replace.xlsx"
        workbook.save(source)

        items = extract_direct_replace_items(source, self.root / "direct-replace")

        self.assertEqual(len(items), 1)
        self.assertIsInstance(items[0], DirectReplaceBatchItem)
        self.assertEqual(items[0].sheet_name, "商品")
        self.assertEqual(items[0].platform, "douyin")
        self.assertEqual(items[0].title, "保温杯")
        self.assertEqual(items[0].manual_skus[0]["sku_name"], "银色 500ml")
        self.assertEqual(items[0].manual_skus[0]["price"], "39.90")

    def test_direct_replace_marks_ambiguous_no_header_row(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.add_image(ExcelImage(self.image), "A1")
        sheet.add_image(ExcelImage(self.second_image), "B1")
        sheet.cell(1, 3, "https://item.jd.com/1002.html")
        source = self.root / "ambiguous.xlsx"
        workbook.save(source)

        items = extract_direct_replace_items(source, self.root / "direct-replace")

        self.assertEqual(len(items), 1)
        self.assertIn("配对冲突", items[0].validation_error)

    def test_direct_collector_uses_store_insight_without_image_search(self) -> None:
        item = DirectLinkBatchItem(
            sequence=1,
            row_number=2,
            source_url="https://item.taobao.com/item.htm?id=4001",
            platform="taobao",
        )
        item_root = self.root / "item"
        output_manifest = item_root / "collected" / "manifest.json"
        output_manifest.parent.mkdir(parents=True)
        output_manifest.write_text(json.dumps({"images": []}), encoding="utf-8")

        class FakeProcess:
            pid = 123
            stdout = iter([f"[collector] manifest: {output_manifest}\n"])

            def wait(self):
                return 0

            def poll(self):
                return None

        collector = DirectLinkCollector(
            self.root,
            self.root / "profile",
            browser_executable="D:/Browser/browser.exe",
        )
        with patch("batch_workflow.subprocess.Popen", return_value=FakeProcess()) as popen:
            result = collector.collect(item, item_root)

        command = popen.call_args.args[0]
        self.assertEqual(result, output_manifest)
        self.assertIn("store_insight_collector.py", " ".join(command))
        self.assertIn(item.source_url, command)
        self.assertIn("--types", command)
        self.assertNotIn("same_item_collector.py", " ".join(command))

    def test_direct_batch_collects_all_main_assets_before_generating_ten(self) -> None:
        runner = BatchRunner(
            None,
            self.root,
            self.root / "profile",
            max_main_images=10,
            batch_mode="direct_link",
        )

        self.assertIsNone(runner.direct_collector.max_main_images)

    def test_direct_collector_resolves_short_link_before_collection(self) -> None:
        item = DirectLinkBatchItem(
            sequence=1,
            row_number=2,
            source_url="https://m.tb.cn/h.abc123",
            platform="taobao",
        )
        item_root = self.root / "short-item"
        output_manifest = item_root / "collected" / "manifest.json"
        output_manifest.parent.mkdir(parents=True)
        output_manifest.write_text(json.dumps({"images": []}), encoding="utf-8")

        class FakeProcess:
            pid = 123
            stdout = iter([f"[collector] manifest: {output_manifest}\n"])

            def wait(self):
                return 0

            def poll(self):
                return None

        collector = DirectLinkCollector(self.root, self.root / "profile")
        with (
            patch("batch_workflow.resolve_direct_item_url", return_value="https://detail.tmall.com/item.htm?id=4002"),
            patch("batch_workflow.subprocess.Popen", return_value=FakeProcess()) as popen,
        ):
            collector.collect(item, item_root)

        self.assertIn("https://detail.tmall.com/item.htm?id=4002", popen.call_args.args[0])

    def test_direct_collector_turns_login_timeout_into_batch_pause(self) -> None:
        item = DirectLinkBatchItem(1, 2, "https://item.taobao.com/item.htm?id=4003", "taobao")

        class FakeProcess:
            pid = 123
            stdout = iter([
                "[collector] 检测到登录页面，请完成登录。\n",
                "[collector] 采集失败：Taobao/Tmall login was not completed before the wait timeout\n",
            ])

            def wait(self):
                return 1

            def poll(self):
                return None

        collector = DirectLinkCollector(self.root, self.root / "profile")
        with patch("batch_workflow.subprocess.Popen", return_value=FakeProcess()):
            with self.assertRaises(BatchCollectionPaused):
                collector.collect(item, self.root / "login-item")

    def test_normalizes_direct_manifest_for_generation_and_export(self) -> None:
        main = self.root / "main.jpg"
        sku = self.root / "sku.jpg"
        main.write_bytes(b"main")
        sku.write_bytes(b"sku")
        source = self.root / "direct-manifest.json"
        source.write_text(
            json.dumps(
                {
                    "source_url": "https://item.jd.com/5001.html",
                    "product_id": "5001",
                    "images": [
                        {"type": "main", "path": str(main)},
                        {
                            "type": "sku",
                            "path": str(sku),
                            "sku_id": "sku-1",
                            "color_text": "红色",
                            "spec_text": "大号",
                            "after_coupon_price": "19.9",
                        },
                    ],
                    "product_parameters": [{"name": "材质", "value": "棉"}],
                    "sku_variants": [{"sku_id": "sku-1", "color_text": "红色"}],
                    "main_video_url": "https://video.example/5001.mp4",
                    "main_video_status": "complete",
                    "missing_asset_types": ["detail"],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        normalized = normalize_direct_manifest(source, self.root / "normalized.json")
        document = json.loads(normalized.read_text(encoding="utf-8"))

        self.assertEqual(document["products"][0]["item_url"], "https://item.jd.com/5001.html")
        self.assertEqual(document["extended_assets"][0]["types"]["main"]["images"][0]["local_path"], str(main))
        self.assertEqual(document["extended_assets"][0]["types"]["sku"]["images"][0]["color_text"], "红色")
        self.assertEqual(document["product_parameters"][0]["name"], "材质")
        self.assertEqual(document["sku_variants"][0]["sku_id"], "sku-1")
        self.assertEqual(document["main_video_url"], "https://video.example/5001.mp4")
        self.assertEqual(document["missing_asset_types"], ["detail"])

    def test_restore_returns_existing_normalized_direct_manifest(self) -> None:
        item_root = self.root / "restore-direct"
        manifest = item_root / "collected" / "direct-manifest.json"
        manifest.parent.mkdir(parents=True)
        manifest.write_text(
            json.dumps(
                {
                    "products": [
                        {
                            "types": {
                                "main": {"images": [{"status": "ok", "local_path": str(self.image)}]},
                            }
                        }
                    ]
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        restored = restore_collected_manifest(item_root)

        self.assertEqual(restored, (manifest, 1))

    def test_direct_link_batch_skips_invalid_rows_and_generates_available_types(self) -> None:
        workbook = self.root / "direct.xlsx"
        workbook.write_bytes(b"workbook")
        output = self.root / "output"
        valid = DirectLinkBatchItem(
            1,
            2,
            "https://item.taobao.com/item.htm?id=6001",
            "taobao",
            "有效商品",
        )
        invalid = DirectLinkBatchItem(
            2,
            3,
            "https://example.com/6002",
            "unsupported",
            "无效商品",
            "不支持的商品链接平台",
        )
        normalized = self.root / "normalized.json"
        normalized.write_text(
            json.dumps(
                {
                    "products": [{"product_id": "6001", "title": "对标标题"}],
                    "extended_assets": [
                        {
                            "types": {
                                "main": {"images": [{"status": "ok", "local_path": str(self.image)}]},
                                "sku": {"images": []},
                                "detail": {"images": []},
                            }
                        }
                    ],
                    "product_parameters": [],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        runner = BatchRunner(None, self.root, self.root, batch_mode="direct_link")
        runner.direct_collector = SimpleNamespace(collect=lambda _item, _root: normalized, cancel=lambda: None)
        generated = [{"category": "main", "ordinal": 1, "status": "completed", "output_path": str(self.image)}]
        workflow = SimpleNamespace(run=Mock(return_value=generated), cancel=lambda: None)

        with (
            patch("batch_workflow.extract_direct_link_items", return_value=[valid, invalid]),
            patch("batch_workflow.normalize_direct_manifest", side_effect=lambda source, target: source),
            patch("batch_workflow.ProductTitleClient.generate", return_value={"long_title": "长标题", "short_title": "短标题"}),
            patch("batch_workflow.WorkflowRunner", return_value=workflow),
            patch("batch_workflow.upload_generation_records", side_effect=lambda records, _uploader: records),
            patch("batch_workflow.export_product_workbook", return_value=self.root / "result.xlsx"),
        ):
            results = runner.run(workbook, output)

        self.assertEqual([item["status"] for item in results], ["completed", "failed"])
        self.assertEqual(results[1]["error"], "不支持的商品链接平台")
        run_kwargs = workflow.run.call_args.kwargs
        self.assertEqual(workflow.run.call_args.args[4], ("main",))
        self.assertEqual(run_kwargs["generation_mode"], "competitor_reference")
        self.assertEqual(run_kwargs["identity_image"], self.image.resolve())

    def test_direct_link_export_keeps_sku_metadata_without_sku_images(self) -> None:
        source = {
            "products": [{"product_id": "7001"}],
            "extended_assets": [{"types": {"sku": {"images": []}}}],
            "sku_variants": [
                {
                    "sku_id": "sku-7001",
                    "sku_label": "红色;大号",
                    "color_text": "红色",
                    "spec_text": "大号",
                    "after_coupon_price": "29.9",
                    "parse_status": "parsed",
                }
            ],
        }

        rows = _sku_rows(source, [], source["products"][0], include_metadata_only=True)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["sku_id"], "sku-7001")
        self.assertEqual(rows[0]["color_text"], "红色")
        self.assertEqual(rows[0]["spec_text"], "大号")
        self.assertEqual(rows[0]["price"], "29.9")
        self.assertEqual(rows[0]["source_path"], "")

    def test_manual_sku_generation_is_mapped_to_metadata_only_export_row(self) -> None:
        source = {
            "products": [{"product_id": "7007"}],
            "extended_assets": [{"types": {"sku": {"images": []}}}],
            "sku_variants": [
                {
                    "source_index": "1",
                    "sku_label": "白色大号",
                    "color_text": "白色",
                    "spec_text": "大号",
                    "list_price": "39.9",
                    "parse_status": "manual",
                }
            ],
        }
        generated = [
            {
                "category": "sku",
                "ordinal": 1,
                "status": "completed",
                "source_path": str(self.image),
                "output_path": str(self.second_image),
                "manual_sku": {
                    "sku_name": "白色大号",
                    "color": "白色",
                    "spec": "大号",
                    "price": "39.9",
                },
            }
        ]

        rows = _sku_rows(source, generated, source["products"][0], include_metadata_only=True)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["output_path"], str(self.second_image))
        self.assertEqual(rows[0]["generation_status"], "生成成功")

    def test_direct_replace_uses_current_row_product_image_as_own_product_identity(self) -> None:
        workbook = self.root / "direct-replace.xlsx"
        workbook.write_bytes(b"workbook")
        output = self.root / "output"
        item = DirectReplaceBatchItem(
            1,
            "商品",
            2,
            self.image,
            "https://item.jd.com/6101.html",
            "jd",
            "我方商品",
        )
        normalized = self.root / "normalized.json"
        normalized.write_text(
            json.dumps(
                {
                    "products": [{"product_id": "6101", "title": "对标标题"}],
                    "extended_assets": [
                        {
                            "types": {
                                "main": {"images": [{"status": "ok", "local_path": str(self.second_image)}]},
                                "sku": {"images": []},
                                "detail": {"images": []},
                            }
                        }
                    ],
                    "product_parameters": [],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        runner = BatchRunner(None, self.root, self.root, batch_mode="direct_replace")
        runner.direct_collector = SimpleNamespace(collect=lambda _item, _root: normalized, cancel=lambda: None)
        generated = [{"category": "main", "ordinal": 1, "status": "completed", "output_path": str(self.image)}]
        workflow = SimpleNamespace(run=Mock(return_value=generated), cancel=lambda: None)

        with (
            patch("batch_workflow.extract_direct_replace_items", return_value=[item]),
            patch("batch_workflow.normalize_direct_manifest", side_effect=lambda source, target: source),
            patch("batch_workflow.ProductTitleClient.generate", return_value={"long_title": "长标题", "short_title": "短标题"}),
            patch("batch_workflow.WorkflowRunner", return_value=workflow),
            patch("batch_workflow.upload_generation_records", side_effect=lambda records, _uploader: records),
            patch("batch_workflow.upload_video_if_needed", side_effect=lambda document, _uploader, _namespace: document) as upload_video,
            patch("batch_workflow.export_product_workbook", return_value=self.root / "result.xlsx"),
        ):
            results = runner.run(workbook, output)

        self.assertEqual(results[0]["status"], "completed")
        self.assertEqual(workflow.run.call_args.args[1], self.image)
        self.assertEqual(workflow.run.call_args.kwargs["generation_mode"], "own_product")
        self.assertEqual(workflow.run.call_args.kwargs["identity_image"], self.image)
        upload_video.assert_called_once()

    def test_video_oss_failure_marks_full_batch_result_failed(self) -> None:
        workbook = self.root / "direct-link.xlsx"
        workbook.write_bytes(b"workbook")
        output = self.root / "output"
        item = DirectLinkBatchItem(1, 2, "https://item.jd.com/6102.html", "jd", "对标商品")
        normalized = self.root / "normalized-video.json"
        normalized.write_text(
            json.dumps(
                {
                    "products": [{"product_id": "6102", "title": "对标商品"}],
                    "extended_assets": [
                        {"types": {"main": {"images": [{"status": "ok", "local_path": str(self.image)}]}}}
                    ],
                    "product_parameters": [],
                    "main_video_local_path": str(self.second_image),
                    "main_video_status": "local_only",
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        runner = BatchRunner(None, self.root, self.root, batch_mode="direct_link")
        runner.direct_collector = SimpleNamespace(collect=lambda _item, _root: normalized, cancel=lambda: None)
        generated = [{"category": "main", "ordinal": 1, "status": "completed", "output_path": str(self.image)}]
        workflow = SimpleNamespace(run=Mock(return_value=generated), cancel=lambda: None)

        with (
            patch("batch_workflow.extract_direct_link_items", return_value=[item]),
            patch("batch_workflow.normalize_direct_manifest", side_effect=lambda source, target: source),
            patch("batch_workflow.ProductTitleClient.generate", return_value={"long_title": "长标题", "short_title": "短标题"}),
            patch("batch_workflow.WorkflowRunner", return_value=workflow),
            patch("batch_workflow.upload_generation_records", side_effect=lambda records, _uploader: records),
            patch(
                "batch_workflow.upload_video_if_needed",
                side_effect=lambda document, _uploader, _namespace: {
                    **document,
                    "main_video_status": "failed",
                    "main_video_error": "OSS 上传失败",
                },
            ),
            patch("batch_workflow.export_product_workbook", return_value=self.root / "result.xlsx"),
        ):
            results = runner.run(workbook, output)

        self.assertEqual(results[0]["status"], "failed")
        self.assertIn("视频", results[0]["error"])

    def test_manual_sku_metadata_is_truthfully_marked_and_can_use_main_as_generation_reference(self) -> None:
        item = DirectReplaceBatchItem(
            1,
            "商品",
            2,
            self.image,
            "https://haohuo.jinritemai.com/ecommerce/trade/detail/index.html?id=6201",
            "douyin",
            manual_skus=(
                {
                    "sku_name": "银色 500ml",
                    "color": "银色",
                    "spec": "500ml",
                    "price": "39.90",
                    "reference_image": "",
                    "source_status": "text_conditioned",
                },
            ),
        )
        source = merge_manual_sku_metadata(
            {"sku_variants": [], "sku_metadata_status": "not_found"},
            item,
        )
        generation_manifest = self.root / "manual-sku-generation.json"
        generation_manifest.write_text(json.dumps({"images": [{"type": "main", "path": str(self.image)}]}), encoding="utf-8")

        add_manual_sku_generation_sources(generation_manifest, item, self.image)
        generation = json.loads(generation_manifest.read_text(encoding="utf-8"))

        self.assertEqual(source["sku_metadata_status"], "manual")
        self.assertEqual(source["sku_variants"][0]["sku_label"], "银色 500ml")
        self.assertEqual(source["sku_variants"][0]["source_status"], "text_conditioned")
        self.assertEqual(source["sku_variants"][0]["list_price"], "39.90")
        self.assertEqual(generation["images"][-1]["type"], "sku")
        self.assertEqual(generation["images"][-1]["path"], str(self.image))
        self.assertEqual(generation["images"][-1]["manual_sku"]["spec"], "500ml")

    def test_manual_sku_metadata_does_not_replace_platform_variants(self) -> None:
        item = DirectReplaceBatchItem(
            1,
            "商品",
            2,
            self.image,
            "https://detail.tmall.com/item.htm?id=6202",
            "tmall",
            manual_skus=(
                {
                    "sku_name": "手工银色 500ml",
                    "color": "银色",
                    "spec": "500ml",
                    "price": "39.90",
                },
            ),
        )
        platform_variants = [
            {
                "sku_id": "platform-1",
                "sku_label": "平台红色 300ml",
                "color_text": "红色",
                "spec_text": "300ml",
                "list_price": "29.90",
                "parse_status": "ok",
            }
        ]

        merged = merge_manual_sku_metadata(
            {"sku_variants": platform_variants, "sku_metadata_status": "ok"},
            item,
        )

        self.assertEqual(merged["sku_variants"], platform_variants)
        self.assertEqual(merged["sku_metadata_status"], "ok")

    def test_screenshot_sku_metadata_replaces_reused_platform_variants(self) -> None:
        screenshot = self.root / "current-sku-screenshot.png"
        screenshot.write_bytes(b"screenshot")
        item = DirectLinkBatchItem(
            2,
            3,
            "https://haohuo.jinritemai.com/views/product/item2.html?id=9002",
            "douyin",
            sku_screenshot=screenshot,
            manual_skus=(
                {
                    "sku_name": "current red",
                    "color": "red",
                    "spec": "1 piece",
                    "price": "29.90",
                    "reference_image": "current-red.png",
                    "source_status": "screenshot_thumbnail",
                },
            ),
        )
        reused_variants = [
            {
                "sku_id": "old-1",
                "sku_label": "old blue",
                "reference_image": "old-blue.png",
            }
        ]

        merged = merge_manual_sku_metadata(
            {"sku_variants": reused_variants, "sku_metadata_status": "ok"},
            item,
        )

        self.assertEqual(merged["sku_variants"][0]["sku_label"], "current red")
        self.assertEqual(merged["sku_variants"][0]["reference_image"], "current-red.png")
        self.assertEqual(merged["sku_metadata_status"], "screenshot")

    def test_manual_sku_images_do_not_replace_platform_generation_sources(self) -> None:
        item = DirectReplaceBatchItem(
            1,
            "商品",
            2,
            self.image,
            "https://detail.tmall.com/item.htm?id=6203",
            "tmall",
            manual_skus=(
                {
                    "sku_name": "手工银色 500ml",
                    "color": "银色",
                    "spec": "500ml",
                    "price": "39.90",
                },
            ),
        )
        generation_manifest = self.root / "platform-sku-generation.json"
        platform_images = [
            {"type": "main", "path": str(self.image)},
            {"type": "sku", "path": str(self.second_image), "source": "platform"},
        ]
        generation_manifest.write_text(json.dumps({"images": platform_images}), encoding="utf-8")

        add_manual_sku_generation_sources(generation_manifest, item, self.image)
        generation = json.loads(generation_manifest.read_text(encoding="utf-8"))

        self.assertEqual(generation["images"], platform_images)

    def test_direct_link_export_uses_neutral_source_labels_and_platform(self) -> None:
        manifest = self.root / "direct-export.json"
        manifest.write_text(
            json.dumps(
                {
                    "products": [
                        {
                            "product_id": "7002",
                            "title": "京东商品",
                            "item_url": "https://item.jd.com/7002.html",
                        }
                    ],
                    "extended_assets": [{"types": {}}],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        item = DirectLinkBatchItem(1, 2, "https://item.jd.com/7002.html", "jd", "京东商品")

        with patch("batch_workflow._export_with_artifact_tool", return_value=self.root / "direct.xlsx") as exporter:
            export_product_workbook(self.root / "direct.xlsx", item, manifest, [])

        overview = dict(exporter.call_args.kwargs["payload"]["overview"][1:])
        self.assertEqual(overview["来源平台"], "京东")
        self.assertEqual(overview["来源商品链接"], item.source_url)
        self.assertIn("商品当前价", overview)
        self.assertNotIn("1688商品链接", overview)
        self.assertNotIn("1688拿货价", overview)

    def test_direct_replace_export_keeps_source_sheet_and_product_identity(self) -> None:
        manifest = self.root / "direct-replace-export.json"
        manifest.write_text(
            json.dumps(
                {
                    "products": [{"product_id": "7008", "title": "抖音商品"}],
                    "extended_assets": [{"types": {}}],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        item = DirectReplaceBatchItem(
            1,
            "待处理商品",
            9,
            self.image,
            "https://haohuo.jinritemai.com/views/product/item2.html?id=7008",
            "douyin",
        )

        with patch("batch_workflow._export_with_artifact_tool", return_value=self.root / "direct-replace.xlsx") as exporter:
            export_product_workbook(self.root / "direct-replace.xlsx", item, manifest, [])

        overview = dict(exporter.call_args.kwargs["payload"]["overview"][1:])
        self.assertEqual(overview["来源平台"], "抖音")
        self.assertEqual(overview["来源工作表"], "待处理商品")
        self.assertEqual(overview["我方商品图"], str(self.image))

    def test_direct_link_export_keeps_collected_images_when_generation_is_empty(self) -> None:
        manifest = self.root / "direct-generated-only.json"
        manifest.write_text(
            json.dumps(
                {
                    "products": [{"product_id": "7003", "item_url": "https://item.jd.com/7003.html"}],
                    "extended_assets": [
                        {
                            "types": {
                                "main": {
                                    "images": [
                                        {"status": "ok", "local_path": str(self.image)},
                                        {"status": "ok", "local_path": str(self.second_image)},
                                    ]
                                }
                            }
                        }
                    ],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        records = []
        item = DirectLinkBatchItem(1, 2, "https://item.jd.com/7003.html", "jd")

        with patch("batch_workflow._export_with_artifact_tool", return_value=self.root / "direct.xlsx") as exporter:
            export_product_workbook(self.root / "direct.xlsx", item, manifest, records)

        rows = exporter.call_args.kwargs["payload"]["main"]
        self.assertEqual(len(rows), 2)
        self.assertTrue(all(row["generation_status"] == "未生成" for row in rows))

    def test_export_supports_flat_collected_manifest_images(self) -> None:
        manifest = self.root / "flat-manifest.json"
        manifest.write_text(
            json.dumps(
                {
                    "product_id": "7004",
                    "product_title": "扁平清单商品",
                    "images": [
                        {"type": "main", "path": str(self.image)},
                        {"type": "detail", "path": str(self.second_image)},
                    ],
                    "product_parameters": [{"name": "材质", "value": "测试"}],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        item = DirectLinkBatchItem(1, 1, "https://item.jd.com/7004.html", "jd")

        with patch("batch_workflow._export_with_artifact_tool", return_value=self.root / "flat.xlsx") as exporter:
            export_product_workbook(self.root / "flat.xlsx", item, manifest, [])

        payload = exporter.call_args.kwargs["payload"]
        self.assertEqual(len(payload["main"]), 1)
        self.assertEqual(len(payload["detail"]), 1)
        self.assertEqual(payload["parameters"][0]["value"], "测试")

    def test_flattens_only_usable_assets(self) -> None:
        manifest = self.root / "source.json"
        manifest.write_text(
            json.dumps(
                {
                    "extended_assets": [
                        {
                            "types": {
                                "main": {"images": [{"status": "ok", "local_path": str(self.image)}]},
                                "sku": {"images": [{"status": "metadata_only", "local_path": ""}]},
                            }
                        }
                    ]
                }
            ),
            encoding="utf-8",
        )

        target = build_generation_manifest(manifest, self.root / "generation.json")
        document = json.loads(target.read_text(encoding="utf-8"))

        self.assertEqual(document["images"], [{"type": "main", "path": str(self.image.resolve())}])

    def test_batch_checkpoint_restores_only_for_the_same_workbook(self) -> None:
        workbook = self.root / "input.xlsx"
        workbook.write_bytes(b"workbook")
        output = self.root / "output"
        expected = [{"sequence": 2, "row": 5, "status": "completed", "workbook": "result.xlsx"}]

        save_batch_results(workbook, output, expected)

        self.assertEqual(load_batch_results(workbook, output), expected)
        self.assertEqual(load_batch_results(self.root / "other.xlsx", output), [])

    def test_resume_skips_completed_items(self) -> None:
        workbook = self.root / "input.xlsx"
        workbook.write_bytes(b"workbook")
        output = self.root / "output"
        items = [
            BatchItem(1, 2, "商品一", self.image, "", "", ""),
            BatchItem(2, 3, "商品二", self.second_image, "", "", ""),
        ]
        completed = [
            {
                "sequence": item.sequence,
                "row": item.row_number,
                "status": "completed",
                "workbook": str(output / f"{item.sequence:03d}-row-{item.row_number:04d}" / f"{item.sequence}.xlsx"),
            }
            for item in items
        ]
        for item, result in zip(items, completed):
            item_root = output / f"{item.sequence:03d}-row-{item.row_number:04d}"
            generated = item_root / "generated" / "main" / "001.png"
            generated.parent.mkdir(parents=True)
            generated.write_bytes(b"generated")
            (item_root / "generated" / "analysis.json").write_text(
                json.dumps({"records": [{"category": "main", "status": "completed", "output_path": str(generated)}]}),
                encoding="utf-8",
            )
            Path(result["workbook"]).write_bytes(b"workbook")
        save_batch_results(workbook, output, completed)
        runner = BatchRunner(None, self.root, self.root)

        with patch("batch_workflow.extract_batch_items", return_value=items), patch.object(runner, "_collect") as collect:
            results = runner.run(workbook, output)

        collect.assert_not_called()
        self.assertEqual(results, completed)

    def test_resume_retries_legacy_completed_item_without_generated_images(self) -> None:
        workbook = self.root / "input.xlsx"
        workbook.write_bytes(b"workbook")
        output = self.root / "output"
        item = BatchItem(1, 2, "商品一", self.image, "", "", "")
        item_root = output / "001-row-0002"
        exported = item_root / "商品一.xlsx"
        exported.parent.mkdir(parents=True)
        exported.write_bytes(b"empty workbook")
        save_batch_results(
            workbook,
            output,
            [{"sequence": 1, "row": 2, "status": "completed", "workbook": str(exported)}],
        )
        runner = BatchRunner(None, self.root, self.root)

        with (
            patch("batch_workflow.extract_batch_items", return_value=[item]),
            patch.object(runner, "_collect", side_effect=RuntimeError("reprocessed")) as collect,
        ):
            results = runner.run(workbook, output)

        collect.assert_called_once()
        self.assertEqual(results[0]["status"], "failed")
        self.assertEqual(results[0]["error"], "reprocessed")

    def test_collect_only_saves_manifest_without_starting_generation(self) -> None:
        workbook = self.root / "input.xlsx"
        workbook.write_bytes(b"workbook")
        output = self.root / "output"
        item = BatchItem(1, 2, "商品一", self.image, "", "", "")
        manifest = output / "001-row-0002" / "collected" / "main-image-manifest.json"
        manifest.parent.mkdir(parents=True)
        manifest.write_text(
            json.dumps(
                {
                    "products": [{"title": "采集标题"}],
                    "extended_assets": [
                        {"types": {"main": {"images": [{"status": "ok", "local_path": str(self.image)}]}}}
                    ],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        runner = BatchRunner(None, self.root, self.root, collect_only=True)

        with (
            patch("batch_workflow.extract_batch_items", return_value=[item]),
            patch.object(runner, "_collect", return_value=manifest),
            patch("batch_workflow.ProductTitleClient") as title_client,
            patch("batch_workflow.WorkflowRunner") as workflow_runner,
            patch("batch_workflow.export_product_workbook") as export_workbook,
        ):
            results = runner.run(workbook, output)

        self.assertEqual(results[0]["status"], "collected")
        self.assertEqual(Path(results[0]["manifest"]), manifest)
        title_client.assert_not_called()
        workflow_runner.assert_not_called()
        export_workbook.assert_not_called()

    def test_full_batch_reuses_collect_only_checkpoint_and_generates(self) -> None:
        workbook = self.root / "input.xlsx"
        workbook.write_bytes(b"workbook")
        output = self.root / "output"
        item = BatchItem(1, 2, "商品一", self.image, "", "", "")
        item_root = output / "001-row-0002"
        manifest = item_root / "collected" / "main-image-manifest.json"
        manifest.parent.mkdir(parents=True)
        manifest.write_text(
            json.dumps(
                {
                    "products": [{"title": "采集标题"}],
                    "extended_assets": [
                        {"types": {"main": {"images": [{"status": "ok", "local_path": str(self.image)}]}}}
                    ],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        save_batch_results(
            workbook,
            output,
            [{"sequence": 1, "row": 2, "status": "collected", "manifest": str(manifest)}],
        )
        runner = BatchRunner(ApiSettings("https://api.example", "vision", "image"), self.root, self.root)

        with (
            patch("batch_workflow.extract_batch_items", return_value=[item]),
            patch.object(runner, "_collect") as collect,
            patch("batch_workflow.ProductTitleClient.generate", return_value={"long_title": "标题", "short_title": "短标题"}),
            patch("batch_workflow.WorkflowRunner.run", return_value=[
                {"category": "main", "ordinal": 1, "status": "completed", "output_path": str(self.image)}
            ]),
            patch("batch_workflow.export_product_workbook", return_value=item_root / "result.xlsx"),
        ):
            results = runner.run(workbook, output)

        collect.assert_not_called()
        self.assertEqual(results[0]["status"], "completed")

    def test_retry_generates_only_failed_ordinals_and_preserves_successful_records(self) -> None:
        workbook = self.root / "input.xlsx"
        workbook.write_bytes(b"workbook")
        output = self.root / "output"
        item = BatchItem(1, 2, "商品一", self.image, "", "", "")
        item_root = output / "001-row-0002"
        manifest = item_root / "collected" / "main-image-manifest.json"
        manifest.parent.mkdir(parents=True)
        manifest.write_text(
            json.dumps(
                {
                    "products": [{"title": "采集标题"}],
                    "extended_assets": [
                        {"types": {"main": {"images": [{"status": "ok", "local_path": str(self.image)}]}}}
                    ],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        generated_root = item_root / "generated"
        first_output = generated_root / "main" / "001.png"
        second_output = generated_root / "main" / "002.png"
        first_output.parent.mkdir(parents=True)
        first_output.write_bytes(b"first")
        second_output.write_bytes(b"second")
        previous_records = [
            {"category": "main", "ordinal": 1, "status": "completed", "output_path": str(first_output)},
            {"category": "main", "ordinal": 2, "status": "failed", "error": "temporary"},
        ]
        (generated_root / "analysis.json").write_text(
            json.dumps({"records": previous_records}, ensure_ascii=False),
            encoding="utf-8",
        )
        save_batch_results(
            workbook,
            output,
            [{"sequence": 1, "row": 2, "status": "failed", "error": "1 image failed"}],
        )
        runner = BatchRunner(
            ApiSettings("https://api.example", "vision", "image"),
            self.root,
            self.root,
            max_main_images=2,
        )
        added_record = {
            "category": "main",
            "ordinal": 2,
            "status": "completed",
            "output_path": str(second_output),
        }

        with (
            patch("batch_workflow.extract_batch_items", return_value=[item]),
            patch.object(runner, "_collect") as collect,
            patch("batch_workflow.ProductTitleClient.generate", return_value={"long_title": "标题", "short_title": "短标题"}),
            patch("batch_workflow.WorkflowRunner.run", return_value=[added_record]) as workflow_run,
            patch("batch_workflow.export_product_workbook", return_value=item_root / "result.xlsx") as export,
        ):
            results = runner.run(workbook, output)

        collect.assert_not_called()
        self.assertEqual(workflow_run.call_args.kwargs["requested_ordinals"], {"main": [2]})
        self.assertEqual(workflow_run.call_args.kwargs["existing_records"], previous_records)
        exported_records = export.call_args.args[3]
        self.assertEqual([(record["category"], record["ordinal"]) for record in exported_records], [("main", 1), ("main", 2)])
        self.assertEqual(results[0]["status"], "completed")

    def test_restores_interrupted_collection_assets_without_new_search(self) -> None:
        collected = self.root / "item" / "collected"
        asset_manifest = collected / "_work" / "extended-competitor-assets.json"
        asset_manifest.parent.mkdir(parents=True)
        asset_manifest.write_text(
            json.dumps(
                {
                    "status": "complete",
                    "products": [
                        {
                            "product_id": "123",
                            "types": {
                                "main": {
                                    "images": [{"status": "ok", "local_path": str(self.image)}],
                                },
                            },
                        },
                    ],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        restored = restore_collected_manifest(collected.parent)

        self.assertIsNotNone(restored)
        manifest, count = restored
        self.assertEqual(count, 1)
        self.assertTrue(manifest.is_file())
        document = json.loads(manifest.read_text(encoding="utf-8"))
        self.assertEqual(len(document["extended_assets"]), 1)

    def test_does_not_restore_explicitly_partial_collection_manifest(self) -> None:
        collected = self.root / "partial-item" / "collected"
        manifest = collected / "main-image-manifest.json"
        manifest.parent.mkdir(parents=True)
        manifest.write_text(
            json.dumps(
                {
                    "status": "partial",
                    "extended_assets": [
                        {"types": {"main": {"images": [{"status": "ok", "local_path": str(self.image)}]}}}
                    ],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        self.assertIsNone(restore_collected_manifest(collected.parent))

    def test_reuses_assets_from_an_earlier_batch_for_the_same_product_image(self) -> None:
        batches = self.root / "batches"
        previous = batches / "previous"
        previous_image = previous / "source-images" / "001-row-0002.png"
        previous_image.parent.mkdir(parents=True)
        shutil.copy2(self.image, previous_image)
        asset_manifest = previous / "001-row-0002" / "collected" / "_work" / "extended-competitor-assets.json"
        asset_manifest.parent.mkdir(parents=True)
        asset_manifest.write_text(
            json.dumps(
                {"products": [{"types": {"main": {"images": [{"status": "ok", "local_path": str(self.image)}]}}}]},
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        current = batches / "current"
        current_image = current / "source-images" / "001-row-0002.png"
        current_image.parent.mkdir(parents=True)
        shutil.copy2(self.image, current_image)

        recovered = find_prior_collected_manifest(current_image, current / "001-row-0002", current)

        self.assertIsNotNone(recovered)
        manifest, count = recovered
        self.assertEqual(count, 1)
        self.assertIn("previous", str(manifest))

    def test_direct_batch_reuses_historical_collection_for_same_product_id(self) -> None:
        batches = self.root / "outputs" / "batches"
        previous = batches / "previous" / "001-row-0002" / "collected"
        previous.mkdir(parents=True)
        previous_manifest = previous / "direct-manifest.json"
        previous_manifest.write_text(
            json.dumps(
                {
                    "source_url": "https://detail.tmall.com/item.htm?id=5001&spm=old",
                    "product_id": "5001",
                    "images": [{"type": "main", "path": str(self.image)}],
                    "requested_asset_types": ["main", "sku", "detail"],
                    "collected_asset_types": ["main"],
                    "missing_asset_types": ["sku", "detail"],
                }
            ),
            encoding="utf-8",
        )
        workbook = self.root / "input.xlsx"
        workbook.write_bytes(b"workbook")

        for mode in ("direct_link", "direct_replace"):
            with self.subTest(mode=mode):
                output = batches / f"current-{mode}"
                if mode == "direct_link":
                    item = DirectLinkBatchItem(
                        1,
                        2,
                        "https://detail.tmall.com/item.htm?id=5001&spm=new",
                        "tmall",
                    )
                    extractor = "batch_workflow.extract_direct_link_items"
                else:
                    item = DirectReplaceBatchItem(
                        1,
                        "Sheet1",
                        2,
                        self.image,
                        "https://detail.tmall.com/item.htm?id=5001&spm=new",
                        "tmall",
                    )
                    extractor = "batch_workflow.extract_direct_replace_items"
                runner = BatchRunner(
                    None,
                    self.root,
                    self.root / "profile",
                    batch_mode=mode,
                    collect_only=True,
                )

                with (
                    patch(extractor, return_value=[item]),
                    patch.object(
                        runner.direct_collector,
                        "collect",
                        side_effect=RuntimeError("collector should not run"),
                    ) as collect,
                ):
                    results = runner.run(workbook, output)

                collect.assert_not_called()
                self.assertEqual(results[0]["status"], "collected")
                reused_manifest = Path(results[0]["manifest"])
                self.assertTrue(reused_manifest.is_file())
                self.assertIn("current-", str(reused_manifest))

    def test_cancel_keeps_the_visible_browser_process_running(self) -> None:
        class Process:
            pid = 12345

            @staticmethod
            def poll():
                return None

        runner = BatchRunner(None, self.root, self.root)
        runner.process = Process()

        with patch("batch_workflow.subprocess.run") as taskkill:
            runner.cancel()

        self.assertEqual(taskkill.call_args.args[0], ["taskkill", "/PID", "12345", "/F"])

    def test_exports_seven_sheets_and_requested_sku_fields(self) -> None:
        manifest = self.root / "source.json"
        manifest.write_text(
            json.dumps(
                {
                    "products": [
                        {
                            "product_id": "123",
                            "title": "淘宝标题",
                            "item_url": "https://item.taobao.com/item.htm?id=123",
                            "sales_text": "1万+人付款",
                        }
                    ],
                    "extended_assets": [
                        {
                            "types": {
                                "main": {
                                    "images": [
                                        {"status": "ok", "local_path": str(self.image)},
                                        {"status": "ok", "local_path": str(self.second_image)},
                                    ]
                                },
                                "sku": {
                                    "images": [
                                        {
                                            "status": "ok",
                                            "local_path": str(self.image),
                                            "product_id": "123",
                                            "source_index": "1",
                                            "sku_id": "sku-1",
                                            "sku_label": "黑色【2件装】",
                                            "spec_text": "2件装",
                                            "color_text": "黑色",
                                            "list_price": "29.9",
                                            "after_coupon_price": "19.9",
                                            "parse_status": "parsed",
                                        }
                                    ]
                                },
                                "detail": {
                                    "images": [
                                        {"status": "ok", "local_path": str(self.image)},
                                        {"status": "ok", "local_path": str(self.second_image)},
                                    ]
                                },
                            }
                        }
                    ],
                    "product_parameters": [
                        {
                            "name": "材质",
                            "value": "棉",
                            "handling": "图片识别，待核验",
                        },
                        {
                            "name": "包装结构",
                            "value": "白色塑料泵瓶，正面印有商品名称和容量标识；" * 20,
                            "handling": "图片识别，待核验",
                        },
                    ],
                    "sku_variants": [
                        {
                            "source_index": "1",
                            "sku_id": "sku-1",
                            "sku_label": "黑色 2件装",
                            "spec_text": "2件装",
                            "color_text": "黑色",
                            "list_price": "29.9",
                            "after_coupon_price": "19.9",
                            "parse_status": "parsed",
                        },
                        {
                            "source_index": "2",
                            "sku_id": "sku-2",
                            "sku_label": "蓝色 3件装",
                            "spec_text": "3件装",
                            "color_text": "蓝色",
                            "list_price": "22.9",
                            "parse_status": "parsed",
                        },
                    ],
                    "main_video_url": "https://cloud.video.taobao.com/play/123.mp4",
                    "main_video_status": "complete",
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        item = BatchItem(1, 2, "输入标题", self.image, "https://detail.1688.com/offer/1.html", 10, 2)
        records = [
            {
                "category": "main",
                "ordinal": 1,
                "status": "completed",
                "source_path": str(self.image),
                "output_path": str(self.image),
                "output_public_url": "https://transform-image.oss-cn-shenzhen.aliyuncs.com/product-workflow/generated/main/result.png",
            },
            {
                "category": "main",
                "ordinal": 2,
                "status": "completed",
                "source_path": str(self.image),
                "output_path": str(self.second_image),
            },
            {
                "category": "main",
                "ordinal": 3,
                "status": "completed",
                "source_path": str(self.image),
                "output_path": str(self.image),
            },
            {
                "category": "sku",
                "ordinal": 1,
                "status": "completed",
                "source_path": str(self.image),
                "output_path": str(self.image),
            },
            {
                "category": "detail",
                "ordinal": 1,
                "status": "completed",
                "source_path": str(self.image),
                "output_path": str(self.second_image),
            },
        ]
        output = export_product_workbook(
            self.root / "result.xlsx",
            item,
            manifest,
            records,
            {"long_title": "长标题" * 20, "short_title": "短标题"},
        )

        workbook = load_workbook(output, data_only=True)
        try:
            self.assertEqual(workbook.sheetnames, ["总览", "主图", "详情图", "SKU", "商品参数", "标题", "视频"])
            main = workbook["主图"]
            self.assertEqual([main.cell(1, column).value for column in range(1, 7)], [
                "序号", "采集图缩略图", "采集图路径", "生成图缩略图", "生成图路径", "生成状态"
            ])
            self.assertEqual(main.max_row, 4)
            self.assertEqual(len(main._images), 6)
            self.assertTrue(str(main["C2"].value).endswith("product.png"))
            self.assertEqual(
                main["E2"].value,
                "https://transform-image.oss-cn-shenzhen.aliyuncs.com/product-workflow/generated/main/result.png",
            )
            self.assertTrue(str(main["C3"].value).endswith("product-2.png"))
            self.assertEqual(main["F4"].value, "生成成功")
            detail = workbook["详情图"]
            self.assertEqual(detail.max_row, 3)
            self.assertEqual(len(detail._images), 3)
            sku = workbook["SKU"]
            self.assertEqual([sku.cell(1, column).value for column in range(1, 13)], [
                "序号", "商品ID", "SKU标签", "规格", "颜色", "价格", "解析状态", "采集图缩略图", "采集图路径", "生成图缩略图", "生成图路径", "生成图状态"
            ])
            self.assertEqual(sku["D2"].value, "2件装")
            self.assertEqual(sku["E2"].value, "黑色")
            self.assertEqual(sku["F2"].value, 19.9)
            self.assertEqual(sku.max_row, 2)
            self.assertEqual(len(sku._images), 2)
            self.assertEqual(workbook["标题"]["B2"].value, "长标题" * 20)
            self.assertGreater(workbook["标题"].row_dimensions[2].height, 22)
            self.assertEqual(workbook["标题"]["C2"].value, "短标题")
            self.assertEqual(workbook["视频"]["C2"].value, "https://cloud.video.taobao.com/play/123.mp4")
            self.assertEqual(workbook["商品参数"]["D2"].value, "图片识别，待核验")
            self.assertGreater(workbook["商品参数"].row_dimensions[3].height, 96)
        finally:
            workbook.close()

    def test_export_keeps_all_collected_images_when_generation_is_limited(self) -> None:
        images = [self.image, self.second_image]
        for index in range(3, 6):
            image = self.root / f"product-{index}.png"
            Image.new("RGB", (120, 120), (index * 30, index * 30, index * 30)).save(image)
            images.append(image)
        manifest = self.root / "source-all.json"
        manifest.write_text(
            json.dumps(
                {
                    "products": [{"product_id": "123"}],
                    "extended_assets": [
                        {
                            "types": {
                                asset_type: {"images": [{"status": "ok", "local_path": str(image)} for image in images]}
                                for asset_type in ("main", "sku", "detail")
                            }
                        }
                    ],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        records = [
            {
                "category": asset_type,
                "ordinal": index,
                "status": "completed",
                "source_path": str(images[index - 1]),
                "output_path": str(self.root / f"generated-{asset_type}-{index}.png"),
            }
            for asset_type in ("main", "sku", "detail")
            for index in (1, 2)
        ]

        output = export_product_workbook(
            self.root / "source-all.xlsx",
            BatchItem(1, 2, "输入标题", self.image, "", "", ""),
            manifest,
            records,
        )
        workbook = load_workbook(output, data_only=True)
        try:
            for sheet_name, source_column in (("主图", "C"), ("SKU", "I"), ("详情图", "C")):
                sheet = workbook[sheet_name]
                self.assertEqual(sheet.max_row, 6)
                exported_sources = [sheet[f"{source_column}{row}"].value for row in range(2, 7)]
                self.assertEqual(len([value for value in exported_sources if value]), 5)
        finally:
            workbook.close()

    def test_export_caps_sku_and_detail_rows_at_generation_limits(self) -> None:
        source_images = []
        for index in range(1, 25):
            image = self.root / f"source-{index:03d}.png"
            Image.new("RGB", (120, 120), (index, index, index)).save(image)
            source_images.append({"status": "ok", "local_path": str(image)})
        manifest = self.root / "source-capped.json"
        manifest.write_text(
            json.dumps(
                {
                    "products": [{"product_id": "123"}],
                    "extended_assets": [
                        {
                            "types": {
                                "main": {"images": source_images},
                                "sku": {"images": source_images},
                                "detail": {"images": source_images},
                            }
                        }
                    ],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        with patch("batch_workflow._export_with_artifact_tool", return_value=self.root / "capped.xlsx") as exporter:
            export_product_workbook(
                self.root / "capped.xlsx",
                BatchItem(1, 2, "输入标题", self.image, "", "", ""),
                manifest,
                [],
            )

        payload = exporter.call_args.kwargs["payload"]
        self.assertEqual(len(payload["main"]), 24)
        self.assertEqual(len(payload["sku"]), 8)
        self.assertEqual(len(payload["detail"]), 15)

    def test_export_caps_metadata_only_sku_rows_at_eight(self) -> None:
        manifest = self.root / "metadata-skus.json"
        manifest.write_text(
            json.dumps(
                {
                    "products": [{"product_id": "123"}],
                    "extended_assets": [{"types": {"sku": {"images": []}}}],
                    "sku_variants": [
                        {"sku_id": str(index), "sku_label": f"SKU-{index}", "parse_status": "parsed"}
                        for index in range(1, 13)
                    ],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        with patch("batch_workflow._export_with_artifact_tool", return_value=self.root / "metadata-skus.xlsx") as exporter:
            export_product_workbook(
                self.root / "metadata-skus.xlsx",
                DirectLinkBatchItem(1, 2, "https://item.jd.com/123.html", "jd"),
                manifest,
                [],
                include_metadata_only_skus=True,
            )

        self.assertEqual(len(exporter.call_args.kwargs["payload"]["sku"]), 8)

    def test_sku_export_drops_generated_rows_without_collected_images(self) -> None:
        source = {
            "products": [{"product_id": "123"}],
            "extended_assets": [
                {
                    "types": {
                        "sku": {
                            "images": [
                                {
                                    "status": "ok",
                                    "local_path": str(self.image),
                                    "sku_id": "sku-1",
                                    "source_index": "1",
                                    "parse_status": "parsed",
                                }
                            ]
                        }
                    }
                }
            ],
        }
        rows = _sku_rows(
            source,
            [
                {"ordinal": 1, "status": "completed", "source_path": str(self.image)},
                {"ordinal": 2, "status": "completed", "source_path": ""},
                {"ordinal": 3, "status": "completed", "source_path": "missing.png"},
            ],
            source["products"][0],
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["sku_id"], "sku-1")

    def test_listing_title_validation_enforces_requested_lengths(self) -> None:
        long_title = "商品" * 30
        result = validate_listing_titles({"long_title": long_title, "short_title": "商品短标题太长示例文字"})

        self.assertEqual(result["long_title_length"], 60)
        self.assertEqual(result["short_title_length"], 10)

    def test_title_client_retries_when_model_returns_short_title(self) -> None:
        valid_long_title = "商品" * 28
        responses = [
            {"choices": [{"message": {"content": json.dumps({"long_title": "短", "short_title": "商品"})}}]},
            {
                "choices": [
                    {
                        "message": {
                            "content": json.dumps(
                                {"long_title": valid_long_title, "short_title": "商品短标题"},
                                ensure_ascii=False,
                            )
                        }
                    }
                ]
            },
        ]
        with patch("image_workflows._request_json", side_effect=responses) as request:
            result = ProductTitleClient(ApiSettings("https://api.example", "vision", "image")).generate(
                self.image,
                "我方商品",
                "对标商品标题",
                [],
            )

        self.assertEqual(result["long_title_length"], 56)
        self.assertEqual(result["short_title_length"], 5)
        self.assertEqual(request.call_count, 2)

    def test_collector_command_is_locked_to_top_sales_product(self) -> None:
        item = BatchItem(1, 2, "输入标题", self.image, "", 10, 2)
        item_root = self.root / "item"
        manifest = item_root / "collected" / "main-image-manifest.json"
        manifest.parent.mkdir(parents=True)
        manifest.write_text("{}", encoding="utf-8")

        class FakeProcess:
            stdout = []
            pid = 123

            @staticmethod
            def wait():
                return 0

        runner = BatchRunner(
            ApiSettings("https://api.example", "vision", "image"),
            self.root,
            self.root / "profile",
            browser_executable=r"C:\Program Files\挖象浏览器\挖象.exe",
        )
        collector_script = self.root / "same_item_collector.py"
        collector_script.write_text("", encoding="utf-8")
        with patch("batch_workflow.subprocess.Popen", return_value=FakeProcess()) as popen:
            result = runner._collect(item, item_root)

        command = popen.call_args.args[0]
        self.assertIn("--top-product-only", command)
        self.assertIn("--collect-main-video", command)
        self.assertIn("--browser-executable", command)
        self.assertEqual(
            command[command.index("--browser-executable") + 1],
            r"C:\Program Files\挖象浏览器\挖象.exe",
        )
        self.assertEqual(command[command.index("--max-count") + 1], "1")
        self.assertEqual(result, manifest)

    def test_collector_pause_stops_the_batch_instead_of_continuing(self) -> None:
        item = BatchItem(1, 2, "item", self.image, "", "", "")
        item_root = self.root / "item"
        paused = item_root / "collected" / "collection-paused.json"
        paused.parent.mkdir(parents=True)
        paused.write_text(json.dumps({"status": "paused", "message": "risk 429"}), encoding="utf-8")

        class FakeProcess:
            stdout = []
            pid = 123

            @staticmethod
            def wait():
                return 2

        runner = BatchRunner(None, self.root, self.root)
        collector_script = self.root / "same_item_collector.py"
        collector_script.write_text("", encoding="utf-8")
        with patch("batch_workflow.subprocess.Popen", return_value=FakeProcess()):
            with self.assertRaisesRegex(BatchCollectionPaused, "risk 429"):
                runner._collect(item, item_root)


if __name__ == "__main__":
    unittest.main()
