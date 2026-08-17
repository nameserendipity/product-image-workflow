from __future__ import annotations

import os
import tempfile
from math import isfinite
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorksheetImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PillowImage


SHEET_NAMES = ("总览", "主图", "详情图", "SKU", "商品参数", "标题", "视频")
HEADER_FILL = "17324D"
BORDER_COLOR = "D7DEE5"
LINK_COLOR = "0563C1"

IMAGE_HEADERS = ["序号", "采集图缩略图", "采集图路径", "生成图缩略图", "生成图路径", "生成状态"]
SKU_HEADERS = [
    "序号",
    "商品ID",
    "SKU标签",
    "规格",
    "颜色",
    "价格",
    "采集图缩略图",
    "采集图路径",
    "生成图缩略图",
    "生成图路径",
    "生成图状态",
]


def _local_link(value: Any, output_dir: Path) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if text.lower().startswith(("http://", "https://")):
        return text
    try:
        return os.path.relpath(text, output_dir).replace("\\", "/")
    except ValueError:
        return Path(text).as_posix()


def _image_link(public_url: Any, local_path: Any, output_dir: Path) -> str:
    return str(public_url or "").strip() or _local_link(local_path, output_dir)


def _as_number_or_text(value: Any) -> int | float | str:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = text.replace("￥", "").replace("$", "").replace(",", "")
    try:
        numeric = float(normalized)
    except ValueError:
        return text
    if not isfinite(numeric):
        return text
    return int(numeric) if numeric.is_integer() else numeric


def _text_units(value: Any) -> int:
    return sum(2 if ord(character) > 127 else 1 for character in str(value or ""))


def _apply_dynamic_row_heights(
    sheet: Any,
    rows: list[list[Any]],
    widths: dict[str, float],
    *,
    minimum: float = 22,
    maximum: float = 96,
) -> None:
    for row_index, row in enumerate(rows[1:], start=2):
        lines = 1
        for column_index, value in enumerate(row, start=1):
            column = get_column_letter(column_index)
            capacity = max(float(widths.get(column, 16)), 8)
            lines = max(lines, int((_text_units(value) + capacity - 1) // capacity))
        sheet.row_dimensions[row_index].height = min(maximum, max(minimum, 18 * lines + 4))


def _apply_table_style(sheet: Any, last_row: int, last_column: int, widths: dict[str, float]) -> None:
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in sheet.iter_rows(min_row=1, max_row=max(last_row, 1), min_col=1, max_col=last_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = border
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 26
    sheet.freeze_panes = "A2"
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row_index in range(2, last_row + 1):
        if sheet.row_dimensions[row_index].height is None:
            sheet.row_dimensions[row_index].height = 22


def _style_link(cell: Any) -> None:
    if not cell.value:
        return
    cell.font = Font(color=LINK_COLOR, underline="single")
    cell.hyperlink = str(cell.value)


def _thumbnail_stream(image_path: Path) -> BytesIO:
    stream = BytesIO()
    with PillowImage.open(image_path) as image:
        image.convert("RGBA" if "A" in image.getbands() else "RGB").save(stream, format="PNG")
    stream.seek(0)
    return stream


def _add_thumbnail(sheet: Any, row: int, column: str, value: Any, streams: list[BytesIO]) -> None:
    image_path = Path(str(value or ""))
    if not str(value or "").strip() or not image_path.is_file():
        return
    try:
        if image_path.suffix.lower() == ".webp":
            stream = _thumbnail_stream(image_path)
            streams.append(stream)
            preview = WorksheetImage(stream)
        else:
            preview = WorksheetImage(str(image_path))
        preview.width = 120
        preview.height = 96
        sheet.add_image(preview, f"{column}{row}")
        sheet.row_dimensions[row].height = max(sheet.row_dimensions[row].height or 0, 76)
    except (OSError, ValueError):
        return


def _write_image_sheet(sheet: Any, records: Iterable[dict[str, Any]], output_dir: Path, streams: list[BytesIO]) -> None:
    records = list(records)
    rows: list[list[Any]] = [IMAGE_HEADERS]
    for index, record in enumerate(records, start=1):
        rows.append(
            [
                index,
                "",
                _image_link(record.get("source_public_url"), record.get("source_path"), output_dir),
                "",
                _image_link(record.get("output_public_url"), record.get("output_path"), output_dir),
                record.get("generation_status") or "未生成",
            ]
        )
    for row in rows:
        sheet.append(row)
    widths = {"A": 10, "B": 16, "C": 72, "D": 16, "E": 72, "F": 16}
    _apply_table_style(sheet, len(rows), 6, widths)
    for row_index in range(2, len(rows) + 1):
        _style_link(sheet.cell(row_index, 3))
        _style_link(sheet.cell(row_index, 5))
    for row_index, record in enumerate(records, start=2):
        _add_thumbnail(sheet, row_index, "B", record.get("source_path"), streams)
        _add_thumbnail(sheet, row_index, "D", record.get("output_path"), streams)


def _write_sku_sheet(sheet: Any, records: Iterable[dict[str, Any]], output_dir: Path, streams: list[BytesIO]) -> None:
    records = list(records)
    rows: list[list[Any]] = [SKU_HEADERS]
    for index, record in enumerate(records, start=1):
        rows.append(
            [
                index,
                str(record.get("product_id") or ""),
                record.get("sku_label") or "",
                record.get("spec_text") or "",
                record.get("color_text") or "",
                _as_number_or_text(record.get("price")),
                "",
                _image_link(record.get("source_public_url"), record.get("source_path"), output_dir),
                "",
                _image_link(record.get("output_public_url"), record.get("output_path"), output_dir),
                record.get("generation_status") or "未生成",
            ]
        )
    for row in rows:
        sheet.append(row)
    widths = {
        "A": 8,
        "B": 18,
        "C": 34,
        "D": 18,
        "E": 16,
        "F": 14,
        "G": 16,
        "H": 72,
        "I": 16,
        "J": 72,
        "K": 16,
    }
    _apply_table_style(sheet, len(rows), 11, widths)
    for row_index in range(2, len(rows) + 1):
        sheet.cell(row_index, 2).number_format = "@"
        _style_link(sheet.cell(row_index, 8))
        _style_link(sheet.cell(row_index, 10))
    for row_index, record in enumerate(records, start=2):
        _add_thumbnail(sheet, row_index, "G", record.get("source_path"), streams)
        _add_thumbnail(sheet, row_index, "I", record.get("output_path"), streams)


def _write_parameters(sheet: Any, parameters: Iterable[dict[str, Any]]) -> None:
    rows: list[list[Any]] = [["类型", "参数名", "参数值", "处理方式"]]
    for item in parameters:
        rows.append(
            [
                item.get("type") or "商品参数",
                item.get("name") or "",
                item.get("value") or "",
                item.get("handling") or "采集原值",
            ]
        )
    for row in rows:
        sheet.append(row)
    widths = {"A": 16, "B": 30, "C": 72, "D": 22}
    _apply_table_style(sheet, len(rows), 4, widths)
    _apply_dynamic_row_heights(sheet, rows, widths, maximum=240)


def _write_titles(sheet: Any, title: dict[str, Any]) -> None:
    rows = [["序号", "长标题", "短标题"], [1, title.get("long_title") or "", title.get("short_title") or ""]]
    for row in rows:
        sheet.append(row)
    widths = {"A": 10, "B": 92, "C": 48}
    _apply_table_style(sheet, len(rows), 3, widths)
    _apply_dynamic_row_heights(sheet, rows, widths)


def _write_videos(sheet: Any, videos: Iterable[dict[str, Any]]) -> None:
    rows: list[list[Any]] = [["序号", "视频名称", "公网播放地址", "访问说明"]]
    for index, video in enumerate(videos, start=1):
        rows.append(
            [
                index,
                video.get("name") or "商品主视频",
                str(video.get("url") or ""),
                video.get("note") or "复制完整地址到浏览器即可播放；链接可能含临时授权参数。",
            ]
        )
    if len(rows) == 1:
        rows.append([1, "商品主视频", "", "未找到可打开的视频 URL"])
    for row in rows:
        sheet.append(row)
    widths = {"A": 10, "B": 30, "C": 104, "D": 58}
    _apply_table_style(sheet, len(rows), 4, widths)
    _apply_dynamic_row_heights(sheet, rows, widths, maximum=76)
    for row_index in range(2, len(rows) + 1):
        _style_link(sheet.cell(row_index, 3))


def _write_overview(sheet: Any, overview: Iterable[Iterable[Any]]) -> None:
    rows = [list(row) for row in overview]
    if not rows:
        rows = [["字段", "值"]]
    for row in rows:
        sheet.append(row)
    widths = {"A": 28, "B": 94}
    _apply_table_style(sheet, len(rows), 2, widths)
    _apply_dynamic_row_heights(sheet, rows, widths, maximum=76)
    for row_index in range(2, len(rows) + 1):
        if str(sheet.cell(row_index, 1).value or "").endswith("商品ID"):
            cell = sheet.cell(row_index, 2)
            cell.value = str(cell.value or "")
            cell.number_format = "@"


def export_workbook_payload(output_path: Path, payload: dict[str, Any]) -> Path:
    """Write one workbook payload atomically and return its output path."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {name: workbook.create_sheet(name) for name in SHEET_NAMES}
    streams: list[BytesIO] = []
    temporary_path: Path | None = None
    try:
        _write_overview(sheets["总览"], payload.get("overview") or [])
        _write_image_sheet(sheets["主图"], payload.get("main") or [], output_path.parent, streams)
        _write_image_sheet(sheets["详情图"], payload.get("detail") or [], output_path.parent, streams)
        _write_sku_sheet(sheets["SKU"], payload.get("sku") or [], output_path.parent, streams)
        _write_parameters(sheets["商品参数"], payload.get("parameters") or [])
        _write_titles(sheets["标题"], payload.get("title") or {})
        _write_videos(sheets["视频"], payload.get("videos") or [])

        with tempfile.NamedTemporaryFile(
            dir=output_path.parent,
            prefix=f".{output_path.stem}.",
            suffix=output_path.suffix,
            delete=False,
        ) as temporary:
            temporary_path = Path(temporary.name)
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
        temporary_path = None
        return output_path
    finally:
        workbook.close()
        for stream in streams:
            stream.close()
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)
