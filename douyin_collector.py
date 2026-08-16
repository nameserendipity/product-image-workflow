from __future__ import annotations

import hashlib
import re
import shutil
import zipfile
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any
from urllib.parse import parse_qs, urlencode, urlparse

from openpyxl import load_workbook


IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".avif"}
VIDEO_EXTENSIONS = {".mp4", ".mov", ".webm"}


def canonical_douyin_url(product_id: str) -> str:
    if not product_id or not product_id.isdigit():
        raise ValueError("抖音商品 ID 必须是数字")
    query = urlencode({"id": product_id, "origin_type": "pc_buyin_group"})
    return f"https://haohuo.jinritemai.com/ecommerce/trade/detail/index.html?{query}"


def douyin_checkbox_states(selected_types: set[str]) -> dict[str, bool]:
    return {
        "main": "main" in selected_types,
        "mainVideo": True,
        "detail": "detail" in selected_types,
        "detailLong": False,
        "productInfo": True,
        "productParam": True,
    }


def _is_all_files_download_label(value: str) -> bool:
    normalized = re.sub(r"\s+", "", value).replace("（", "(").replace("）", ")")
    return "下载全部" in normalized and "多文件" in normalized


_legacy_is_all_files_download_label = _is_all_files_download_label


def _is_all_files_download_label(value: str) -> bool:
    normalized = re.sub(r"[\s\u200b\u200c\u200d\ufeff]+", "", value)
    return _legacy_is_all_files_download_label(value) or normalized in {
        "\u591a\u6587\u4ef6",
        "\u591a\u6587\u4ef6\u4e0b\u8f7d",
    } or "\u591a\u6587\u4ef6" in normalized


def _click_all_files_download(menu_items: Any) -> None:
    menu_items.nth(0).wait_for(state="visible", timeout=5000)
    labels: list[str] = []
    for index in range(menu_items.count()):
        item = menu_items.nth(index)
        label = item.inner_text().strip()
        labels.append(label)
        if _is_all_files_download_label(label) and item.is_visible():
            item.evaluate("element => element.click()")
            return
    observed = "、".join(label or "<空白>" for label in labels) or "<无>"
    raise RuntimeError(f"店透视抖音下载菜单未显示“下载全部（多文件）”；当前菜单：{observed}")


def _wait_for_resource_refresh(page: Any, timeout_ms: int) -> None:
    page.evaluate("delete window.__storeInsightResourceIdleSince")
    page.wait_for_timeout(500)
    page.wait_for_function(
        """
        () => {
          const visible = (element) => {
            const style = getComputedStyle(element);
            return style.display !== 'none'
              && style.visibility !== 'hidden'
              && style.opacity !== '0'
              && element.getClientRects().length > 0;
          };
          const busy = Array.from(document.querySelectorAll('.el-loading-mask')).some(visible);
          const button = Array.from(document.querySelectorAll('button')).find(
            (element) => (element.textContent || '').includes('下载图频/数据')
          );
          const disabled = !button || button.disabled || button.classList.contains('is-disabled');
          if (busy || disabled) {
            window.__storeInsightResourceIdleSince = 0;
            return false;
          }
          const now = Date.now();
          if (!window.__storeInsightResourceIdleSince) {
            window.__storeInsightResourceIdleSince = now;
          }
          return now - window.__storeInsightResourceIdleSince >= 1500;
        }
        """,
        timeout=timeout_ms,
        polling=250,
    )


def _missing_requested_image_types(archive_path: Path, selected_types: set[str]) -> set[str]:
    requested_image_types = selected_types.intersection({"main", "detail"})
    if not requested_image_types:
        return set()
    with zipfile.ZipFile(archive_path) as archive:
        available = {
            asset_type
            for member in archive.infolist()
            if not member.is_dir()
            and (asset_type := _archive_asset_type(Path(member.filename))) in requested_image_types
        }
    return requested_image_types - available


def download_douyin_package(
    page: Any,
    product_id: str,
    downloads_dir: Path,
    timeout_ms: int,
    selected_types: set[str],
) -> Path:
    effective_product_id = str(product_id or "").strip()
    if not effective_product_id:
        page_url = getattr(page, "url", "")
        if isinstance(page_url, str):
            effective_product_id = parse_qs(urlparse(page_url).query).get("id", [""])[0].strip()
    if not effective_product_id:
        raise RuntimeError("店透视抖音下载缺少商品 ID，且当前页面地址未提供 id")

    query_input = page.locator('input[placeholder="请输入商品ID/口令"]')
    download_button = page.locator("button").filter(has_text="下载图频/数据")
    query_input.wait_for(state="attached", timeout=timeout_ms)
    download_button.wait_for(state="attached", timeout=timeout_ms)

    if query_input.input_value() != effective_product_id:
        query_input.fill(effective_product_id)
        page.get_by_role("button", name="查询").evaluate("element => element.click()")
        page.wait_for_timeout(1500)
        page.locator(".el-loading-mask").wait_for(state="hidden", timeout=timeout_ms)
    if query_input.input_value() != effective_product_id:
        raise RuntimeError("店透视未加载当前抖音商品 ID")

    for value, checked in douyin_checkbox_states(selected_types).items():
        checkbox = page.locator(f'input[type="checkbox"][value="{value}"]')
        checkbox.wait_for(state="attached", timeout=timeout_ms)
        if checkbox.is_checked() != checked:
            checkbox.evaluate("element => element.click()")
        if checkbox.is_checked() != checked:
            raise RuntimeError(f"店透视抖音下载选项设置失败：{value}")

    _wait_for_resource_refresh(page, timeout_ms)

    downloads_dir.mkdir(parents=True, exist_ok=True)
    with page.expect_download(timeout=timeout_ms) as download_info:
        download_button.wait_for(state="visible", timeout=timeout_ms)
        download_button.evaluate("element => element.click()")
        menu_items = page.locator("li.el-dropdown-menu__item")
        _click_all_files_download(menu_items)
    download = download_info.value
    target = downloads_dir / download.suggested_filename
    download.save_as(target)
    if not target.is_file() or target.stat().st_size == 0:
        raise RuntimeError("店透视抖音资料包下载为空")
    missing_types = _missing_requested_image_types(target, selected_types)
    if missing_types:
        labels = [label for key, label in (("main", "主图"), ("detail", "详情图")) if key in missing_types]
        raise RuntimeError(f"店透视多文件下载包缺少已选择图片：{'、'.join(labels)}")
    return target


def _safe_extract(archive_path: Path, target_dir: Path) -> list[Path]:
    root = target_dir.resolve()
    extracted: list[Path] = []
    with zipfile.ZipFile(archive_path) as archive:
        for member in archive.infolist():
            normalized = member.filename.replace("\\", "/")
            parts = [part for part in normalized.split("/") if part not in {"", "."}]
            if normalized.startswith("/") or any(part == ".." for part in parts):
                raise RuntimeError(f"Unsafe ZIP member path: {member.filename}")
            destination = target_dir.joinpath(*parts).resolve()
            if destination != root and root not in destination.parents:
                raise RuntimeError(f"Unsafe ZIP member path: {member.filename}")
            if member.is_dir():
                destination.mkdir(parents=True, exist_ok=True)
                continue
            destination.parent.mkdir(parents=True, exist_ok=True)
            with archive.open(member) as source, destination.open("wb") as output:
                shutil.copyfileobj(source, output)
            extracted.append(destination)
    return extracted


def _archive_asset_type(path: Path) -> str | None:
    if path.suffix.lower() not in IMAGE_EXTENSIONS:
        return None
    normalized = path.as_posix().lower()
    if "sku" in normalized or "sku图" in normalized:
        return "sku"
    if "详情图" in normalized or "detail" in normalized:
        return "detail"
    if "主图" in normalized or "页面图" in normalized or "main" in normalized:
        return "main"
    return None


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _normalized_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).lower()


def _parse_product_workbook(path: Path) -> dict[str, Any]:
    metadata: dict[str, Any] = {}
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if "商品数据" in workbook.sheetnames:
            sheet = workbook["商品数据"]
            rows = list(sheet.iter_rows(min_row=1, max_row=2, values_only=True))
            if len(rows) == 2:
                values = {
                    _normalized_header(name): value
                    for name, value in zip(rows[0], rows[1])
                    if _normalized_header(name)
                }
                metadata.update(
                    {
                        "product_title": str(values.get("商品标题") or "").strip(),
                        "product_id": str(values.get("商品id") or "").strip(),
                        "brand": str(values.get("商品品牌") or "").strip(),
                        "category": str(values.get("商品类目") or "").strip(),
                        "current_price": str(values.get("商品价格") or "").strip(),
                        "sales_30_days": str(values.get("30天销量") or "").strip(),
                        "shop_name": str(values.get("店铺名称") or "").strip(),
                    }
                )
        parameters: list[dict[str, str]] = []
        if "商品参数" in workbook.sheetnames:
            sheet = workbook["商品参数"]
            rows = list(sheet.iter_rows(min_row=1, max_row=2, values_only=True))
            if len(rows) == 2:
                parameters = [
                    {"name": str(name).strip(), "value": str(value).strip()}
                    for name, value in zip(rows[0], rows[1])
                    if str(name or "").strip() and str(value or "").strip()
                ]
        metadata["product_parameters"] = parameters
        metadata["parameter_status"] = "complete" if parameters else "not_found"
        metadata["parameter_error"] = "" if parameters else "抖音资料表未提供商品参数"
        return metadata
    finally:
        workbook.close()


def materialize_douyin_package(
    archive_path: Path,
    output_root: Path,
    selected_types: set[str],
    max_main_images: int | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    records: list[dict[str, Any]] = []
    counters: dict[str, int] = {}
    metadata: dict[str, Any] = {
        "product_title": "",
        "current_price": "",
        "product_parameters": [],
        "parameter_status": "not_found",
        "parameter_error": "抖音资料表未提供商品参数",
        "main_video_requested": True,
        "main_video_url": "",
        "main_video_local_path": "",
        "main_video_status": "not_found",
        "main_video_error": "抖音资料包未包含主图视频",
        "sku_metadata_status": "not_found",
        "sku_metadata_error": "抖音资料表未提供可验证的 SKU 数据",
        "sku_variants": [],
    }
    with TemporaryDirectory(prefix="douyin-product-extract-") as temporary:
        files = _safe_extract(archive_path, Path(temporary))
        for source in files:
            asset_type = _archive_asset_type(source)
            if asset_type is not None:
                if asset_type not in selected_types:
                    continue
                if asset_type == "main" and max_main_images is not None and counters.get("main", 0) >= max_main_images:
                    continue
                counters[asset_type] = counters.get(asset_type, 0) + 1
                destination_root = output_root / asset_type
                destination_root.mkdir(parents=True, exist_ok=True)
                destination = destination_root / f"{counters[asset_type]:03d}{source.suffix.lower()}"
                shutil.copy2(source, destination)
                records.append(
                    {
                        "type": asset_type,
                        "path": str(destination.resolve()),
                        "source_name": source.name,
                        "sha256": _sha256(destination),
                    }
                )
                continue
            if source.suffix.lower() in VIDEO_EXTENSIONS and not metadata["main_video_local_path"]:
                video_root = output_root / "video"
                video_root.mkdir(parents=True, exist_ok=True)
                destination = video_root / f"main-video{source.suffix.lower()}"
                shutil.copy2(source, destination)
                metadata.update(
                    {
                        "main_video_local_path": str(destination.resolve()),
                        "main_video_status": "local_only",
                        "main_video_error": "",
                    }
                )
                continue
            if source.suffix.lower() == ".xlsx":
                parsed = _parse_product_workbook(source)
                metadata.update({key: value for key, value in parsed.items() if value not in (None, "")})

    metadata["requested_asset_types"] = sorted(selected_types)
    metadata["collected_asset_types"] = [
        asset_type for asset_type in ("main", "sku", "detail")
        if any(record["type"] == asset_type for record in records)
    ]
    metadata["missing_asset_types"] = [
        asset_type for asset_type in sorted(selected_types)
        if asset_type not in metadata["collected_asset_types"]
    ]
    return records, metadata
